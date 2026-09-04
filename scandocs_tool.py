#!/usr/bin/env python3
"""
Speedy Scandocs
Automatically classifies and renames scanned law firm documents
using a local AI model (OpenWebUI / Ollama).

Usage:
    python scandocs_tool.py

Requirements:
    pip install -r requirements.txt
"""

import os
import re
import sys
import json
import base64
import difflib
import hashlib
import logging
import threading
import queue
import csv
import datetime
import subprocess
import socket
import getpass
import uuid
import time

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _XLSX_AVAILABLE = True
except ImportError:
    _XLSX_AVAILABLE = False
from dataclasses import dataclass, field
from typing import Optional, List

import tkinter as tk
from tkinter import messagebox, filedialog
import ttkbootstrap as ttk

try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None

try:
    import requests
except ImportError:
    requests = None

try:
    import pytesseract
except ImportError:
    pytesseract = None

try:
    from PIL import Image as PILImage, ImageTk as PILImageTk
except ImportError:
    PILImage = None
    PILImageTk = None

# Optional — only used by DocumentExtractor._prepare_photo() when
# reading.deskew_photos is on. NOT in requirements.txt: the production app
# must work identically without them, so every use is behind this flag and
# degrades gracefully (see _prepare_photo) when either is missing.
try:
    import cv2
    import numpy as np
    _CV2_AVAILABLE = True
except ImportError:
    cv2 = None
    np = None
    _CV2_AVAILABLE = False


# ─────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(SCRIPT_DIR, "assets")

# ── Version + auto-update ──────────────────────────────────────────────────
# APP_VERSION is bumped by build/release.py — keep it in sync with the
# installer.iss AppVersion. Auto-update checks GitHub Releases on UPDATE_REPO
# and compares the latest tag (vX.Y.Z) against APP_VERSION.
APP_VERSION = "2.0.0"
UPDATE_REPO = "treyj1/Speedy-Scandocs"
UPDATE_API_URL = f"https://api.github.com/repos/{UPDATE_REPO}/releases/latest"
UPDATE_CHECK_INTERVAL_SEC = 24 * 60 * 60   # 24 hours

# ── Build flavor ────────────────────────────────────────────────────────────
# "production" or "test". A test build uses a separate app-data directory
# (so it never touches a real installation's config/logs/reports/client
# list) and disables auto-update entirely. Do not change this by editing the
# installer/spec files in this pass — it's a source-level switch for now.
BUILD_FLAVOR = "production"   # "production" or "test"
IS_TEST_BUILD = (BUILD_FLAVOR == "test")
APP_TITLE = "Speedy Scandocs" + (" [TEST BUILD]" if IS_TEST_BUILD else "")
_APP_DATA_DIRNAME = "SpeedyScandocs-Test" if IS_TEST_BUILD else "SpeedyScandocs"

# ── Colour-palette definitions ─────────────────────────────────────────────
# App primary color — fixed, no user-selectable palette
_APP_PRIMARY   = "#1565c0"
_APP_LIGHT     = "#e3f2fd"
_APP_MID       = "#1976d2"

# ── Typography ─────────────────────────────────────────────────────────────
# Single app-wide font family.
APP_FONT = "Times New Roman"

# ── User-writable data directory ───────────────────────────────────────────
# When installed to Program Files / Applications the app bundle is read-only,
# so config, logs, and reports are stored in the platform user-data folder.
if getattr(sys, "frozen", False):
    if sys.platform == "win32":
        _appdata = os.environ.get("APPDATA", os.path.expanduser("~"))
    elif sys.platform == "darwin":
        _appdata = os.path.join(os.path.expanduser("~"), "Library", "Application Support")
    else:
        _appdata = os.environ.get("XDG_DATA_HOME", os.path.join(os.path.expanduser("~"), ".local", "share"))
    _USER_DATA_DIR = os.path.join(_appdata, _APP_DATA_DIRNAME)
    os.makedirs(_USER_DATA_DIR, exist_ok=True)
else:
    _USER_DATA_DIR = SCRIPT_DIR

CONFIG_PATH            = os.path.join(_USER_DATA_DIR, "config.json")
LOG_PATH               = os.path.join(_USER_DATA_DIR, "scandocs_tool.log")
DEFAULT_REPORTS_FOLDER = os.path.join(_USER_DATA_DIR, "Reports")
RENAME_LOG_PATH         = os.path.join(_USER_DATA_DIR, "renames.jsonl")

# Instance lock (folder lock) tuning — see FolderLock below.
LOCK_FILENAME               = ".speedyscandocs.lock"
LOCK_STALE_SECONDS          = 120   # heartbeat older than this = treat as dead, take over
LOCK_HEARTBEAT_INTERVAL_SEC = 30    # how often run_batch refreshes the heartbeat


def _open_file(path: str):
    """Open a file with the default system viewer — cross-platform."""
    if sys.platform == "win32":
        os.startfile(path)
    elif sys.platform == "darwin":
        subprocess.run(["open", path])
    else:
        subprocess.run(["xdg-open", path])


# ── PyInstaller bundle: point pytesseract at the bundled Tesseract binary ──
_bundle_dir = getattr(sys, "_MEIPASS", None)
if _bundle_dir:
    try:
        import pytesseract as _pt
        # Windows bundle: Tesseract-OCR/tesseract.exe
        _bundled_tess_win = os.path.join(_bundle_dir, "Tesseract-OCR", "tesseract.exe")
        # Mac bundle: tesseract (no extension)
        _bundled_tess_mac = os.path.join(_bundle_dir, "tesseract")
        if os.path.isfile(_bundled_tess_win):
            _pt.pytesseract.tesseract_cmd = _bundled_tess_win
            os.environ["TESSDATA_PREFIX"] = os.path.join(_bundle_dir, "Tesseract-OCR", "tessdata")
        elif os.path.isfile(_bundled_tess_mac):
            _pt.pytesseract.tesseract_cmd = _bundled_tess_mac
            os.environ["TESSDATA_PREFIX"] = os.path.join(_bundle_dir, "tessdata")
    except ImportError:
        pass
else:
    # Not bundled — try known system install locations (Windows + Mac Homebrew)
    _TESS_CANDIDATES = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        "/opt/homebrew/bin/tesseract",   # Mac Apple Silicon (Homebrew)
        "/usr/local/bin/tesseract",      # Mac Intel (Homebrew)
    ]
    try:
        import pytesseract as _pt
        for _p in _TESS_CANDIDATES:
            if os.path.isfile(_p):
                _pt.pytesseract.tesseract_cmd = _p
                break
    except ImportError:
        pass


# ── Windows taskbar identity ───────────────────────────────────────────────
# Without an explicit AppUserModelID, Windows groups the running app under
# Python's default ID and the taskbar icon falls back to Python's, not the
# embedded EXE icon. Must be set before any Tk window is mapped.
if sys.platform == "win32":
    try:
        import ctypes as _ctypes
        _ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
            "com.gdj.speedyscandocs"
        )
    except Exception:
        pass


# ── Bundled font registration ──────────────────────────────────────────────
# Must run before any Tk window is created — GDI/CoreText caches font lists
# at Tk init time, so a late registration won't be visible to tkinter.
def _load_bundled_fonts() -> None:
    fonts_dir = os.path.join(ASSETS_DIR, "fonts")
    if not os.path.isdir(fonts_dir):
        return
    ttfs = [os.path.join(fonts_dir, f) for f in os.listdir(fonts_dir)
            if f.lower().endswith(".ttf")]
    if not ttfs:
        return
    if sys.platform == "win32":
        try:
            import ctypes
            FR_PRIVATE = 0x10
            gdi32 = ctypes.WinDLL("gdi32")
            gdi32.AddFontResourceExW.argtypes = [
                ctypes.c_wchar_p, ctypes.c_ulong, ctypes.c_void_p]
            gdi32.AddFontResourceExW.restype = ctypes.c_int
            for path in ttfs:
                if gdi32.AddFontResourceExW(path, FR_PRIVATE, 0) == 0:
                    logging.info(f"AddFontResourceExW failed: {path}")
        except Exception as e:
            logging.info(f"Windows font load failed: {e}")
    elif sys.platform == "darwin":
        try:
            import ctypes
            from ctypes import c_void_p, c_bool, c_long, c_uint32, c_char_p
            cf = ctypes.CDLL("/System/Library/Frameworks/CoreFoundation.framework/CoreFoundation")
            ct = ctypes.CDLL("/System/Library/Frameworks/CoreText.framework/CoreText")
            cf.CFStringCreateWithCString.restype = c_void_p
            cf.CFStringCreateWithCString.argtypes = [c_void_p, c_char_p, c_uint32]
            cf.CFURLCreateWithFileSystemPath.restype = c_void_p
            cf.CFURLCreateWithFileSystemPath.argtypes = [c_void_p, c_void_p, c_long, c_bool]
            cf.CFRelease.argtypes = [c_void_p]
            ct.CTFontManagerRegisterFontsForURL.restype = c_bool
            ct.CTFontManagerRegisterFontsForURL.argtypes = [c_void_p, c_uint32, c_void_p]
            kCFStringEncodingUTF8 = 0x08000100
            kCTFontManagerScopeProcess = 1
            for path in ttfs:
                cfstr = cf.CFStringCreateWithCString(
                    None, path.encode("utf-8"), kCFStringEncodingUTF8)
                if not cfstr:
                    continue
                cfurl = cf.CFURLCreateWithFileSystemPath(None, cfstr, 0, False)
                if cfurl:
                    ct.CTFontManagerRegisterFontsForURL(
                        cfurl, kCTFontManagerScopeProcess, None)
                    cf.CFRelease(cfurl)
                cf.CFRelease(cfstr)
        except Exception as e:
            logging.info(f"macOS font load failed: {e}")

DEFAULT_CONFIG: dict = {
    "paths": {
        "scandocs_folder": "",
        "client_list_file": os.path.join(_USER_DATA_DIR, "client_list.txt"),
        "document_types_file": os.path.join(_USER_DATA_DIR, "document_types.txt"),
        "providers_file": os.path.join(_USER_DATA_DIR, "providers.txt"),
    },
    "api": {
        "openwebui_url": "http://localhost:3000",
        "ollama_url": "http://localhost:11434",
        "model": "llama3.2-vision",
        "api_key": "",
        "timeout_connect": 10,
        "timeout_read": 120,
    },
    "processing": {
        "fuzzy_threshold": 0.82,
        "max_ocr_chars": 8000,
        "require_high_confidence": True,
        "max_pages": 5,
        "skip_already_processed": True,
        "audit_mode": True,
        "file_mode": False,
        "file_mode_auto": True,     # enable File Mode's panel on the Process tab
        "file_mode_manual": True,   # enable File Mode's panel on the Manual Correction panel
        "file_mode_destination": "",
        "suggest_location_enabled": False,
        "suggest_location_parent_folder": "",
        "auto_commit_moves": False,
        "candidate_list_size": 10,
        "show_manual_entry_tab": False,  # legacy standalone Manual Entry tab
        "extraction_method": "ocr",   # "ocr" or "vision"
        "max_vision_pages": 2,          # pages sent to vision model per doc
        "ocr_preprocess": True,         # upscale/binarize/autocontrast before OCR
    },
    "reports": {
        "auto_save": True,
        "report_folder": DEFAULT_REPORTS_FOLDER,
    },
    "updates": {
        "check_on_startup": True,
        "last_check_iso": "",
        "skip_version": "",
    },
    "ui": {
        "preview_popup_width": 0,   # 0 = not yet set by the user; use the default large size
        "preview_popup_height": 0,
    },
    # ── Foundation for later passes — all defaults below reproduce today's
    # behavior exactly. No UI wired to these yet.
    "naming": {
        "preserve_acronyms": True,          # safe, pure improvement — on by default
        "include_recipient": False,
        "include_doc_date": False,
        "use_templates": False,             # False = legacy "{client} - {desc}" construction, byte-for-byte
        "date_format": "%m-%d-%y",          # e.g. "PPR 07-15-26", matching the office's handwriting convention
        "split_unknown_states": False,      # False = both cases collapse to no_client_label, as today
        # doc_type -> template string. [bracketed] segments drop whole when
        # any placeholder inside them is empty (see NameTemplate). Only
        # takes effect when use_templates is True.
        "templates": {
            "Reduction Request":            "{client} - {doc_type}[ to {recipient}]",
            "Balance Verification Request": "{client} - {doc_type}[ to {recipient}]",
            "Medical Records Request":      "{client} - {doc_type}[ to {recipient}]",
            "PPR":                          "{client} - {doc_type}[ {doc_date}]",
            "Progress Report":              "{client} - {doc_type}[ {doc_date}]",
            "Disability Certification":     "{client} - {doc_type}[ {doc_date}]",
        },
        "default_template": "{client} - {doc_type}",
        "date_disambiguation": False,       # False = legacy "(1)" behavior
        "unknown_client_label": "A-UNKNOWN CLIENT",  # a name was read but isn't on the client list
        "no_client_label": "A-NEEDS REVIEW",         # no name could be read at all — keep legacy default
    },
    "reading": {
        "skip_fax_cover_pages": False,
        "deskew_photos": False,
        "vision_escalation": False,
        "extract_claim_numbers": False,
        "escalation_threshold": 0.35,   # text-quality score below which vision_escalation kicks in
    },
    "learning": {
        "log_corrections": True,            # foundation; harmless, just writes a log
        "document_types": "off",            # "off" | "suggest" | "auto"
        "client_relationships": "off",      # "off" | "suggest" | "auto"
        "claim_linking": "off",             # "off" | "suggest" | "auto"
        "observations_required": 3,
        "retroactive_rename": "off",        # "off" | "preview"
        "few_shot_examples": False,         # append similar past corrections to the prompt
    },
    "automation": {
        "dry_run": False,
        "watch_folder": False,
        "watch_poll_seconds": 15,
    },
    "safety": {
        # Deliberately NOT user-toggleable. Present for diagnostics/logging only.
        "instance_lock": True,
        "undo_log": True,
        "recheck_before_rename": True,
    },
    # ── PASS 4: classification quality — vocabularies, structured output,
    # grounding, evidence-based confidence. Every flag below defaults to
    # False/off, so an unmodified config reproduces today's exact prompt,
    # parsing, and confidence behavior. See ClientListManager,
    # DocumentTypeManager, ProviderManager, APIClient, FileProcessor.
    "classification": {
        "structured_output": False,      # off = today's exact prompt and parsing
        "use_document_types": False,
        "use_providers": False,
        "extract_recipient": False,
        "grounding_check": False,
        "evidence_confidence": False,
        "use_candidate_shortlist": False,
    },
}

SUPPORTED_EXTENSIONS = {".pdf", ".jpg", ".jpeg"}
ILLEGAL_CHARS_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

# ── Vision model allowlist ───────────────────────────────────────────
# Substring-matched against the lowercased Ollama model name. Any model
# whose tag contains one of these prefixes is treated as vision-capable,
# which unlocks "Use Vision Model" in Settings. Excluded tags (cloud
# variants) send document images off-device, which isn't acceptable for
# law office client documents, so we force those to OCR mode.
VISION_MODEL_PREFIXES = [
    "llama3.2-vision",  # 11b and 90b
    "gemma4",           # latest, e2b, e4b, 26b, 31b (local variants)
]
VISION_MODEL_EXCLUSIONS = ["-cloud", ":cloud"]

# ── Document type vocabulary (PASS 4) ───────────────────────────────────
# Seeds document_types.txt on first run when classification.use_document_types
# is turned on and the file doesn't exist yet. Each entry is written verbatim
# as a line in "Canonical Name | alias1 | alias2" format — see
# DocumentTypeManager.load / load_alias_map.
DEFAULT_DOCUMENT_TYPES = [
    "Reduction Request | Compromise Offer | Request for Reduction | Reduction Letter",
    "Balance Verification Request",
    "PPR | Physician Progress Report | Physicians and Chiropractors Progress Report | "
    "Physician Chiropractor Progress Report",
    "Progress Report",
    "Medical Records Request",
    "Medical Bills",
    "Letter of Representation",
    "Demand Letter",
    "Lien",
    "Certified Mail Receipt",
    "Insurance Claim Payment Check",
    "Notice of Intention to Close Claim",
    "Permanent Partial Disability Award Notice",
    "TTD Benefits Appeal Notice",
    "Workers Compensation Hearing Appeal",
    "Retainer Agreement",
    "Settlement Statement",
    "Subpoena",
    "Police Report",
    "Medical Authorization",
    "Disability Certification",
    "IME Report",
]

# providers.txt ships empty (it fills in from use) — this is only the
# explanatory comment header written on first run.
_PROVIDERS_FILE_HEADER = [
    "# Providers (facilities, clinics, hospitals, insurers, etc.), one per line.",
    "# This file starts empty and is meant to fill in as recipients/senders are",
    "# recognized through use. Add a name manually to have it recognized",
    "# immediately by ProviderManager.normalize.",
]


def _disable_combobox_scroll(widget) -> None:
    """Stop a ttk.Combobox from cycling values when the mouse wheel scrolls
    over it. Otherwise an accidental scroll while hovering Settings can flip
    Use OCR ↔ Use Vision Model without the user realizing."""
    for seq in ("<MouseWheel>", "<Button-4>", "<Button-5>"):
        widget.bind(seq, lambda _e: "break")


def model_supports_vision(model_name: str) -> bool:
    """Return True if the given Ollama model tag is a local vision model
    in our allowlist. Cloud variants are intentionally excluded."""
    if not model_name:
        return False
    name = model_name.lower()
    if any(excl in name for excl in VISION_MODEL_EXCLUSIONS):
        return False
    return any(prefix in name for prefix in VISION_MODEL_PREFIXES)


# ─────────────────────────────────────────────────────────────
# Settings tab — display <-> config mapping and dependency logic (PASS 7)
# ─────────────────────────────────────────────────────────────
# Pure functions, deliberately free of any tkinter/UI dependency so they
# can be unit-tested on their own. The Settings tab calls these; nothing
# else in the app should need to.

LEARNING_MODE_CONFIG_TO_DISPLAY = {
    "off": "Off", "suggest": "Suggest only", "auto": "Automatic",
}
LEARNING_MODE_DISPLAY_TO_CONFIG = {
    v: k for k, v in LEARNING_MODE_CONFIG_TO_DISPLAY.items()
}
LEARNING_MODE_VALUES = list(LEARNING_MODE_CONFIG_TO_DISPLAY.values())


def learning_mode_to_display(value: str) -> str:
    """Config string ("off"/"suggest"/"auto") -> combobox display string.
    An unrecognized or legacy value falls back to "Off" rather than
    raising, so a hand-edited config.json can never crash the UI."""
    return LEARNING_MODE_CONFIG_TO_DISPLAY.get((value or "").strip().lower(), "Off")


def learning_mode_to_config(display: str) -> str:
    """Combobox display string -> config string. An unrecognized display
    value (shouldn't happen from the UI itself) falls back to "off"."""
    return LEARNING_MODE_DISPLAY_TO_CONFIG.get((display or "").strip(), "off")


RETROACTIVE_MODE_CONFIG_TO_DISPLAY = {"off": "Off", "preview": "Preview only"}
RETROACTIVE_MODE_DISPLAY_TO_CONFIG = {
    v: k for k, v in RETROACTIVE_MODE_CONFIG_TO_DISPLAY.items()
}
RETROACTIVE_MODE_VALUES = list(RETROACTIVE_MODE_CONFIG_TO_DISPLAY.values())


def retroactive_mode_to_display(value: str) -> str:
    """Only two states exist here on purpose — there is no "Automatic":
    a law firm must never have files renamed with no human in the loop."""
    return RETROACTIVE_MODE_CONFIG_TO_DISPLAY.get((value or "").strip().lower(), "Off")


def retroactive_mode_to_config(display: str) -> str:
    return RETROACTIVE_MODE_DISPLAY_TO_CONFIG.get((display or "").strip(), "off")


def compute_settings_dependency_state(cfg: dict) -> dict:
    """Given a config dict, return which dependent Settings controls should
    be disabled right now, and why.

    Returns {dependency_key: (disabled: bool, reason: str)} — `reason` is
    "" whenever disabled is False. Keys match the setting each gates:
      "classification.use_document_types"  requires structured_output
      "classification.use_providers"       requires structured_output
      "classification.extract_recipient"   requires structured_output
      "naming.include_recipient"           requires classification.extract_recipient
      "reading.vision_escalation"          requires a vision-capable model
      "reading.deskew_photos"              requires OpenCV/numpy in the build
    """
    classification = (cfg or {}).get("classification", {}) or {}
    api = (cfg or {}).get("api", {}) or {}
    structured_on = bool(classification.get("structured_output", False))
    extract_recipient_on = bool(classification.get("extract_recipient", False))
    model = (api.get("model") or "").strip()
    vision_ok = model_supports_vision(model)

    need_structured = "Turn on 'Ask the model for structured details' first"
    need_recipient = "Turn on 'Identify who the document is going to' first"
    if vision_ok:
        vision_reason = ""
    elif model:
        vision_reason = f"Needs a vision-capable model — currently set to {model}"
    else:
        vision_reason = "Needs a vision-capable model — no model is set"

    return {
        "classification.use_document_types": (not structured_on, "" if structured_on else need_structured),
        "classification.use_providers":      (not structured_on, "" if structured_on else need_structured),
        "classification.extract_recipient":  (not structured_on, "" if structured_on else need_structured),
        "naming.include_recipient":          (not extract_recipient_on, "" if extract_recipient_on else need_recipient),
        "reading.vision_escalation":         (not vision_ok, vision_reason),
        # OpenCV/numpy are deliberately absent from requirements.txt, so in
        # a normal build _prepare_photo can only log a warning and fall
        # back. Say that in the UI rather than offering a switch that
        # silently does nothing.
        "reading.deskew_photos": (
            not _CV2_AVAILABLE,
            "" if _CV2_AVAILABLE else
            "Not available in this build — needs OpenCV, which isn't included",
        ),
    }


# ─────────────────────────────────────────────────────────────
# Data classes
# ─────────────────────────────────────────────────────────────

@dataclass
class ExtractionResult:
    content_type: str   # "text" or "image"
    content: str        # text string or first base64 image (back-compat)
    mime_type: str = "image/png"
    method: str = ""    # "pymupdf", "tesseract", or "vision"
    # For multi-page vision extraction. When non-empty, APIClient sends
    # every base64 image in this list to the model. `content` mirrors
    # images[0] for back-compat with single-image code paths.
    images: List[str] = field(default_factory=list)


@dataclass
class ProcessResult:
    original_name: str
    final_name: str
    status: str         # "renamed", "needs_review", "skipped", "error"
    client: str = ""
    description: str = ""
    confidence: str = ""
    error_message: Optional[str] = None
    renamed_at: Optional[str] = None
    extraction_method: str = ""
    # Audit fields — filled in by employee after processing
    audit_correct: bool = False
    audit_wrong_client: bool = False
    audit_bad_description: bool = False
    audit_failed_client: bool = False
    audit_should_review: bool = False
    audit_corrected_name: str = ""  # what the employee said it should be named
    # File mode
    moved_to: str = ""              # destination path after a successful move
    pending_dest: str = ""          # destination staged via "Apply to Selected", not yet moved
    # Foundation fields for later passes (learning, structured naming, etc.)
    raw_client: str = ""            # what the model returned, before fuzzy matching
    raw_confidence: str = ""
    extracted_text: str = ""        # first 2000 chars of what was actually read
    doc_hash: str = ""              # sha256 of file bytes, for dedupe
    doc_type: str = ""              # structured field, populated in a later pass
    recipient: str = ""
    doc_date: str = ""
    direction: str = ""             # "incoming" | "outgoing" | ""
    claim_number: str = ""
    was_dry_run: bool = False
    skip_reason: str = ""           # why a file was skipped, for reporting
    # PASS 6 (learning): how final_client was determined — "" (not resolved
    # / needs review), "fuzzy" (normal client-list match), "claim" (learned
    # claim-number linking), "alias" (learned client-relationship linking).
    match_source: str = ""


# ─────────────────────────────────────────────────────────────
# ConfigManager
# ─────────────────────────────────────────────────────────────

class ConfigManager:
    def __init__(self):
        self.config = self._load()

    def _load(self) -> dict:
        if not os.path.exists(CONFIG_PATH):
            self._write(DEFAULT_CONFIG)
            return self._deep_copy(DEFAULT_CONFIG)
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            return self._deep_merge(self._deep_copy(DEFAULT_CONFIG), data)
        except Exception as e:
            logging.warning(f"Could not load config.json: {e}. Using defaults.")
            return self._deep_copy(DEFAULT_CONFIG)

    # Config dicts that are USER DATA, not a set of named settings: a
    # saved value replaces the default outright instead of being merged
    # key-by-key over it. Without this, recursive merging re-adds any entry
    # the user deleted — every app start would resurrect the stock
    # naming.templates rows they had removed, with no way to get rid of them.
    REPLACE_WHOLESALE_PATHS = frozenset({"naming.templates"})

    @staticmethod
    def _deep_merge(defaults: dict, saved: dict, _path: str = "") -> dict:
        """Recursively merge `saved` (loaded from disk) over `defaults`.

        - A saved scalar overrides the default.
        - A key present in defaults but missing from saved keeps the default.
        - A key present in saved but unknown to defaults is preserved as-is
          (never silently dropped — e.g. a section added by a newer version
          of the app, or user data from a future config format).
        - Nested dicts are merged recursively rather than replaced wholesale.

        This replaces a previous implementation that merged section-by-section
        with hardcoded lines (`merged["paths"].update(...)`, etc.) — any
        section added to DEFAULT_CONFIG without a matching hardcoded line
        there would have its saved values silently discarded on every load.

        `_path` tracks position so REPLACE_WHOLESALE_PATHS (below) can opt
        specific dicts out of recursive merging.
        """
        if not isinstance(saved, dict):
            return defaults
        merged = dict(defaults)
        for key, saved_val in saved.items():
            child_path = f"{_path}.{key}" if _path else key
            if child_path in ConfigManager.REPLACE_WHOLESALE_PATHS \
                    and isinstance(saved_val, dict):
                merged[key] = saved_val
            elif key in merged and isinstance(merged[key], dict) and isinstance(saved_val, dict):
                merged[key] = ConfigManager._deep_merge(merged[key], saved_val, child_path)
            else:
                merged[key] = saved_val
        return merged

    def save(self, data: dict = None):
        if data is None:
            data = self.config
        self._write(data)
        self.config = data

    def _write(self, data: dict):
        tmp = CONFIG_PATH + ".tmp"
        try:
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            os.replace(tmp, CONFIG_PATH)
        except Exception as e:
            logging.error(f"Failed to write config: {e}")
            raise

    def validate(self) -> list:
        errors = []
        scandocs = self.config["paths"].get("scandocs_folder", "")
        if not scandocs:
            errors.append("Scandocs folder is not configured.")
        elif not os.path.isdir(scandocs):
            errors.append(f"Scandocs folder not found:\n  {scandocs}")
        client_file = self.config["paths"].get("client_list_file", "")
        if not client_file:
            errors.append("Client list file path is not configured.")
        return errors

    @staticmethod
    def _deep_copy(d: dict) -> dict:
        return json.loads(json.dumps(d))


# ─────────────────────────────────────────────────────────────
# FolderLock — cross-instance / cross-machine batch lock
# ─────────────────────────────────────────────────────────────
#
# The scandocs folder is Dropbox-synced and more than one machine (or more
# than one copy of the app) can point at it. Without coordination, two
# concurrent batches will both grab the same file list, and the second one
# to reach a given file finds it already renamed out from under it — the
# paired OK/ERROR row bug this pass fixes. FolderLock is a best-effort
# cooperative lock file living inside the scandocs folder itself, so every
# instance pointed at that folder sees it regardless of machine.

class FolderLockHeld(Exception):
    """Raised by FolderLock.acquire() when another instance's heartbeat
    proves it is alive right now. This is the one case that must actually
    block starting a batch — every other lock problem is swallowed."""

    def __init__(self, message: str, info: Optional[dict] = None):
        super().__init__(message)
        self.info = info or {}


class FolderLock:
    """Cooperative, best-effort lock file (`.speedyscandocs.lock`) inside a
    scandocs folder. Use as a context manager:

        with FolderLock(scandocs_folder) as lock:
            ... run the batch, calling lock.heartbeat() periodically ...

    Acquisition is atomic (O_CREAT | O_EXCL) so two processes racing to
    create the file cannot both "win". A lock whose heartbeat is fresh
    (< LOCK_STALE_SECONDS old) blocks acquisition by raising
    FolderLockHeld; a stale one (crash, or a Stop that somehow skipped
    release) is logged and taken over. Any other OS-level problem
    (permissions, a read-only/offline share, etc.) is logged as a warning
    and acquisition is treated as a no-op success — a lock we can't take
    is not worth blocking real work over.
    """

    def __init__(self, folder: str):
        self.folder = folder
        self.path = os.path.join(folder, LOCK_FILENAME)
        self.acquired = False
        self._last_heartbeat_monotonic: float = 0.0

    # ── internals ────────────────────────────────────────────

    def _read(self) -> Optional[dict]:
        try:
            with open(self.path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None

    @staticmethod
    def _is_stale(info: dict) -> bool:
        hb = info.get("heartbeat") or info.get("started")
        if not hb:
            return True
        try:
            hb_time = datetime.datetime.fromisoformat(hb)
        except Exception:
            return True
        return (datetime.datetime.now() - hb_time).total_seconds() > LOCK_STALE_SECONDS

    @staticmethod
    def _describe(info: dict) -> str:
        host = info.get("host") or "another computer"
        started = info.get("started") or ""
        started_disp = started
        try:
            dt = datetime.datetime.fromisoformat(started)
            started_disp = dt.strftime("%I:%M %p").lstrip("0") or dt.strftime("%H:%M")
        except Exception:
            pass
        return (
            f"Another copy of Speedy Scandocs is processing this folder "
            f"right now ({host}, started {started_disp}). "
            "Wait for it to finish, or close it and try again."
        )

    def _new_info(self) -> dict:
        now = datetime.datetime.now().isoformat()
        return {
            "host": socket.gethostname(),
            "pid": os.getpid(),
            "user": getpass.getuser(),
            "started": now,
            "heartbeat": now,
        }

    # ── public API ───────────────────────────────────────────

    def acquire(self) -> None:
        """Take the lock. Raises FolderLockHeld only when another instance
        is provably alive right now; every other failure is logged and
        swallowed so a lock problem never blocks real work."""
        try:
            existing = self._read()
            if existing is not None:
                if self._is_stale(existing):
                    logging.warning(
                        "Stale folder lock found (host=%s pid=%s heartbeat=%s) — taking over.",
                        existing.get("host"), existing.get("pid"), existing.get("heartbeat"),
                    )
                    try:
                        os.remove(self.path)
                    except FileNotFoundError:
                        pass
                else:
                    raise FolderLockHeld(self._describe(existing), info=existing)

            info = self._new_info()
            try:
                fd = os.open(self.path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            except FileExistsError:
                # Race: someone else created it between our staleness check
                # and our own atomic create. Re-read and decide again rather
                # than assuming either outcome.
                existing = self._read() or {}
                if existing and not self._is_stale(existing):
                    raise FolderLockHeld(self._describe(existing), info=existing)
                logging.warning("Folder lock race on acquire — proceeding without a lock.")
                self.acquired = False
                return
            with os.fdopen(fd, "w", encoding="utf-8") as f:
                json.dump(info, f)
            self.acquired = True
            self._last_heartbeat_monotonic = time.monotonic()
        except FolderLockHeld:
            raise
        except Exception as e:
            logging.warning(f"Could not acquire folder lock ({e}) — proceeding without it.")
            self.acquired = False

    def heartbeat(self, force: bool = False) -> None:
        """Refresh the lock's heartbeat timestamp if it's been roughly
        LOCK_HEARTBEAT_INTERVAL_SEC since the last refresh (or always, if
        force=True). Meant to be called periodically by the processing
        thread while a batch runs — no-op if this instance doesn't hold
        the lock. Never raises."""
        if not self.acquired:
            return
        now = time.monotonic()
        if not force and (now - self._last_heartbeat_monotonic) < LOCK_HEARTBEAT_INTERVAL_SEC:
            return
        try:
            info = self._read() or self._new_info()
            info["heartbeat"] = datetime.datetime.now().isoformat()
            tmp = self.path + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(info, f)
            os.replace(tmp, self.path)
            self._last_heartbeat_monotonic = now
        except Exception as e:
            logging.warning(f"Could not refresh folder lock heartbeat: {e}")

    def release(self) -> None:
        """Release the lock. Safe to call even if acquire() never
        succeeded, and safe to call more than once."""
        if not self.acquired:
            return
        try:
            os.remove(self.path)
        except Exception as e:
            logging.warning(f"Could not remove folder lock: {e}")
        self.acquired = False

    def __enter__(self) -> "FolderLock":
        self.acquire()
        return self

    def __exit__(self, exc_type, exc, tb) -> bool:
        self.release()
        return False


# ─────────────────────────────────────────────────────────────
# RenameLog — undo log for every rename/move the app performs
# ─────────────────────────────────────────────────────────────

def log_rename(batch_id: str, action: str, src: str, dst: str, source: str,
                log_path: Optional[str] = None) -> None:
    """Append one JSON-line entry to the undo log. Never raises — a failure
    to write the undo log must never block or break a real rename/move.
    `action` is "rename" or "move"; `source` is "auto" | "correction" |
    "audit" | "move"."""
    try:
        entry = {
            "ts": datetime.datetime.now().isoformat(),
            "batch_id": batch_id or "manual",
            "action": action,
            "from": os.path.abspath(src),
            "to": os.path.abspath(dst),
            "source": source,
        }
        path = log_path or RENAME_LOG_PATH
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry) + "\n")
    except Exception as e:
        logging.warning(f"Could not write undo log entry: {e}")


class RenameLog:
    """Reads back what log_rename() wrote, and can undo a whole batch."""

    def __init__(self, path: Optional[str] = None):
        self.path = path or RENAME_LOG_PATH

    def append(self, entry: dict) -> None:
        try:
            os.makedirs(os.path.dirname(self.path), exist_ok=True)
            with open(self.path, "a", encoding="utf-8") as f:
                f.write(json.dumps(entry) + "\n")
        except Exception as e:
            logging.warning(f"Could not append to rename log: {e}")

    def _read_all(self) -> List[dict]:
        entries: List[dict] = []
        try:
            with open(self.path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        entries.append(json.loads(line))
                    except Exception:
                        continue
        except FileNotFoundError:
            pass
        except Exception as e:
            logging.warning(f"Could not read rename log: {e}")
        return entries

    def last_batches(self, n: int = 10) -> List[dict]:
        """Return up to `n` most-recent batches, most-recent-first:
        [{"batch_id", "count", "first_ts", "last_ts"}, ...]."""
        by_batch: dict = {}
        order: List[str] = []
        for e in self._read_all():
            bid = e.get("batch_id", "")
            if bid not in by_batch:
                by_batch[bid] = {
                    "batch_id": bid, "count": 0,
                    "first_ts": e.get("ts"), "last_ts": e.get("ts"),
                }
                order.append(bid)
            rec = by_batch[bid]
            rec["count"] += 1
            rec["last_ts"] = e.get("ts")
        return [by_batch[b] for b in order][-n:][::-1]

    def undo_batch(self, batch_id: str) -> tuple:
        """Reverse a batch's renames/moves, most-recent-first. An entry is
        skipped (not an error) when the file is no longer at its recorded
        `to` path — someone else has since moved or renamed it — or when
        something already occupies the original `from` path. Returns
        (undone, skipped, errors)."""
        entries = [e for e in self._read_all() if e.get("batch_id") == batch_id]
        undone = 0
        skipped = 0
        errors: List[str] = []
        for e in reversed(entries):
            current_path = e.get("to", "")
            original_path = e.get("from", "")
            if not current_path or not original_path:
                skipped += 1
                continue
            if not os.path.isfile(current_path):
                skipped += 1
                continue
            if os.path.exists(original_path):
                skipped += 1
                errors.append(
                    f"{os.path.basename(original_path)}: a file already exists at the original path"
                )
                continue
            try:
                os.rename(current_path, original_path)
                undone += 1
            except Exception as ex:
                errors.append(f"{os.path.basename(current_path)}: {ex}")
        return (undone, skipped, errors)


# ─────────────────────────────────────────────────────────────
# ClientListManager
# ─────────────────────────────────────────────────────────────

class ClientListManager:

    @staticmethod
    def load(path: str) -> list:
        """Load client list and automatically add space-variant for any hyphenated names.
        e.g. 'GARCIA-TELLEZ, Miguel' also registers as 'GARCIA TELLEZ, Miguel' internally
        so OCR output (which often drops hyphens) still fuzzy-matches correctly."""
        if not path or not os.path.isfile(path):
            return []
        try:
            with open(path, "r", encoding="utf-8") as f:
                names = [
                    line.strip()
                    for line in f
                    if line.strip() and not line.startswith("#")
                ]
            # Add space-variants for hyphenated names (kept in order, deduped)
            result = []
            seen = set()
            for name in names:
                if name not in seen:
                    result.append(name)
                    seen.add(name)
                if '-' in name:
                    variant = name.replace('-', ' ')
                    if variant not in seen:
                        result.append(variant)
                        seen.add(variant)
            return result
        except Exception as e:
            logging.error(f"Could not load client list from {path}: {e}")
            return []

    @staticmethod
    def save(path: str, clients: list):
        with open(path, "w", encoding="utf-8") as f:
            for name in sorted(set(clients)):
                f.write(name + "\n")

    @staticmethod
    def _invert_name(name: str) -> str:
        """Try to convert 'First Last' → 'LAST, First' for matching.
        If the name already contains a comma, returns it unchanged."""
        if ',' in name:
            return name
        parts = name.strip().split()
        if len(parts) >= 2:
            last = parts[-1].upper()
            first = ' '.join(parts[:-1])
            return f"{last}, {first}"
        return name

    @staticmethod
    def _normalize(name: str) -> str:
        """Lowercase and replace hyphens with spaces for comparison.
        Allows 'Garcia-Tellez' to match 'GARCIA TELLEZ' and vice versa."""
        return name.lower().replace('-', ' ')

    @staticmethod
    def _all_inversions(name: str) -> list:
        """Generate all possible LAST, First splits for a multi-word name.
        e.g. "Julio Solano Trujillo" →
             ["TRUJILLO, Julio Solano",   ← standard (last word = last name)
              "SOLANO TRUJILLO, Julio"]   ← two-word last name
        Only adds extra splits for 3+ word names; single-last-word is always first.
        Excludes splits where first-name portion would be empty."""
        if ',' in name:
            return [name]   # already in LAST, First form
        parts = name.strip().split()
        if len(parts) < 2:
            return [name]
        results = []
        # Try each split point from right to left (most common last-name first)
        for split in range(len(parts) - 1, 0, -1):
            last  = ' '.join(parts[split:]).upper()
            first = ' '.join(parts[:split])
            results.append(f"{last}, {first}")
        return results

    @staticmethod
    def fuzzy_match(candidate: str, client_list: list, threshold: float = 0.82) -> Optional[str]:
        if not candidate or not client_list:
            return None
        candidate = candidate.strip().strip(".,;:'\"").strip()

        # Build normalized index (lowercase + hyphens→spaces)
        norm_list = [ClientListManager._normalize(e) for e in client_list]

        # 1. Exact match after normalization — always preferred
        candidate_norm = ClientListManager._normalize(candidate)
        for i, entry_norm in enumerate(norm_list):
            if entry_norm == candidate_norm:
                return client_list[i]

        # 2. Try all inversion forms — standard split first, then multi-word last names.
        inversions = ClientListManager._all_inversions(candidate)

        # Also generate truncated-last-name forms:
        # "SOLANO TRUJILLO, Julio" → also try "SOLANO, Julio"
        # This handles AI returning 3-word names where only one surname is in the list.
        truncated = set()
        for inv in inversions:
            if ',' in inv:
                last_part, first_part = inv.split(',', 1)
                last_words = last_part.strip().split()
                if len(last_words) > 1:
                    truncated.add(f"{last_words[0]}, {first_part.strip()}")

        for inv in list(inversions) + list(truncated):
            inv_norm = ClientListManager._normalize(inv)
            if inv_norm == candidate_norm:
                continue
            for i, entry_norm in enumerate(norm_list):
                if entry_norm == inv_norm:
                    return client_list[i]   # exact match on this inversion

        # 3. Fuzzy match — score all inversion forms including truncated, keep best
        all_forms = {candidate_norm} | {
            ClientListManager._normalize(inv) for inv in list(inversions) + list(truncated)
        }
        best_entry = None
        best_score = 0.0
        for form in all_forms:
            for i, entry_norm in enumerate(norm_list):
                score = difflib.SequenceMatcher(None, form, entry_norm).ratio()
                if score > best_score:
                    best_score = score
                    best_entry = client_list[i]

        if best_entry and best_score >= threshold:
            return best_entry

        # 4. Compound surname prefix match — handles OCR merging like
        #    "Delgadillocuellar" where OCR drops the space between two last names.
        #    Extract the last-name portion of the candidate (before comma, or last word).
        if ',' in candidate:
            cand_last = candidate.split(',')[0].strip()
        else:
            parts = candidate.strip().split()
            cand_last = parts[-1] if parts else candidate
        cand_last_norm = ClientListManager._normalize(cand_last)

        # cand_last must be strictly longer than the entry's last name — otherwise
        # this is just a plain last-name match, not a merged compound surname.
        if len(cand_last_norm) >= 7:
            # Extract candidate's first name for cross-checking
            if ',' in candidate:
                cand_first_norm = ClientListManager._normalize(candidate.split(',')[1].strip())
            else:
                parts = candidate.strip().split()
                cand_first_norm = ClientListManager._normalize(parts[0]) if len(parts) > 1 else ""

            prefix_matches = []
            for i, entry in enumerate(client_list):
                entry_last = ClientListManager._normalize(
                    entry.split(',')[0].strip() if ',' in entry else entry.split()[-1]
                )
                # Require cand_last is longer (genuinely merged surname, not same last name)
                if len(entry_last) >= 6 and len(cand_last_norm) > len(entry_last) and cand_last_norm.startswith(entry_last):
                    # Cross-check first name if we have one
                    if cand_first_norm:
                        entry_first_norm = ClientListManager._normalize(
                            entry.split(',')[1].strip() if ',' in entry else ""
                        )
                        # First names must share a common start (handles abbreviations/nicknames)
                        if entry_first_norm and not (
                            entry_first_norm.startswith(cand_first_norm) or
                            cand_first_norm.startswith(entry_first_norm)
                        ):
                            continue  # First name mismatch — skip
                    prefix_matches.append(i)

            # Only act when exactly one client matches — ambiguity → NEEDS_REVIEW
            if len(prefix_matches) == 1:
                logging.info(
                    f"fuzzy_match: prefix match '{cand_last_norm}' → "
                    f"'{client_list[prefix_matches[0]]}'"
                )
                return client_list[prefix_matches[0]]

        return None

    @staticmethod
    def is_valid_format(name: str) -> bool:
        """Accepts 'LAST, First' or 'LAST, First Middle' etc."""
        return bool(re.match(r"^[A-Za-z\-\'\. ]+,\s+[A-Za-z\-\'\. ]+$", name.strip()))

    @staticmethod
    def filter_candidates(doc_text: str, client_list: list, top_n: int = 10) -> list:
        """Pre-filter the client list to the top_n most likely candidates.

        Strategy: only extract text from labeled keyword regions (Client:, RE:,
        Insured:, etc.) — NOT the full document.  Using the full document caused
        the original anchoring failure (facility name 'Velazquez Pain Summerlin'
        → model picked VELAZQUEZ, Aida).  Restricting to labeled regions gives a
        clean signal of who the document is actually about.

        Falls back to the full list when no labeled regions are found.
        """
        if not doc_text or not client_list:
            return client_list

        # ── Step 1: extract text near labeled keywords only ───────────────────
        # Each pattern captures up to ~60 chars after the label.
        _label_re = re.compile(
            r"(?:Client(?:\s*/\s*Patient)?(?:\s+Name)?|Claimant|Clmt|Injured(?:\s+(?:Worker|Party))?|"
            r"Patient(?:\s+Name)?|Insured(?:\s+Name)?|Named\s+Insured|Employee|"
            r"RE|Re|Regarding|Subject)\s*[:\-\.]\s*"
            r"([A-Za-z][A-Za-z ,\-\'\.]{1,60})",
            re.IGNORECASE,
        )
        labeled_phrases = [m.group(1).strip() for m in _label_re.finditer(doc_text)]

        # ── Step 2: score every client against the labeled phrases ────────────
        if labeled_phrases:
            scored: list = []
            for client in client_list:
                client_lower = client.lower()
                last_name = client.split(",")[0].strip().lower()
                best = 0.0
                for phrase in labeled_phrases:
                    phrase_lower = phrase.lower()
                    # Exact last-name hit in the phrase → strong signal
                    if len(last_name) >= 3 and last_name in phrase_lower:
                        best = max(best, 0.90)
                        continue
                    ratio = difflib.SequenceMatcher(None, phrase_lower, client_lower).ratio()
                    if ratio > best:
                        best = ratio
                scored.append((best, client))

            scored.sort(key=lambda x: x[0], reverse=True)
            top_score = scored[0][0] if scored else 0.0

            if top_score >= 0.50:
                short_list = [name for _score, name in scored[:top_n]]
                logging.info(
                    f"filter_candidates: {len(client_list)} → {len(short_list)} candidates "
                    f"(top score {top_score:.2f}, phrases: {labeled_phrases})"
                )
                return short_list

        # No labeled regions found or no strong score → fall back to full list
        logging.info(
            f"filter_candidates: no labeled keyword regions found, using full list "
            f"({len(client_list)} clients)"
        )
        return client_list


# ─────────────────────────────────────────────────────────────
# DocumentTypeManager / ProviderManager (PASS 4)
# ─────────────────────────────────────────────────────────────
#
# Same load/save pattern as ClientListManager: a plain newline-delimited
# text file, "#" comments and blank lines ignored. Both are inert unless a
# classification.use_document_types / classification.use_providers flag
# turns them on — see FileProcessor.process_file.

class DocumentTypeManager:

    # Minimum alias length eligible for whole-word containment matching.
    _CONTAINMENT_MIN_LEN = 10

    @staticmethod
    def _seed_if_missing(path: str) -> None:
        """Write DEFAULT_DOCUMENT_TYPES to `path` if nothing is there yet —
        mirrors how client_list.txt is expected to already exist, except
        this file is safe to auto-populate since the seed list is generic
        firm vocabulary, not client data."""
        if not path or os.path.isfile(path):
            return
        try:
            dirpath = os.path.dirname(path)
            if dirpath:
                os.makedirs(dirpath, exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                f.write("# Document types, one per line.\n")
                f.write("# Canonical Name | alias1 | alias2 ...\n")
                f.write("# Lines starting with # are ignored.\n")
                for line in DEFAULT_DOCUMENT_TYPES:
                    f.write(line + "\n")
        except Exception as e:
            logging.error(f"Could not seed document types file at {path}: {e}")

    @staticmethod
    def load(path: str) -> list:
        """Return the canonical document type names (first field of each
        'Canonical | alias...' line). Seeds the file with
        DEFAULT_DOCUMENT_TYPES on first run if it doesn't exist yet."""
        if not path:
            return []
        DocumentTypeManager._seed_if_missing(path)
        if not os.path.isfile(path):
            return []
        try:
            names = []
            with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    canonical = line.split("|", 1)[0].strip()
                    if canonical:
                        names.append(canonical)
            return names
        except Exception as e:
            logging.error(f"Could not load document types from {path}: {e}")
            return []

    @staticmethod
    def load_alias_map(path: str) -> dict:
        """Return {lowercased alias-or-canonical: canonical name} for every
        name on every line, so 'physician progress report' and 'ppr' both
        resolve to 'PPR'. Seeds the file on first run, same as load()."""
        if not path:
            return {}
        DocumentTypeManager._seed_if_missing(path)
        if not os.path.isfile(path):
            return {}
        alias_map = {}
        try:
            with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    parts = [p.strip() for p in line.split("|") if p.strip()]
                    if not parts:
                        continue
                    canonical = parts[0]
                    for name in parts:
                        alias_map[name.lower()] = canonical
            return alias_map
        except Exception as e:
            logging.error(f"Could not load document type aliases from {path}: {e}")
            return {}

    @staticmethod
    def _split_file(path: str) -> tuple:
        """Read `path` into (header_comment_lines, {canonical: whole line}).

        The returned line is kept VERBATIM — aliases, spacing and all — so
        save() can write an untouched entry back exactly as the user wrote
        it. Missing/unreadable file yields ([], {})."""
        header: List[str] = []
        entries: dict = {}
        if not path or not os.path.isfile(path):
            return header, entries
        try:
            with open(path, "r", encoding="utf-8") as f:
                seen_entry = False
                for raw in f:
                    line = raw.rstrip("\n")
                    stripped = line.strip()
                    if not stripped or stripped.startswith("#"):
                        # Only the comment block ABOVE the first entry is a
                        # header; comments further down belong to whatever
                        # the user was annotating and are dropped on rewrite.
                        if not seen_entry:
                            header.append(line)
                        continue
                    seen_entry = True
                    canonical = stripped.split("|", 1)[0].strip()
                    if canonical:
                        entries[canonical] = stripped
        except Exception as e:
            logging.error(f"Could not read document types from {path}: {e}")
            return [], {}
        return header, entries

    @staticmethod
    def save(path: str, canonical_names: list) -> None:
        """Rewrite the file with exactly `canonical_names`, PRESERVING each
        surviving entry's aliases and the explanatory comment header.

        This used to be a flat rewrite of canonical names only, which
        silently destroyed every alias in the file — so the first press of
        the Document Types tab's Save button, or the first accepted
        doc-type suggestion, wiped the vocabulary that
        DocumentTypeManager.normalize matches against ("Compromise Offer"
        stopped resolving to "Reduction Request"). Nothing surfaced the
        loss: no error, no log line, and the tab still listed the same
        canonical names — only naming quality quietly degraded.

        A name dropped from `canonical_names` takes its aliases with it:
        the file format is "Canonical | alias | alias", so an alias cannot
        outlive the canonical it points at.
        """
        header, existing = DocumentTypeManager._split_file(path)
        if not header:
            header = [
                "# Document types, one per line.",
                "# Canonical Name | alias1 | alias2 ...",
                "# Lines starting with # are ignored.",
            ]
        dirpath = os.path.dirname(path)
        if dirpath:
            os.makedirs(dirpath, exist_ok=True)
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            for line in header:
                f.write(line + "\n")
            for name in sorted(set(canonical_names)):
                # Write the original line (aliases intact) when we have one.
                f.write(existing.get(name.strip(), name.strip()) + "\n")
        os.replace(tmp, path)

    @staticmethod
    def normalize(raw: str, alias_map: dict, threshold: float = 0.85) -> str:
        """Map a raw model-returned doc type string to its canonical name.

        Exact alias/canonical hit first (case-insensitive), then a difflib
        fuzzy match against every known alias/canonical string. Returns ""
        when nothing is close enough — the caller should then keep the
        model's raw printed title rather than discard it."""
        if not raw or not alias_map:
            return ""
        raw_norm = raw.strip().lower()
        if not raw_norm:
            return ""
        if raw_norm in alias_map:
            return alias_map[raw_norm]

        # Containment pass — the model often returns the real title padded with
        # extra description, e.g. "Compromise Offer Letter Medical Bill" or
        # "Physician Chiropractor Progress Report Disability Certification".
        # A whole-string fuzzy ratio scores those below threshold, so look for a
        # known alias appearing as a whole-word substring. Longest match wins so
        # "Physician Chiropractor Progress Report" beats the shorter, more
        # generic "Progress Report". Keys under _CONTAINMENT_MIN_LEN are skipped
        # so short ones like "Lien" or "PPR" can't match inside unrelated titles.
        contained = [
            key for key in alias_map
            if len(key) >= DocumentTypeManager._CONTAINMENT_MIN_LEN
            and re.search(r"\b" + re.escape(key) + r"\b", raw_norm)
        ]
        if contained:
            return alias_map[max(contained, key=len)]

        best_key = None
        best_score = 0.0
        for key in alias_map:
            score = difflib.SequenceMatcher(None, raw_norm, key).ratio()
            if score > best_score:
                best_score = score
                best_key = key
        if best_key is not None and best_score >= threshold:
            return alias_map[best_key]
        return ""


class ProviderManager:

    @staticmethod
    def _seed_if_missing(path: str) -> None:
        """Write only the explanatory comment header — providers.txt ships
        empty and fills in from use, unlike document_types.txt."""
        if not path or os.path.isfile(path):
            return
        try:
            dirpath = os.path.dirname(path)
            if dirpath:
                os.makedirs(dirpath, exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                for line in _PROVIDERS_FILE_HEADER:
                    f.write(line + "\n")
        except Exception as e:
            logging.error(f"Could not seed providers file at {path}: {e}")

    @staticmethod
    def load(path: str) -> list:
        if not path:
            return []
        ProviderManager._seed_if_missing(path)
        if not os.path.isfile(path):
            return []
        try:
            with open(path, "r", encoding="utf-8") as f:
                # Test the STRIPPED line for the comment marker — an
                # indented "  # note" is a comment too, not a provider named
                # "# note".
                return [
                    line.strip() for line in f
                    if line.strip() and not line.strip().startswith("#")
                ]
        except Exception as e:
            logging.error(f"Could not load providers from {path}: {e}")
            return []

    @staticmethod
    def save(path: str, providers: list) -> None:
        """Rewrite the provider list, keeping the explanatory comment
        header. providers.txt has no aliases, so unlike
        DocumentTypeManager.save there is nothing else to preserve — but a
        flat rewrite still threw the header away, leaving a user who opened
        the file with no idea what it was for."""
        header: List[str] = []
        if path and os.path.isfile(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    for raw in f:
                        line = raw.rstrip("\n")
                        stripped = line.strip()
                        if not stripped:
                            continue
                        if not stripped.startswith("#"):
                            break   # reached the first real entry
                        header.append(line)
            except Exception as e:
                logging.warning(f"Could not read providers header from {path}: {e}")
                header = []
        if not header:
            header = list(_PROVIDERS_FILE_HEADER)
        dirpath = os.path.dirname(path)
        if dirpath:
            os.makedirs(dirpath, exist_ok=True)
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            for line in header:
                f.write(line + "\n")
            for name in sorted(set(providers)):
                f.write(name.strip() + "\n")
        os.replace(tmp, path)

    # Trailing address noise: a comma, or a run of digits, onward — e.g.
    # "Chiropractic Works, 5105 E. Sahara Ave Ste 144" -> "Chiropractic Works",
    # "First Care Industrial Medicine 89142" -> "First Care Industrial Medicine".
    _ADDRESS_NOISE_RE = re.compile(r"(,|\s+\d).*$", re.DOTALL)

    @staticmethod
    def strip_address(raw: str) -> str:
        """Strip trailing address noise from a raw recipient/sender string."""
        if not raw:
            return raw
        cleaned = ProviderManager._ADDRESS_NOISE_RE.sub("", raw).strip()
        return cleaned or raw.strip()

    @staticmethod
    def normalize(raw: str, providers: list, threshold: float = 0.85) -> str:
        """Strip address noise from `raw`, then exact/fuzzy match it against
        `providers`. Returns "" when no known provider is close enough —
        the caller should keep the raw (address-stripped) value rather than
        discard it."""
        if not raw:
            return ""
        cleaned = ProviderManager.strip_address(raw)
        if not providers or not cleaned:
            return ""
        cleaned_norm = cleaned.lower()
        for p in providers:
            if p.strip().lower() == cleaned_norm:
                return p
        best = None
        best_score = 0.0
        for p in providers:
            score = difflib.SequenceMatcher(None, cleaned_norm, p.strip().lower()).ratio()
            if score > best_score:
                best_score = score
                best = p
        if best is not None and best_score >= threshold:
            return best
        return ""


# ─────────────────────────────────────────────────────────────
# LearningStore (PASS 6) — the learning loop
# ─────────────────────────────────────────────────────────────
#
# The problem this exists to solve: a named client (Mary) can have other
# people attached to her matter — e.g. George was a passenger in Mary's
# accident. George receives mail addressed to him, but it has to be filed
# under Mary, because she's the client. An employee knows this; the machine
# doesn't, and there are too many such relationships to enter by hand.
#
# Before this pass, FileProcessor.process_file already computed
# `raw_client` (the model's free-text guess, e.g. "George Martinez") but
# threw it away the moment fuzzy_match against the client list failed —
# only the unknown-client label survived. This pass captures that guess
# instead, backlogs it every time an employee corrects a file (via the
# Manual Correction panel, the office's actual workflow — NOT the Audit
# checkboxes, which the office doesn't use), and once enough distinct
# documents agree on the same raw-name -> client pairing, surfaces it as a
# suggestion for a human to confirm. Confirmed pairings are then applied
# automatically to the *next* George document, and can be swept
# retroactively across every George file the machine already produced.
#
# Storage is a single append-only JSONL log (corrections.jsonl — one line
# per correction/confirmation/observation, human-inspectable, never
# rewritten) plus a derived JSON index (learning_index.json) rebuilt from
# it. No sqlite — this is intentionally simple.
#
# Guards on alias promotion (all four matter — see rebuild_index):
#   1. Surname required — a bare first name ("George") can never become an
#      alias. Common first names recur constantly in this client base; a
#      bare-first-name alias would misfile everything.
#   2. Distinct documents only — the same file logged twice (has happened
#      in production) must not manufacture evidence. Counted by doc_hash.
#   3. Agreement — >= observations_required distinct docs AND >= 90% of
#      them pointing at ONE client. Otherwise "ambiguous" and never
#      surfaced — George may be a passenger in two different accidents, or
#      there may be two Georges.
#   4. Never auto-confirmed — promotion only ever produces a *suggestion*.
#      Only confirm_alias() (called by a human, via the Pass 7 UI) can set
#      status "confirmed" and make lookup_alias() start returning it.

class LearningStore:
    """Correction log + derived lookup index, both under the app's user
    data directory. See the module comment above for the design."""

    LOG_FILENAME = "corrections.jsonl"
    INDEX_FILENAME = "learning_index.json"

    # ── Growth control ───────────────────────────────────────────────────
    # The log is append-only and grows for the life of an installation:
    # log_observation() writes one line per unresolved file per batch, so a
    # busy office adds thousands of lines a week. Two guards keep that from
    # turning into a slow, then unusable, app:
    #
    #   MAX_FEWSHOT_SCAN  find_similar_corrections runs a difflib ratio per
    #                     entry and is called ONCE PER DOCUMENT during a
    #                     batch. Unbounded, that measured 0.24s at 500
    #                     entries, 0.93s at 2k and 4.87s at 10k — per file,
    #                     getting worse every run. Only the most recent
    #                     entries are scanned; older corrections are still
    #                     kept, indexed and counted as evidence, they just
    #                     stop being candidate few-shot examples.
    #
    #   MAX_LOG_ENTRIES   hard ceiling on the log itself. Trimming keeps the
    #                     newest entries; alias/claim decisions already
    #                     confirmed by a human live in the index's
    #                     *_decisions maps and survive trimming (see
    #                     rebuild_index), so nothing a human decided is lost.
    MAX_FEWSHOT_SCAN = 2000
    MAX_LOG_ENTRIES = 50000
    TRIM_TO_ENTRIES = 40000

    # Business names harvested out of a corrected description's "to X" /
    # "from X" phrasing (e.g. "Reduction Request to Chiropractic Works").
    _PROVIDER_RE = re.compile(
        r"\b(?:to|from)\s+([A-Z][A-Za-z.&'\-]+(?:\s+[A-Z][A-Za-z.&'\-]+){0,3})"
    )

    def __init__(self, user_data_dir: str, observations_required: int = 3,
                  sentinel_labels: Optional[list] = None):
        self.dir = user_data_dir
        self.log_path = os.path.join(user_data_dir, self.LOG_FILENAME)
        self.index_path = os.path.join(user_data_dir, self.INDEX_FILENAME)
        self.observations_required = max(1, int(observations_required or 3))
        # Firm-customized unknown-client labels (naming.unknown_client_label /
        # naming.no_client_label). Checked alongside the built-in sentinels
        # so a renamed placeholder can't slip through as a real client.
        self.sentinel_labels = list(sentinel_labels or [])
        # This store is reached from BOTH the UI thread (correction commits,
        # Suggestions tab accept/reject) and the batch worker thread
        # (alias/claim lookups, observations). Every log append and every
        # index mutation goes through this lock so a rebuild can't race an
        # append and silently drop evidence.
        self._lock = threading.RLock()
        self._observations_since_trim_check = 0
        self.index = self._load_index()

    def is_sentinel(self, name: str) -> bool:
        """True if `name` is an unresolved-placeholder label rather than a
        real client — see FileProcessor.is_sentinel_client."""
        return FileProcessor.is_sentinel_client(name, self.sentinel_labels)

    # ── normalization helpers ───────────────────────────────────────────

    @staticmethod
    def _normalize_name(name: str) -> str:
        return re.sub(r"\s+", " ", (name or "").strip().lower())

    @staticmethod
    def _normalize_for_similarity(text: str) -> str:
        return re.sub(r"\s+", " ", (text or "")).strip().lower()

    # ── index persistence ────────────────────────────────────────────────

    @staticmethod
    def _empty_index() -> dict:
        return {
            "alias_candidates": {},
            "doc_type_candidates": {},
            "provider_candidates": {},
            "claim_index": {},
            # Human decisions (confirm_alias/reject_alias), keyed the same
            # as alias_candidates. Kept separate from the computed stats so
            # a rebuild never silently forgets a decision a human already
            # made.
            "alias_decisions": {},
            # Same idea for the Pass 7 Suggestions tab's doc-type/provider
            # rows (accept_doc_type_candidate / reject_doc_type_candidate
            # and their provider equivalents) — without this an "Ignore"
            # click would just reappear on the next rebuild.
            "doc_type_decisions": {},
            "provider_decisions": {},
        }

    def _load_index(self) -> dict:
        empty = self._empty_index()
        if not os.path.isfile(self.index_path):
            return empty
        try:
            with open(self.index_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                return empty
            for key, default_val in empty.items():
                data.setdefault(key, default_val)
            return data
        except Exception as e:
            logging.warning(f"Could not load learning index: {e}")
            return empty

    def _write_index(self) -> None:
        tmp = self.index_path + ".tmp"
        try:
            os.makedirs(self.dir, exist_ok=True)
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(self.index, f, indent=2)
            os.replace(tmp, self.index_path)
        except Exception as e:
            logging.warning(f"Could not write learning index: {e}")

    # ── the log itself ───────────────────────────────────────────────────

    def _read_all_corrections(self) -> List[dict]:
        entries: List[dict] = []
        try:
            with open(self.log_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        entries.append(json.loads(line))
                    except Exception:
                        continue
        except FileNotFoundError:
            pass
        except Exception as e:
            logging.warning(f"Could not read corrections log: {e}")
        return entries

    def _append_entry(self, entry: dict) -> None:
        entry = dict(entry)
        entry.setdefault("ts", datetime.datetime.now().isoformat())
        with self._lock:
            os.makedirs(self.dir, exist_ok=True)
            with open(self.log_path, "a", encoding="utf-8") as f:
                f.write(json.dumps(entry) + "\n")

    def _trim_log_if_needed(self) -> bool:
        """Keep corrections.jsonl bounded (see MAX_LOG_ENTRIES). Rewrites
        the file with only the newest TRIM_TO_ENTRIES lines, atomically.
        Returns True if a trim happened. Never raises — a trim failure is
        a housekeeping problem, not a reason to break a correction."""
        try:
            with self._lock:
                if not os.path.isfile(self.log_path):
                    return False
                with open(self.log_path, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                if len(lines) <= self.MAX_LOG_ENTRIES:
                    return False
                keep = lines[-self.TRIM_TO_ENTRIES:]
                tmp = self.log_path + ".tmp"
                with open(tmp, "w", encoding="utf-8") as f:
                    f.writelines(keep)
                os.replace(tmp, self.log_path)
                logging.info(
                    f"Trimmed corrections log from {len(lines)} to {len(keep)} entries."
                )
                return True
        except Exception as e:
            logging.warning(f"Could not trim corrections log: {e}")
            return False

    def log_correction(self, entry: dict) -> None:
        """Append one correction/confirmation/audit entry and refresh the
        derived index. Never raises — a logging failure must never block or
        break the actual rename. This is human-paced (one call per
        employee action in the Manual Correction panel or Audit submit),
        so rebuilding the index on every call is cheap in practice.

        Expected shape (see module comment for field meanings):
            {"ts", "doc_hash", "original_name", "predicted_client",
             "raw_client", "predicted_desc", "predicted_doc_type",
             "predicted_recipient", "corrected_client", "corrected_desc",
             "claim_number", "doc_date", "text_excerpt",
             "changed_client", "changed_desc", "source"}
        `source` is "correction" | "confirmation" | "audit" | "observation".
        """
        try:
            self._append_entry(entry)
        except Exception as e:
            logging.warning(f"Could not log correction: {e}")
            return
        self._trim_log_if_needed()
        try:
            self.rebuild_index()
        except Exception as e:
            logging.warning(f"Could not rebuild learning index: {e}")

    def log_observation(self, doc_hash: str, raw_client: str, original_name: str,
                         claim_number: str = "", predicted_desc: str = "",
                         text_excerpt: str = "") -> None:
        """Record the model's raw guess for a file that stayed unresolved,
        WITHOUT any employee correction — learning.client_relationships or
        claim_linking set to "suggest" (never applies, only records), or
        "auto" that found nothing to apply. This is the literal "backlog"
        the owner described: not visible anywhere, just recorded, so that a
        later confirmed alias can retroactively sweep every other unrenamed
        file that carried the same raw guess (see plan_retroactive_renames).

        Deliberately does NOT rebuild the index — this can fire once per
        unresolved file in a batch of thousands, and observations carry no
        corrected_client so they never feed alias_candidates evidence
        anyway (see rebuild_index). Never raises."""
        try:
            self._append_entry({
                "doc_hash": doc_hash,
                "original_name": original_name,
                "predicted_client": "",
                "raw_client": raw_client,
                "predicted_desc": predicted_desc,
                "predicted_doc_type": "",
                "predicted_recipient": "",
                "corrected_client": "",
                "corrected_desc": "",
                "claim_number": claim_number,
                "doc_date": "",
                "text_excerpt": (text_excerpt or "")[:600],
                "changed_client": False,
                "changed_desc": False,
                "source": "observation",
            })
            # Observations are the fastest-growing source of log lines (one
            # per unresolved file per batch). Checking the size on every
            # append would re-read the whole file per document, so amortize
            # it — a few hundred lines of overshoot past MAX_LOG_ENTRIES is
            # harmless.
            self._observations_since_trim_check += 1
            if self._observations_since_trim_check >= 500:
                self._observations_since_trim_check = 0
                self._trim_log_if_needed()
        except Exception as e:
            logging.warning(f"Could not log observation: {e}")

    # ── rebuilding the index ─────────────────────────────────────────────

    def rebuild_index(self) -> dict:
        """Walk the whole JSONL log and recompute alias_candidates,
        doc_type_candidates, provider_candidates, and claim_index from
        scratch, then persist atomically. Human decisions recorded via
        confirm_alias/reject_alias (alias_decisions) always take priority
        over the freshly-computed guard result for that name.

        Holds the store lock for the whole read-compute-write cycle so a
        concurrent append from the batch worker can't land between the read
        and the write and be lost."""
        with self._lock:
            return self._rebuild_index_locked()

    def _rebuild_index_locked(self) -> dict:
        entries = self._read_all_corrections()
        alias_stats: dict = {}
        doc_type_stats: dict = {}
        provider_stats: dict = {}
        claim_stats: dict = {}

        for e in entries:
            raw_client = (e.get("raw_client") or "").strip()
            corrected_client = (e.get("corrected_client") or "").strip()
            predicted_client = (e.get("predicted_client") or "").strip()
            doc_hash = (e.get("doc_hash") or "").strip()
            ts = e.get("ts") or ""

            # Alias evidence only counts when the file was genuinely
            # unresolved beforehand (predicted_client empty or one of the
            # model's own NEEDS_REVIEW-style sentinels) — otherwise a
            # no-op "confirmation" of an already-correct match (raw_client
            # "Mary Smith" confirmed against canonical "SMITH, Mary") would
            # manufacture a pointless alias suggestion for a client that
            # was never actually unresolved.
            was_unresolved = (not predicted_client) or (
                predicted_client.upper() in FileProcessor._CLIENT_SENTINELS
            )
            # The corrected side must be a REAL client. A no-change commit
            # on a still-unresolved row (browsing rows with the Manual
            # Correction panel open does exactly this) records
            # corrected_client == "A-NEEDS REVIEW"; counting that as
            # evidence would eventually surface 'file <raw name> under
            # A-NEEDS REVIEW' as a suggestion, and accepting it would
            # rename real documents to the placeholder while reporting them
            # as successfully renamed.
            corrected_is_real = bool(corrected_client) and not self.is_sentinel(corrected_client)
            if raw_client and corrected_is_real and was_unresolved \
                    and FileProcessor._looks_like_person_name(raw_client):
                norm = self._normalize_name(raw_client)
                rec = alias_stats.setdefault(norm, {
                    "raw_display": raw_client, "resolved": {}, "doc_hashes": set(),
                    "first_seen": ts, "last_seen": ts,
                })
                rec["resolved"][corrected_client] = rec["resolved"].get(corrected_client, 0) + 1
                if doc_hash:
                    rec["doc_hashes"].add(doc_hash)
                if ts:
                    if not rec["first_seen"] or ts < rec["first_seen"]:
                        rec["first_seen"] = ts
                    if not rec["last_seen"] or ts > rec["last_seen"]:
                        rec["last_seen"] = ts

            # Doc-type candidates: a genuinely changed description.
            corrected_desc = (e.get("corrected_desc") or "").strip()
            if corrected_desc and e.get("changed_desc"):
                norm_desc = self._normalize_name(corrected_desc)
                drec = doc_type_stats.setdefault(norm_desc, {"count": 0, "examples": []})
                drec["count"] += 1
                if corrected_desc not in drec["examples"] and len(drec["examples"]) < 5:
                    drec["examples"].append(corrected_desc)

                # Provider candidates: harvested from "to X" / "from X" in
                # the same corrected description.
                for m in self._PROVIDER_RE.finditer(corrected_desc):
                    prov = m.group(1).strip().rstrip(".")
                    if len(prov) < 3:
                        continue
                    norm_prov = self._normalize_name(prov)
                    prec = provider_stats.setdefault(norm_prov, {"count": 0, "examples": []})
                    prec["count"] += 1
                    if prov not in prec["examples"] and len(prec["examples"]) < 5:
                        prec["examples"].append(prov)

            # Same gate for claim linking — a claim number must never be
            # learned as pointing at an unresolved-placeholder label.
            claim_number = (e.get("claim_number") or "").strip()
            if claim_number and corrected_is_real:
                crec = claim_stats.setdefault(claim_number, {})
                crec[corrected_client] = crec.get(corrected_client, 0) + 1

        decisions = self.index.get("alias_decisions", {})
        alias_candidates: dict = {}
        for norm, rec in alias_stats.items():
            doc_hashes = sorted(rec["doc_hashes"])
            resolved = rec["resolved"]
            decision = decisions.get(norm)
            if decision:
                status = decision["status"]
                confirmed_client = decision.get("client", "")
            else:
                confirmed_client = ""
                if resolved:
                    best_client, best_count = max(resolved.items(), key=lambda kv: kv[1])
                    total_obs = sum(resolved.values())
                    agreement = (best_count / total_obs) if total_obs else 0.0
                else:
                    agreement = 0.0
                # NOTE: the distinct-document-count guard (>=
                # observations_required) is enforced by pending_suggestions(),
                # not here — this keeps "status" a pure read of agreement,
                # while "surfaced as pending" additionally requires enough
                # evidence. A same-hash-3x entry can be status "pending"
                # with only 1 distinct doc; pending_suggestions() filters
                # it out until more distinct documents agree.
                status = "pending" if agreement >= 0.90 else "ambiguous"
            entry = {
                "raw_display": rec["raw_display"],
                "resolved": resolved,
                "doc_hashes": doc_hashes,
                "first_seen": rec["first_seen"],
                "last_seen": rec["last_seen"],
                "status": status,
            }
            if confirmed_client:
                entry["confirmed_client"] = confirmed_client
            alias_candidates[norm] = entry

        # A decision made on a name with no current supporting stats (e.g.
        # rejected, then all its corrections happen to age out of a future
        # trimmed log) must not vanish from the index.
        for norm, decision in decisions.items():
            if norm in alias_candidates:
                continue
            entry = {
                "raw_display": decision.get("raw_display", norm),
                "resolved": {}, "doc_hashes": [], "first_seen": "", "last_seen": "",
                "status": decision["status"],
            }
            if decision.get("client"):
                entry["confirmed_client"] = decision["client"]
            alias_candidates[norm] = entry

        self.index = {
            "alias_candidates": alias_candidates,
            "doc_type_candidates": doc_type_stats,
            "provider_candidates": provider_stats,
            "claim_index": claim_stats,
            "alias_decisions": decisions,
            "doc_type_decisions": self.index.get("doc_type_decisions", {}),
            "provider_decisions": self.index.get("provider_decisions", {}),
        }
        self._write_index()
        return self.index

    # ── suggestions / decisions (Pass 7 UI calls these) ────────────────

    def pending_suggestions(self) -> list:
        """Alias, doc-type, and provider candidates that have cleared their
        guards but have not yet been confirmed or rejected by a human.
        Never mutates anything — only confirm_alias/reject_alias do."""
        out = []
        for norm, rec in self.index.get("alias_candidates", {}).items():
            if rec.get("status") != "pending":
                continue
            doc_hashes = rec.get("doc_hashes", [])
            if len(doc_hashes) < self.observations_required:
                continue
            resolved = rec.get("resolved", {})
            if not resolved:
                continue
            best_client, _ = max(resolved.items(), key=lambda kv: kv[1])
            # An index written before the sentinel guard existed can still
            # hold placeholder-valued candidates — never offer one.
            if self.is_sentinel(best_client):
                continue
            out.append({
                "kind": "alias",
                "raw_name": rec.get("raw_display", norm),
                "resolved_client": best_client,
                "observations": len(doc_hashes),
                "doc_hashes": list(doc_hashes),
                "first_seen": rec.get("first_seen", ""),
                "last_seen": rec.get("last_seen", ""),
            })

        doc_type_decisions = self.index.get("doc_type_decisions", {})
        for norm, rec in self.index.get("doc_type_candidates", {}).items():
            if norm in doc_type_decisions:
                continue
            if rec.get("count", 0) >= self.observations_required:
                examples = rec.get("examples", [])
                out.append({
                    "kind": "doc_type",
                    "name": examples[0] if examples else norm,
                    "count": rec.get("count", 0),
                    "examples": examples,
                })

        provider_decisions = self.index.get("provider_decisions", {})
        for norm, rec in self.index.get("provider_candidates", {}).items():
            if norm in provider_decisions:
                continue
            if rec.get("count", 0) >= self.observations_required:
                examples = rec.get("examples", [])
                out.append({
                    "kind": "provider",
                    "name": examples[0] if examples else norm,
                    "count": rec.get("count", 0),
                    "examples": examples,
                })

        return out

    def confirm_alias(self, raw_name: str, client: str) -> None:
        """Human confirmation (Pass 7 UI) — the only thing that can make
        lookup_alias() start returning a value for this raw name."""
        norm = self._normalize_name(raw_name)
        if not norm:
            return
        # Refuse to confirm a placeholder as a client, even if a stale
        # index (written before this guard existed) offered one.
        if not client or self.is_sentinel(client):
            logging.warning(
                f"Refusing to confirm alias '{raw_name}' -> '{client}': "
                "that is an unresolved-placeholder label, not a client."
            )
            return
        with self._lock:
            self.index.setdefault("alias_decisions", {})[norm] = {
                "status": "confirmed", "client": client, "raw_display": raw_name,
            }
            self.rebuild_index()

    def reject_alias(self, raw_name: str) -> None:
        """Human rejection (Pass 7 UI) — permanently dismisses this raw
        name so it stops appearing in pending_suggestions()."""
        norm = self._normalize_name(raw_name)
        if not norm:
            return
        with self._lock:
            existing = self.index.get("alias_decisions", {}).get(norm, {})
            self.index.setdefault("alias_decisions", {})[norm] = {
                "status": "rejected", "client": "",
                "raw_display": existing.get("raw_display", raw_name),
            }
            self.rebuild_index()

    def accept_doc_type_candidate(self, name: str) -> None:
        """Human accepted a suggested document type (Pass 7 Suggestions
        tab). Marks it decided so it stops appearing as pending — the UI
        is responsible for actually adding `name` to document_types.txt
        via DocumentTypeManager."""
        norm = self._normalize_name(name)
        if not norm:
            return
        with self._lock:
            self.index.setdefault("doc_type_decisions", {})[norm] = {
                "status": "accepted", "name": name,
            }
            self.rebuild_index()

    def reject_doc_type_candidate(self, name: str) -> None:
        """Human dismissed a suggested document type — permanently stops
        it from reappearing in pending_suggestions()."""
        norm = self._normalize_name(name)
        if not norm:
            return
        with self._lock:
            self.index.setdefault("doc_type_decisions", {})[norm] = {
                "status": "rejected", "name": name,
            }
            self.rebuild_index()

    def accept_provider_candidate(self, name: str) -> None:
        """Human accepted a suggested provider. Marks it decided; the UI
        adds `name` to providers.txt via ProviderManager."""
        norm = self._normalize_name(name)
        if not norm:
            return
        with self._lock:
            self.index.setdefault("provider_decisions", {})[norm] = {
                "status": "accepted", "name": name,
            }
            self.rebuild_index()

    def reject_provider_candidate(self, name: str) -> None:
        """Human dismissed a suggested provider."""
        norm = self._normalize_name(name)
        if not norm:
            return
        with self._lock:
            self.index.setdefault("provider_decisions", {})[norm] = {
                "status": "rejected", "name": name,
            }
            self.rebuild_index()

    # ── lookups (FileProcessor.process_file calls these) ────────────────

    def lookup_alias(self, raw_name: str) -> str:
        """The confirmed client for a raw name, or "" if unconfirmed
        (pending, ambiguous, rejected, or never seen)."""
        norm = self._normalize_name(raw_name)
        if not norm:
            return ""
        rec = self.index.get("alias_candidates", {}).get(norm)
        if not rec or rec.get("status") != "confirmed":
            return ""
        client = rec.get("confirmed_client", "")
        # Last line of defence: never hand back a placeholder as a client,
        # however it got into the index.
        return "" if self.is_sentinel(client) else client

    def lookup_claim(self, claim_number: str) -> str:
        """The client for a claim number, when every logged correction for
        that claim number agrees on a single client. "" when the claim is
        unknown or split across more than one client."""
        claim_number = (claim_number or "").strip()
        if not claim_number:
            return ""
        rec = self.index.get("claim_index", {}).get(claim_number, {})
        if len(rec) == 1:
            client = next(iter(rec))
            return "" if self.is_sentinel(client) else client
        return ""

    # ── few-shot retrieval (learning.few_shot_examples) ─────────────────

    def find_similar_corrections(self, text: str, doc_type: str = "", limit: int = 3) -> list:
        """Past corrections whose text_excerpt is most similar to `text`,
        most-similar first. Similarity is a difflib ratio over a
        whitespace-collapsed, lowercased comparison; a matching doc_type
        adds a small bonus. Used for few-shot prompting, but implemented
        and testable independently of that wiring."""
        norm_text = self._normalize_for_similarity(text)
        if not norm_text:
            return []
        scored = []
        # Newest-first, capped — see MAX_FEWSHOT_SCAN. This runs once per
        # document in a batch, so it must not scale with the whole log.
        recent = self._read_all_corrections()[-self.MAX_FEWSHOT_SCAN:]
        for e in recent:
            excerpt = e.get("text_excerpt", "")
            if not excerpt:
                continue
            norm_excerpt = self._normalize_for_similarity(excerpt)
            if not norm_excerpt:
                continue
            ratio = difflib.SequenceMatcher(
                None, norm_text[:2000], norm_excerpt[:2000]
            ).ratio()
            entry_doc_type = (e.get("predicted_doc_type") or e.get("corrected_desc") or "")
            if doc_type and entry_doc_type and doc_type.strip().lower() == entry_doc_type.strip().lower():
                ratio += 0.1
            scored.append((ratio, e))
        scored.sort(key=lambda t: t[0], reverse=True)
        results = []
        for ratio, e in scored[:max(0, limit)]:
            results.append({
                "text_excerpt": e.get("text_excerpt", ""),
                "corrected_client": e.get("corrected_client", ""),
                "corrected_desc": e.get("corrected_desc", ""),
                "doc_type": e.get("predicted_doc_type", ""),
                "score": round(ratio, 4),
            })
        return results

    # ── retroactive rename planning (Pass 7 wires the preview UI) ───────

    # Sentinel unknown-client filename prefixes — a file starting with one
    # of these was named by THE MACHINE, never by a human. Kept in sync
    # with naming.unknown_client_label / naming.no_client_label defaults;
    # callers may pass additional labels via `extra_unknown_labels` when a
    # firm has customized them in Settings.
    _DEFAULT_UNKNOWN_LABELS = ("A-UNKNOWN CLIENT", "A-NEEDS REVIEW")

    def plan_retroactive_renames(self, raw_name: str, client: str, folders: list,
                                  extra_unknown_labels: Optional[list] = None) -> list:
        """Given a newly confirmed alias (raw_name -> client), scan
        `folders` for files that:
          1. the MACHINE named — the filename starts with an unknown-client
             label (never a file a human already named), and
          2. were logged (correction OR observation) with a raw_client
             matching `raw_name`, identified by doc_hash so this doesn't
             depend on the file's current name.

        Returns [{"path", "current_name", "proposed_name"}, ...]. Performs
        NO renames — Pass 7 adds a preview UI with checkboxes and routes
        the actual renames through RenameLog so they're undoable."""
        norm = self._normalize_name(raw_name)
        if not norm or not client:
            return []

        matching_hashes = set()
        for e in self._read_all_corrections():
            if self._normalize_name(e.get("raw_client", "")) == norm:
                h = (e.get("doc_hash") or "").strip()
                if h:
                    matching_hashes.add(h)
        if not matching_hashes:
            return []

        unknown_labels = list(self._DEFAULT_UNKNOWN_LABELS)
        for lbl in (extra_unknown_labels or []):
            if lbl and lbl not in unknown_labels:
                unknown_labels.append(lbl)

        proposals = []
        for folder in folders or []:
            if not folder or not os.path.isdir(folder):
                continue
            try:
                names = os.listdir(folder)
            except Exception as e:
                logging.warning(f"plan_retroactive_renames: could not list {folder}: {e}")
                continue
            for name in names:
                path = os.path.join(folder, name)
                if not os.path.isfile(path):
                    continue
                # SAFETY (non-negotiable): only ever propose a file the
                # machine itself named — never a file a human already named.
                if not any(name.upper().startswith(lbl.upper()) for lbl in unknown_labels):
                    continue
                file_hash = FileProcessor._file_hash(path)
                if not file_hash or file_hash not in matching_hashes:
                    continue
                ext = os.path.splitext(name)[1]
                m = re.match(r"^.+? - (.+)\.[^.]+$", name)
                desc = m.group(1) if m else os.path.splitext(name)[0]
                proposals.append({
                    "path": path,
                    "current_name": name,
                    "proposed_name": f"{client} - {desc}{ext}",
                })
        return proposals


# Module-level singleton — shared by FileProcessor (forward alias/claim
# application during a batch) and APIClient (few-shot examples) so neither
# reloads/rebuilds the index once per file. observations_required is
# re-applied from config on every call in case Settings changed it
# mid-session.
_learning_store_singleton: Optional["LearningStore"] = None
_learning_store_lock = threading.Lock()


def _get_learning_store(config: dict) -> "LearningStore":
    global _learning_store_singleton
    learning_cfg = (config.get("learning", {}) or {})
    naming_cfg = (config.get("naming", {}) or {})
    obs_required = learning_cfg.get("observations_required", 3)
    # The firm's own unknown-client labels count as sentinels too, so a
    # customized placeholder can't be learned as a client either.
    sentinel_labels = [
        naming_cfg.get("unknown_client_label", ""),
        naming_cfg.get("no_client_label", ""),
    ]
    with _learning_store_lock:
        if _learning_store_singleton is None:
            _learning_store_singleton = LearningStore(
                _USER_DATA_DIR, observations_required=obs_required,
                sentinel_labels=sentinel_labels,
            )
        else:
            _learning_store_singleton.observations_required = max(1, int(obs_required or 3))
            _learning_store_singleton.sentinel_labels = sentinel_labels
        return _learning_store_singleton


# ─────────────────────────────────────────────────────────────
# DocumentExtractor
# ─────────────────────────────────────────────────────────────

class DocumentExtractor:
    """
    Extracts content from a document for AI classification.
    Returns an ExtractionResult with content_type="text" or "image".
    """

    IMAGE_RENDER_SCALE = 300 / 72  # ~4.17x — renders at 300 DPI for better OCR accuracy

    @staticmethod
    def extract(file_path: str, max_chars: int = 4000, max_pages: int = 5,
                vision_mode: bool = False, max_vision_pages: int = 2,
                ocr_preprocess: bool = True, skip_fax_cover_pages: bool = False,
                deskew_photos: bool = False) -> ExtractionResult:
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".pdf":
            if vision_mode:
                return DocumentExtractor._from_pdf_vision(
                    file_path, max_vision_pages,
                    skip_fax_cover_pages=skip_fax_cover_pages,
                )
            return DocumentExtractor._from_pdf(file_path, max_chars, max_pages,
                                                ocr_preprocess=ocr_preprocess,
                                                skip_fax_cover_pages=skip_fax_cover_pages,
                                                deskew_photos=deskew_photos)
        elif ext in (".jpg", ".jpeg"):
            return DocumentExtractor._from_jpeg(file_path, deskew_photos=deskew_photos)
        else:
            raise ValueError(f"Unsupported file type: {ext}")

    # ── Fax / transmission cover page detection (reading.skip_fax_cover_pages) ──
    #
    # A fax machine or MFP's own "Send Result Report" / "Transmission
    # Verification" page is frequently scanned in as page 1 of a document,
    # ahead of the actual letter. Left in, its boilerplate gets mixed into
    # whatever text/image is sent to the model and can corrupt the reading
    # of the real document (see PASS 3 item A). Detection below requires at
    # least two independent signals so a real letter that merely mentions
    # "VIA FACSIMILE" or a fax number is never misclassified.

    _TRANSMISSION_PHRASES_RE = re.compile(
        r"send result report|transmission report|transmission verification|"
        r"communication result|activity report|journal report|fax confirmation",
        re.IGNORECASE,
    )
    _TRANSMISSION_VENDOR_RE = re.compile(
        r"KYOCERA|TASKalfa|Canon\s+iR|RICOH|Brother\s+MFC|HP\s+OfficeJet|"
        r"Xerox\s+WorkCentre",
        re.IGNORECASE,
    )
    _TRANSMISSION_TABLE_HEADERS_RE = [
        re.compile(r"Job\s*No", re.IGNORECASE),
        re.compile(r"Total\s*Time", re.IGNORECASE),
        re.compile(r"Resolution\s*/\s*ECM", re.IGNORECASE),
        re.compile(r"\bResult\b", re.IGNORECASE),
        re.compile(r"Destination", re.IGNORECASE),
        re.compile(r"Firmware\s*Version", re.IGNORECASE),
        re.compile(r"Page\(s\)", re.IGNORECASE),
    ]

    @staticmethod
    def _is_transmission_page(page_text: str) -> bool:
        """True if `page_text` looks like a fax machine / MFP transmission
        artifact (a "Send Result Report", activity log, confirmation sheet,
        etc.) rather than real document content.

        Requires at least two independent signals so a real letter that
        merely mentions a fax number or "VIA FACSIMILE" is never flagged:
          - a known report-title phrase ("Send Result Report", "Transmission
            Verification", ...)
          - a vendor/MFP banner (KYOCERA, RICOH, Brother MFC, ...)
          - two or more result-table column headers together (Job No, Total
            Time, Destination, Result, ...) — any single one of these words
            can appear innocently elsewhere, but the combination is
            distinctive of a machine-generated report
          - very low text volume (< 400 chars), but ONLY as a tie-breaker on
            top of one of the above — short text alone proves nothing
        """
        if not page_text or not page_text.strip():
            return False
        text = page_text
        signals = 0
        if DocumentExtractor._TRANSMISSION_PHRASES_RE.search(text):
            signals += 1
        if DocumentExtractor._TRANSMISSION_VENDOR_RE.search(text):
            signals += 1
        header_hits = sum(
            1 for pat in DocumentExtractor._TRANSMISSION_TABLE_HEADERS_RE
            if pat.search(text)
        )
        if header_hits >= 2:
            signals += 1
        if signals >= 1 and len(text.strip()) < 400:
            signals += 1
        return signals >= 2

    @staticmethod
    def _drop_transmission_pages(file_path: str, indices: list, texts: list) -> list:
        """Given parallel lists of original page indices and their text,
        return the subset of indices that are NOT transmission artifacts —
        logging each drop. Safety net: if dropping would remove every page,
        the original list is returned unchanged (with a log line) rather
        than ever handing the model zero pages of content."""
        kept = [
            idx for idx, txt in zip(indices, texts)
            if not DocumentExtractor._is_transmission_page(txt)
        ]
        if not kept:
            logging.info(
                f"{os.path.basename(file_path)}: fax/transmission drop would "
                "remove all pages — keeping originals"
            )
            return indices
        for idx in indices:
            if idx not in kept:
                logging.info(
                    f"{os.path.basename(file_path)}: dropped page {idx + 1} "
                    "as fax/transmission cover"
                )
        return kept

    @staticmethod
    def _from_pdf_vision(file_path: str, max_vision_pages: int,
                          skip_fax_cover_pages: bool = False) -> ExtractionResult:
        """Render the first N pages of a PDF as PNG images and return them
        as a base64 image list for the vision model. Bypasses OCR entirely."""
        if fitz is None:
            raise ImportError("PyMuPDF is not installed. Run: pip install PyMuPDF")
        doc = fitz.open(file_path)
        candidate_indices = list(range(doc.page_count))

        if skip_fax_cover_pages:
            # Quick text probe per page — cheap even for a scanned/image-only
            # PDF (returns "" there and _is_transmission_page just says No)
            # — so a fax cover isn't rendered and doesn't eat into the
            # vision model's limited page budget.
            probe_texts = []
            for i in candidate_indices:
                try:
                    probe_texts.append(doc[i].get_text("text"))
                except Exception:
                    probe_texts.append("")
            candidate_indices = DocumentExtractor._drop_transmission_pages(
                file_path, candidate_indices, probe_texts
            )

        page_indices = candidate_indices[:max(1, max_vision_pages)]
        scale = DocumentExtractor.IMAGE_RENDER_SCALE
        mat = fitz.Matrix(scale, scale)
        images_b64: List[str] = []
        for i in page_indices:
            pix = doc[i].get_pixmap(matrix=mat)
            images_b64.append(base64.b64encode(pix.tobytes("png")).decode("utf-8"))
        doc.close()
        logging.info(
            f"{os.path.basename(file_path)}: vision mode — sending "
            f"{len(images_b64)} page(s) to model"
        )
        return ExtractionResult(
            content_type="image",
            content=images_b64[0] if images_b64 else "",
            mime_type="image/png",
            method="vision",
            images=images_b64,
        )

    # Labels that reliably identify the client when they appear in document text
    _CLIENT_LABEL_RE = re.compile(
        r"(?:Client(?:\s*/\s*Patient)?(?:\s+Name)?|Claimant|Clmt|"
        r"Injured(?:\s+(?:Worker|Party))?|Patient(?:\s+Name)?|"
        r"Insured(?:\s+Name)?|Named\s+Insured|Employee|"
        r"Your\s+Client|RE|Re|Regarding|Subject)\s*[:\-\.]\s*"
        r"([A-Za-z][A-Za-z ,\-\'\.]{1,60})",
        re.IGNORECASE,
    )

    # Words that signal the end of a name (stop words for phrase truncation)
    _NAME_STOP_RE = re.compile(
        r"\b(?:Law|Office|Attorney|Atty|DOB|Date|vs?\.?|and|LLC|Inc|"
        r"Insurance|Clinic|Hospital|Medical|Center|Corp|PC|LLP|PA)\b",
        re.IGNORECASE,
    )

    @staticmethod
    def _extract_labeled_snippets(page_text: str) -> list:
        """Extract name phrases following client-identifying labels on a page.
        Truncates each phrase at stop words so surrounding text doesn't bleed in."""
        snippets = []
        for m in DocumentExtractor._CLIENT_LABEL_RE.finditer(page_text):
            phrase = m.group(1).strip()
            # Truncate at stop words
            stop = DocumentExtractor._NAME_STOP_RE.search(phrase)
            if stop:
                phrase = phrase[:stop.start()].strip()
            phrase = phrase.strip(".,;: ")
            if len(phrase) >= 3:
                snippets.append(phrase)
        return snippets

    @staticmethod
    def _from_pdf(file_path: str, max_chars: int, max_pages: int = 5,
                  ocr_preprocess: bool = True, skip_fax_cover_pages: bool = False,
                  deskew_photos: bool = False) -> ExtractionResult:
        if fitz is None:
            raise ImportError("PyMuPDF is not installed. Run: pip install PyMuPDF")
        doc = fitz.open(file_path)
        page_limit = min(doc.page_count, max_pages)
        if doc.page_count > max_pages:
            logging.info(
                f"{os.path.basename(file_path)}: {doc.page_count} pages — "
                f"capping at {max_pages}"
            )

        # ── Pass 1: collect per-page text via PyMuPDF ─────────────────────
        page_texts = []
        all_native_text = ""
        for page in doc.pages(0, page_limit):
            pt = page.get_text("text").strip()
            page_texts.append(pt)
            all_native_text += pt

        has_native_text = len(all_native_text.strip()) >= 50

        if has_native_text:
            # ── Pass 1b: drop fax/transmission cover pages before anything
            # downstream (labeling, char budget) ever sees them — dropping
            # AFTER classification would be too late, the cover text has
            # already had a chance to corrupt what gets sent to the model.
            page_indices = list(range(len(page_texts)))
            if skip_fax_cover_pages:
                page_indices = DocumentExtractor._drop_transmission_pages(
                    file_path, page_indices, page_texts
                )

            # ── Pass 2: find pages containing client-identifying labels ────
            labeled_pages = []
            unlabeled_pages = []
            for i in page_indices:
                pt = page_texts[i]
                snippets = DocumentExtractor._extract_labeled_snippets(pt)
                if snippets:
                    labeled_pages.append((i, pt, snippets))
                else:
                    unlabeled_pages.append((i, pt))

            if labeled_pages:
                # Build content: labeled pages first (with snippet annotation),
                # then remaining pages up to the char budget.
                parts = []
                for i, pt, snippets in labeled_pages:
                    parts.append(f"[Page {i+1} — labeled fields: {'; '.join(snippets)}]\n{pt}")
                for i, pt in unlabeled_pages:
                    parts.append(f"[Page {i+1}]\n{pt}")
                raw_text = "\n\n".join(parts).strip()[:max_chars]
                logging.info(
                    f"{os.path.basename(file_path)}: labeled pages found "
                    f"({[i+1 for i,_,_ in labeled_pages]}), sending those first"
                )
            else:
                # No labeled pages — send remaining pages sequentially
                raw_text = "".join(page_texts[i] for i in page_indices).strip()[:max_chars]

            doc.close()
            return ExtractionResult(content_type="text", content=raw_text, method="pymupdf")

        # ── Image-only PDF: OCR every page up to limit, prioritize labeled ──
        doc.close()
        ocr_results = []  # (index, text) for pages that OCR'd successfully
        for i in range(page_limit):
            ocr_text = DocumentExtractor._ocr_pdf_page(file_path, page_index=i,
                                                        max_chars=max_chars,
                                                        preprocess=ocr_preprocess,
                                                        deskew_photos=deskew_photos)
            if ocr_text:
                ocr_results.append((i, ocr_text))

        if skip_fax_cover_pages and ocr_results:
            ocr_indices = [i for i, _ in ocr_results]
            ocr_texts = [t for _, t in ocr_results]
            kept = set(DocumentExtractor._drop_transmission_pages(
                file_path, ocr_indices, ocr_texts
            ))
            ocr_results = [(i, t) for i, t in ocr_results if i in kept]

        ocr_labeled = []
        ocr_unlabeled = []
        for i, ocr_text in ocr_results:
            snippets = DocumentExtractor._extract_labeled_snippets(ocr_text)
            if snippets:
                ocr_labeled.append((i, ocr_text, snippets))
            else:
                ocr_unlabeled.append((i, ocr_text))

        if ocr_labeled or ocr_unlabeled:
            parts = []
            for i, txt, snippets in ocr_labeled:
                parts.append(f"[Page {i+1} — labeled fields: {'; '.join(snippets)}]\n{txt}")
            for i, txt in ocr_unlabeled:
                parts.append(f"[Page {i+1}]\n{txt}")
            combined = "\n\n".join(parts).strip()[:max_chars]
            logging.info(
                f"OCR succeeded on {file_path} ({len(combined)} chars, "
                f"labeled pages: {[i+1 for i,_,_ in ocr_labeled]})"
            )
            return ExtractionResult(content_type="text", content=combined, method="tesseract")

        # Last resort: send page 1 as image to vision model
        return DocumentExtractor._render_pdf_page(file_path)

    @staticmethod
    def _ocr_pdf_page(file_path: str, page_index: int, max_chars: int,
                      preprocess: bool = True, deskew_photos: bool = False) -> str:
        """Run Tesseract OCR on a single PDF page rendered to an image.
        Returns extracted text if >= 50 characters were found, otherwise empty string.
        Returns empty string silently on any error so the caller can fall back gracefully."""
        if pytesseract is None or PILImage is None:
            return ""
        try:
            doc = fitz.open(file_path)
            page_index = min(page_index, doc.page_count - 1)
            pix = doc[page_index].get_pixmap(
                matrix=fitz.Matrix(DocumentExtractor.IMAGE_RENDER_SCALE,
                                   DocumentExtractor.IMAGE_RENDER_SCALE)
            )
            doc.close()
            import io
            pil = PILImage.open(io.BytesIO(pix.tobytes("png")))
            if preprocess:
                if deskew_photos:
                    pil = DocumentExtractor._prepare_photo(pil)
                else:
                    pil = DocumentExtractor._preprocess_for_ocr(pil)
            text = pytesseract.image_to_string(pil).strip()
            return text[:max_chars] if len(text) >= 50 else ""
        except Exception as e:
            logging.warning(f"OCR failed on {file_path} page {page_index}: {e}")
            return ""

    @staticmethod
    def _preprocess_for_ocr(img):
        """Upscale, contrast-normalize, and binarize a page image so Tesseract
        has the cleanest possible input. Falls back to the original image on
        any error — preprocessing must never break OCR."""
        try:
            from PIL import ImageOps
            g = img.convert("L")
            w, h = g.size
            # Tesseract is trained on ~300 DPI text; upscale small scans.
            target = 2000
            if max(w, h) < target:
                scale = target / max(w, h)
                g = g.resize((int(w * scale), int(h * scale)), PILImage.LANCZOS)
            g = ImageOps.autocontrast(g, cutoff=2)
            # Otsu threshold — picks the split that best separates ink from paper.
            hist = g.histogram()[:256]
            total = sum(hist)
            if total == 0:
                return g
            sum_total = sum(i * hist[i] for i in range(256))
            sum_b = 0.0
            w_b = 0
            var_max = 0.0
            threshold = 127
            for t in range(256):
                w_b += hist[t]
                if w_b == 0:
                    continue
                w_f = total - w_b
                if w_f == 0:
                    break
                sum_b += t * hist[t]
                m_b = sum_b / w_b
                m_f = (sum_total - sum_b) / w_f
                var_between = w_b * w_f * (m_b - m_f) ** 2
                if var_between > var_max:
                    var_max = var_between
                    threshold = t
            return g.point(lambda p: 255 if p > threshold else 0, mode="1")
        except Exception as e:
            logging.warning(f"OCR preprocessing failed, using raw image: {e}")
            return img

    # One-time flag so a whole batch running without OpenCV logs the
    # "falling back" notice once instead of once per page/photo.
    _cv2_unavailable_logged = False

    @staticmethod
    def _prepare_photo(img):
        """Perspective-correct, deskew, and adaptively threshold a phone
        photo of a page (reading.deskew_photos). Used in place of
        `_preprocess_for_ocr`'s global-Otsu pass, which is actively harmful
        on a phone photo: paper on a dark/textured background, skewed, with
        uneven lighting pushes a single whole-frame threshold to the wrong
        split and destroys the page before OCR ever sees it (see PASS 3
        item B).

        Pipeline, OpenCV/numpy available:
          1. Border check — sample a strip around the frame edge and see how
             much of it is dark/non-white. A clean scan's border is page-
             white; if this frame doesn't look like "paper on a background",
             leave it to the existing `_preprocess_for_ocr` unchanged.
          2. Page detection — Canny edges + contours; take the largest
             convex 4-point quadrilateral covering 15%-95% of the frame.
          3. If found, perspective-correct it (getPerspectiveTransform /
             warpPerspective) to a deskewed rectangle.
          4. If no quadrilateral is found, fall back to deskew-only:
             minAreaRect over thresholded text pixels, rotating by the
             detected angle when it's a plausible small skew (< 20 degrees).
          5. Adaptive thresholding (Gaussian, generous block size) instead
             of a single global Otsu split, so shadows/uneven lighting
             across the photo don't wipe out text in the darker half.

        Returns a PIL Image in every case. Safe by construction:
          - cv2/numpy missing → logs once at INFO, falls back to
            `_preprocess_for_ocr` (today's behavior) unchanged.
          - any exception anywhere in the OpenCV pipeline → logs a warning
            and returns the ORIGINAL image untouched — page prep must never
            break OCR.
        """
        if not _CV2_AVAILABLE:
            if not DocumentExtractor._cv2_unavailable_logged:
                logging.info(
                    "reading.deskew_photos is on but OpenCV/numpy are not "
                    "installed — falling back to standard OCR preprocessing."
                )
                DocumentExtractor._cv2_unavailable_logged = True
            return DocumentExtractor._preprocess_for_ocr(img)

        try:
            arr = np.array(img.convert("RGB"))
            gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
            h, w = gray.shape[:2]

            # 1. Border check: does this look like a photo of a page sitting
            # on a background, rather than a clean flatbed/ADF scan?
            border = max(2, min(h, w) // 40)
            border_pixels = np.concatenate([
                gray[:border, :].ravel(),
                gray[-border:, :].ravel(),
                gray[:, :border].ravel(),
                gray[:, -border:].ravel(),
            ])
            dark_fraction = float(np.mean(border_pixels < 180)) if border_pixels.size else 0.0
            if dark_fraction < 0.15:
                # Border reads as page-white — treat as a normal clean scan
                # and don't touch it.
                return DocumentExtractor._preprocess_for_ocr(img)

            # 2. Page detection via contours.
            blurred = cv2.GaussianBlur(gray, (5, 5), 0)
            edges = cv2.Canny(blurred, 50, 150)
            edges = cv2.dilate(edges, np.ones((5, 5), np.uint8), iterations=1)
            contours, _ = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)

            frame_area = float(h * w)
            quad = None
            best_area = 0.0
            for c in contours:
                peri = cv2.arcLength(c, True)
                approx = cv2.approxPolyDP(c, 0.02 * peri, True)
                if len(approx) == 4 and cv2.isContourConvex(approx):
                    area = cv2.contourArea(approx)
                    frac = area / frame_area if frame_area else 0.0
                    if 0.15 <= frac <= 0.95 and area > best_area:
                        best_area = area
                        quad = approx.reshape(4, 2).astype("float32")

            # 3. Perspective-correct if a page quad was found.
            if quad is not None:
                working = DocumentExtractor._warp_quad(arr, quad)
            else:
                # 4. No quad — deskew-only via minAreaRect on text pixels.
                working = arr
                _, thresh = cv2.threshold(
                    gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU
                )
                coords = cv2.findNonZero(thresh)
                if coords is not None:
                    rect = cv2.minAreaRect(coords)
                    angle = rect[-1]
                    # cv2.minAreaRect's angle convention wraps at -90; fold
                    # into [-45, 45] so "small skew" means what it says.
                    if angle < -45:
                        angle = 90 + angle
                    if abs(angle) < 20:
                        center = (w / 2, h / 2)
                        rot_m = cv2.getRotationMatrix2D(center, angle, 1.0)
                        working = cv2.warpAffine(
                            arr, rot_m, (w, h), flags=cv2.INTER_CUBIC,
                            borderMode=cv2.BORDER_REPLICATE,
                        )

            # 5. Adaptive threshold — robust to uneven lighting/shadows,
            # unlike a single global Otsu split over the whole frame.
            working_gray = (
                cv2.cvtColor(working, cv2.COLOR_RGB2GRAY) if working.ndim == 3 else working
            )
            block_size = max(15, (min(working_gray.shape[:2]) // 8) | 1)  # odd, generous
            adaptive = cv2.adaptiveThreshold(
                working_gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY, block_size, 15,
            )
            return PILImage.fromarray(adaptive)
        except Exception as e:
            logging.warning(f"Photo page preparation failed, using original image: {e}")
            return img

    @staticmethod
    def _warp_quad(arr, quad):
        """Perspective-warp `arr` so the 4-point quad (any order) becomes a
        flat, deskewed rectangle sized from the quad's own edge lengths."""
        tl, tr, br, bl = DocumentExtractor._order_quad_points(quad)
        width_a = np.linalg.norm(br - bl)
        width_b = np.linalg.norm(tr - tl)
        max_width = max(int(width_a), int(width_b), 1)
        height_a = np.linalg.norm(tr - br)
        height_b = np.linalg.norm(tl - bl)
        max_height = max(int(height_a), int(height_b), 1)
        dst = np.array([
            [0, 0],
            [max_width - 1, 0],
            [max_width - 1, max_height - 1],
            [0, max_height - 1],
        ], dtype="float32")
        src = np.array([tl, tr, br, bl], dtype="float32")
        m = cv2.getPerspectiveTransform(src, dst)
        return cv2.warpPerspective(arr, m, (max_width, max_height))

    @staticmethod
    def _order_quad_points(pts):
        """Order 4 arbitrary points as (top-left, top-right, bottom-right,
        bottom-left) by summed/differenced coordinates."""
        s = pts.sum(axis=1)
        diff = np.diff(pts, axis=1).ravel()
        tl = pts[np.argmin(s)]
        br = pts[np.argmax(s)]
        tr = pts[np.argmin(diff)]
        bl = pts[np.argmax(diff)]
        return tl, tr, br, bl

    @staticmethod
    def _render_pdf_page(file_path: str) -> ExtractionResult:
        if fitz is None:
            raise ImportError("PyMuPDF is not installed.")
        doc = fitz.open(file_path)
        page = doc[0]
        scale = DocumentExtractor.IMAGE_RENDER_SCALE
        mat = fitz.Matrix(scale, scale)
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        doc.close()
        b64 = base64.b64encode(img_bytes).decode("utf-8")
        return ExtractionResult(content_type="image", content=b64, mime_type="image/png", method="vision")

    @staticmethod
    def _from_jpeg(file_path: str, deskew_photos: bool = False) -> ExtractionResult:
        # JPEGs (phone photos) always go to the model as an image — there's
        # no OCR path for them. When reading.deskew_photos is on, run the
        # same perspective-correct/deskew prep used for photographed PDF
        # pages before handing the image over, so a skewed photo on a dark
        # background gets straightened out for whichever model reads it.
        # Any failure anywhere in this path falls back to the untouched
        # original bytes — page prep must never break extraction.
        if deskew_photos and PILImage is not None:
            try:
                img = PILImage.open(file_path)
                img.load()
                prepared = DocumentExtractor._prepare_photo(img)
                import io
                buf = io.BytesIO()
                prepared.convert("RGB").save(buf, format="PNG")
                b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
                return ExtractionResult(content_type="image", content=b64,
                                        mime_type="image/png", method="vision")
            except Exception as e:
                logging.warning(f"Photo prep failed for {file_path}, using original image: {e}")

        with open(file_path, "rb") as f:
            img_bytes = f.read()
        b64 = base64.b64encode(img_bytes).decode("utf-8")
        return ExtractionResult(content_type="image", content=b64, mime_type="image/jpeg", method="vision")

    # ── Claim number / date-of-injury / DOB extraction (reading.extract_claim_numbers) ──

    _CLAIM_LABEL_RE = re.compile(
        r"(?:Claim\s*(?:Number|No\.?|#)|Claim\s*ID|File\s*(?:Number|No\.?|#)|"
        r"WCS?\s*(?:Claim|No))\s*[:\-]?\s*"
        r"([A-Za-z0-9][A-Za-z0-9 \-]{3,40}?)(?=\s{2,}|[\n\r]|$)",
        re.IGNORECASE,
    )
    # Reject a captured claim value that's actually a date (M/D/YY, MM-DD-YYYY, ...)
    _CLAIM_LOOKS_LIKE_DATE_RE = re.compile(r"^\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}$")
    # ...or a phone number (with or without a hyphen/space separator).
    _CLAIM_LOOKS_LIKE_PHONE_RE = re.compile(r"^\(?\d{3}\)?[\-\s]?\d{3}[\-\s]?\d{4}$")

    _DATE_VALUE_PATTERN = (
        r"(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}|[A-Za-z]{3,9}\.?\s+\d{1,2},?\s+\d{4})"
    )
    _DOI_LABEL_RE = re.compile(
        r"(?:Date\s+of\s+Injury|D\.?\s*O\.?\s*I\.?|Injury\s+Date|Accident\s+Date|"
        r"Date\s+of\s+Accident)\s*[:\-]?\s*" + _DATE_VALUE_PATTERN,
        re.IGNORECASE,
    )
    _DOB_LABEL_RE = re.compile(
        r"(?:D\.?\s*O\.?\s*B\.?|Date\s+of\s+Birth|Birth\s+Date)\s*[:\-]?\s*"
        + _DATE_VALUE_PATTERN,
        re.IGNORECASE,
    )

    @staticmethod
    def _clean_claim_value(raw: str) -> str:
        """Trim trailing punctuation and collapse internal spaces — a claim
        number OCR'd/typed as "1E01 E018 695852" normalizes to
        "1E01E018695852". Dashes are left in place (some claim numbers use
        them structurally, e.g. "501-216-260260334")."""
        val = raw.strip().strip(".,;:")
        return val.replace(" ", "")

    @staticmethod
    def _normalize_date(raw: str) -> str:
        """Normalize a matched date string to YYYY-MM-DD when it can be
        parsed confidently; otherwise return the raw matched text unchanged
        rather than guess. Two-digit years: 00-40 -> 20xx, 41-99 -> 19xx."""
        raw = raw.strip().rstrip(",")
        m = re.match(r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})$", raw)
        if m:
            mo_s, da_s, yr_s = m.groups()
            mo, da = int(mo_s), int(da_s)
            if len(yr_s) == 2:
                yr_i = int(yr_s)
                yr = 2000 + yr_i if yr_i <= 40 else 1900 + yr_i
            else:
                yr = int(yr_s)
            try:
                return datetime.date(yr, mo, da).strftime("%Y-%m-%d")
            except ValueError:
                return raw
        m = re.match(r"^([A-Za-z]{3,9})\.?\s+(\d{1,2}),?\s+(\d{4})$", raw)
        if m:
            month_name, day_s, year_s = m.groups()
            for fmt in ("%B", "%b"):
                try:
                    parsed_month = datetime.datetime.strptime(month_name.title(), fmt).month
                    return datetime.date(int(year_s), parsed_month, int(day_s)).strftime("%Y-%m-%d")
                except ValueError:
                    continue
            return raw
        return raw

    @staticmethod
    def extract_identifiers(text: str) -> dict:
        """Best-effort extraction of a claim number, date of injury, and
        date of birth from document text, via labeled regex matches
        (see PASS 3 item C). Returns
        {"claim_number": str, "date_of_injury": str, "dob": str} with empty
        strings for anything not confidently found.

        Deliberately conservative: a wrong claim number is worse than none.
        A claim-number candidate is only accepted when it's 6-30 characters
        after normalizing (stripping internal spaces) and doesn't look like
        a date or a phone number — both of which sit right next to claim
        number labels on real intake forms and are easy to mis-capture.
        Dates are normalized to YYYY-MM-DD only when confidently parseable;
        otherwise the raw matched text is kept rather than guessed at.
        """
        result = {"claim_number": "", "date_of_injury": "", "dob": ""}
        if not text:
            return result

        for m in DocumentExtractor._CLAIM_LABEL_RE.finditer(text):
            candidate = DocumentExtractor._clean_claim_value(m.group(1))
            if not (6 <= len(candidate) <= 30):
                continue
            if DocumentExtractor._CLAIM_LOOKS_LIKE_DATE_RE.match(candidate):
                continue
            if DocumentExtractor._CLAIM_LOOKS_LIKE_PHONE_RE.match(candidate):
                continue
            if not re.search(r"[A-Za-z0-9]", candidate):
                continue
            result["claim_number"] = candidate
            break

        m = DocumentExtractor._DOI_LABEL_RE.search(text)
        if m:
            result["date_of_injury"] = DocumentExtractor._normalize_date(m.group(1))

        m = DocumentExtractor._DOB_LABEL_RE.search(text)
        if m:
            result["dob"] = DocumentExtractor._normalize_date(m.group(1))

        return result

    # ── OCR -> vision escalation ladder (reading.vision_escalation) ──

    _WORD_TOKEN_RE = re.compile(r"^[A-Za-z]{2,}$")
    _VOWEL_RE = re.compile(r"[aeiouAEIOU]")

    @staticmethod
    def assess_text_quality(text: str) -> float:
        """Heuristic 0.0-1.0 score for whether `text` looks like real,
        usable document text rather than OCR noise (see PASS 3 item D).
        Used to decide whether reading.vision_escalation should re-read a
        document with the vision model instead of trusting a bad OCR pass.

        Combines, with no single factor able to carry the score on its own:
          - length: very short text can't be classified reliably regardless
            of how clean it looks (scaled up to a 300-char cap)
          - "real word" ratio: the fraction of whitespace-split tokens that
            are 2+ letters and contain a vowel. Typical OCR garbage like
            "1|11 ,, ~~ lll ]{ " produces almost no such tokens, while a
            clean sentence produces close to 100%.
          - alphabetic ratio: fraction of all characters that are letters.
            Numbers/punctuation/whitespace are normal in real documents and
            aren't penalized alone, but combined with a low word ratio it's
            a garbage signal.
          - a flat bonus if a client-identifying label (_CLIENT_LABEL_RE)
            was found — a strong sign the OCR captured something meaningful.

        Not a proof of correctness, just cheap enough to run on every
        document.
        """
        if not text or not text.strip():
            return 0.0
        stripped = text.strip()

        length_score = min(1.0, len(stripped) / 300.0)

        tokens = stripped.split()
        if not tokens:
            return 0.0
        word_like = 0
        for t in tokens:
            core = t.strip(".,;:()[]{}\"'-")
            if DocumentExtractor._WORD_TOKEN_RE.match(core) and DocumentExtractor._VOWEL_RE.search(core):
                word_like += 1
        word_ratio = word_like / len(tokens)

        alpha_chars = sum(1 for c in stripped if c.isalpha())
        alpha_ratio = alpha_chars / len(stripped)

        label_bonus = 0.15 if DocumentExtractor._CLIENT_LABEL_RE.search(stripped) else 0.0

        score = (0.25 * length_score) + (0.45 * word_ratio) + (0.25 * alpha_ratio) + label_bonus
        return max(0.0, min(1.0, score))


# ─────────────────────────────────────────────────────────────
# APIClient
# ─────────────────────────────────────────────────────────────

class APIClient:

    # JSON schema handed to Ollama's `format` field (Ollama accepts a
    # JSON-schema object there, not just the string "json") when
    # classification.structured_output is on. Keeps malformed responses
    # rare for the richer shape. The OpenWebUI/OpenAI-compatible path
    # deliberately does NOT get an equivalent response_format — not every
    # local gateway accepts that field, so the payload shape there is
    # unchanged from before this pass.
    _STRUCTURED_SCHEMA = {
        "type": "object",
        "properties": {
            "client": {"type": "string"},
            "doc_type": {"type": "string"},
            "recipient": {"type": "string"},
            "direction": {"type": "string"},
            "doc_date": {"type": "string"},
            "confidence": {"type": "string"},
        },
        "required": ["client", "doc_type", "confidence"],
    }

    @staticmethod
    def _build_prompt(extraction: ExtractionResult, config: Optional[dict] = None,
                       document_types: Optional[list] = None,
                       candidates: Optional[list] = None,
                       few_shot_examples: Optional[list] = None) -> str:
        """Build the classification prompt.

        With classification.structured_output off (the default, and the
        behavior when `config` is omitted) this returns the ORIGINAL prompt
        byte-for-byte — see _build_legacy_prompt. On, it returns a richer
        prompt asking for doc_type/recipient/direction/doc_date alongside
        client — see _build_structured_prompt.

        `few_shot_examples` (learning.few_shot_examples, off by default) is
        only ever honored on the structured prompt — see
        LearningStore.find_similar_corrections.
        """
        cls_cfg = (config or {}).get("classification", {})
        if not cls_cfg.get("structured_output", False):
            return APIClient._build_legacy_prompt(extraction)
        return APIClient._build_structured_prompt(
            extraction, cls_cfg, document_types, candidates, few_shot_examples
        )

    @staticmethod
    def _build_legacy_prompt(extraction: ExtractionResult) -> str:
        """The prompt as it existed before PASS 4. Kept verbatim — including
        the "Incoming Document" escape hatch this pass otherwise removes —
        so classification.structured_output=False reproduces today's exact
        behavior. AI extracts the client name freely from the document; no
        client list in the prompt. fuzzy_match (in FileProcessor) maps the
        raw name to the authoritative list entry."""
        if extraction.content_type == "text":
            doc_section = f"Document text:\n{extraction.content}"
        else:
            doc_section = "[See attached image]"

        return (
            "You are a document classifier for a law firm.\n"
            "Your job is to identify which client this document belongs to "
            "and write a short description of what it is.\n\n"
            f"{doc_section}\n\n"
            "RULES — read carefully before responding:\n\n"
            "CLIENT IDENTIFICATION:\n"
            "- Scan the document for labels that introduce the client's name, "
            "in this priority order:\n"
            "    1. 'Client:', 'Client Name:', 'Client/Patient Name:'\n"
            "    2. 'Claimant:', 'Injured:', 'Injured Party:', 'Injured Worker:'\n"
            "    3. 'Patient:', 'Patient Name:'\n"
            "    4. 'Insured:', 'Insured Name:', 'Named Insured:'\n"
            "    5. 'Employee:' (workers comp)\n"
            "    6. 'RE:', 'Re:', 'Regarding:', 'Subject:'\n"
            "    7. Case captions (plaintiff or defendant the firm represents)\n"
            "- Do NOT identify a business, facility, clinic, hospital, insurance "
            "company, opposing party, or attorney as the client.\n"
            "- Return the client's full name exactly as it appears in the document.\n"
            "- If you cannot clearly identify the client, return NEEDS_REVIEW.\n\n"
            "DESCRIPTION:\n"
            "- 2-5 separate words with spaces between them, title case, no special characters "
            "(e.g. \"Retainer Agreement\", \"Motion to Dismiss\", \"Invoice\").\n"
            "- IMPORTANT: Always use spaces between words. Never combine words "
            "(e.g. write \"Medical Record Request\" NOT \"Medicalrecordrequest\").\n"
            "- NEVER describe a fax cover sheet or fax wrapper — describe the actual "
            "document content. If no real content is visible, use \"Incoming Document\".\n\n"
            "Return ONLY valid JSON with no extra text:\n"
            '{"client": "LAST, First", "desc": "Retainer Agreement", "confidence": "high|medium|low"}'
        )

    @staticmethod
    def _build_structured_prompt(extraction: ExtractionResult, cls_cfg: dict,
                                  document_types: Optional[list],
                                  candidates: Optional[list],
                                  few_shot_examples: Optional[list] = None) -> str:
        """Structured-output prompt (classification.structured_output=True).

        Differs from the legacy prompt in four deliberate ways, each fixing
        a specific production failure:
          - Prefers the document's own printed title/subject/form name over
            a paraphrased summary (fixes "Reduction Request" being renamed
            to a made-up "Compromise Offer Letter Medical Bill").
          - Removes the "Incoming Document" escape hatch entirely — the
            model must always attempt a real description.
          - Optionally includes the firm's document-type vocabulary so the
            model can return a canonical, collision-resistant type name.
          - Optionally asks for the recipient/sender organization — the
            business a business-can't-be-client rule used to train the
            model away from ever naming, even though the office wants it
            in filenames like "Reduction Request to Chiropractic Works".
        """
        if extraction.content_type == "text":
            doc_section = f"Document text:\n{extraction.content}"
        else:
            doc_section = "[See attached image]"

        lines = [
            "You are a document classifier for a law firm.",
            "Your job is to identify which client this document belongs to, "
            "what kind of document it is, and a few other structured facts "
            "about it.",
            "",
            doc_section,
            "",
            "RULES — read carefully before responding:",
            "",
            "CLIENT IDENTIFICATION:",
            "- Scan the document for labels that introduce the client's name, "
            "in this priority order:",
            "    1. 'Client:', 'Client Name:', 'Client/Patient Name:'",
            "    2. 'Claimant:', 'Injured:', 'Injured Party:', 'Injured Worker:'",
            "    3. 'Patient:', 'Patient Name:'",
            "    4. 'Insured:', 'Insured Name:', 'Named Insured:'",
            "    5. 'Employee:' (workers comp)",
            "    6. 'RE:', 'Re:', 'Regarding:', 'Subject:'",
            "    7. Case captions (plaintiff or defendant the firm represents)",
            "- A business, facility, clinic, hospital, insurance company, "
            "opposing party, or attorney is NEVER the CLIENT. That said, one "
            "of those IS the expected value for RECIPIENT below — being "
            "ineligible as the client does not mean ignore it.",
            "- Return the client's full name exactly as it appears in the document.",
            "- If you cannot clearly identify the client, return NEEDS_REVIEW.",
        ]

        if candidates:
            lines.append("")
            lines.append(
                "LIKELY CANDIDATES — the client is probably, but not "
                "certainly, one of these. Still return NEEDS_REVIEW if none "
                "of them genuinely fit:"
            )
            lines.append(", ".join(candidates))

        lines.append("")
        lines.append("DOCUMENT TYPE (doc_type):")
        lines.append(
            "- Prefer the document's own printed title. If the document "
            "displays its own title, subject line, or form name (e.g. "
            "'Reduction Request', \"PHYSICIAN'S AND CHIROPRACTOR'S PROGRESS "
            "REPORT\"), use that. Do not invent a summary of the contents."
        )
        if document_types:
            lines.append(
                "- Here is a list of document types this firm commonly "
                "sees. If the document plainly matches one of them, return "
                "that exact name. Otherwise, return the document's own "
                "printed title verbatim:"
            )
            lines.append(", ".join(document_types))
        lines.append(
            "- Always give your best description of what the document IS, "
            "even if you cannot identify the client. Never answer "
            "\"Incoming Document\" or \"Unknown Document\"."
        )

        lines.append("")
        lines.append("DIRECTION (direction):")
        lines.append(
            "- \"outgoing\" if the document is on the firm's own letterhead "
            "(\"Law Offices of Greg D. Jensen\" / \"GREG D. JENSEN\"). "
            "Otherwise \"incoming\"."
        )

        if cls_cfg.get("extract_recipient", False):
            lines.append("")
            lines.append("RECIPIENT (recipient):")
            lines.append(
                "- For an outgoing document, give the addressee "
                "organization — the \"to\" of the letter (e.g. the clinic, "
                "hospital, or insurer it was sent to)."
            )
            lines.append(
                "- For an incoming document, give the sender organization "
                "instead."
            )
            lines.append(
                "- This is exactly the kind of business/facility/clinic/"
                "hospital/insurance-company value that is never a valid "
                "CLIENT — naming it here is expected and wanted."
            )
            lines.append("- Leave it empty if no organization is identifiable.")

        lines.append("")
        lines.append("DATE (doc_date):")
        lines.append(
            "- The date printed on the document itself (letter date, date "
            "of exam, date of visit) in YYYY-MM-DD format. Leave empty if "
            "unclear."
        )

        # learning.few_shot_examples (off by default) — worked examples
        # pulled from LearningStore.find_similar_corrections for documents
        # that read similarly to this one, so the model can see how staff
        # actually resolved comparable cases.
        if few_shot_examples:
            lines.append("")
            lines.append(
                "WORKED EXAMPLES — similar documents corrected by staff "
                "previously (for reference only; base your answer on the "
                "document above, not on these):"
            )
            for ex in few_shot_examples:
                excerpt = (ex.get("text_excerpt") or "")[:200].replace("\n", " ").strip()
                lines.append(
                    f"- Excerpt: \"{excerpt}...\" -> "
                    f"client: \"{ex.get('corrected_client', '')}\", "
                    f"doc_type: \"{ex.get('corrected_desc', '')}\""
                )

        lines.append("")
        lines.append("Return ONLY valid JSON with no extra text, in exactly this shape:")
        lines.append(
            '{"client": "LAST, First", "doc_type": "Reduction Request", '
            '"recipient": "Chiropractic Works", "direction": "outgoing", '
            '"doc_date": "2026-08-10", "confidence": "high|medium|low"}'
        )
        return "\n".join(lines)

    @staticmethod
    def classify(extraction: ExtractionResult, client_list: list, config: dict,
                 document_types: Optional[list] = None,
                 candidates: Optional[list] = None) -> dict:
        # No client list in the prompt — AI extracts the name freely from the document.
        # fuzzy_match (called in FileProcessor) maps the raw name to the authoritative list.
        cls_cfg = config.get("classification", {})
        structured = cls_cfg.get("structured_output", False)

        # learning.few_shot_examples (off by default): only meaningful on
        # the structured prompt, and only when there's text to compare
        # against past corrections (vision mode reads images, not text).
        few_shot_examples = None
        if structured and config.get("learning", {}).get("few_shot_examples", False) \
                and extraction.content_type == "text" and extraction.content:
            try:
                store = _get_learning_store(config)
                few_shot_examples = store.find_similar_corrections(extraction.content, limit=3) or None
            except Exception as e:
                logging.warning(f"Few-shot example lookup failed: {e}")

        prompt = APIClient._build_prompt(
            extraction, config, document_types, candidates, few_shot_examples
        )

        if config.get("processing", {}).get("debug_log_prompt", False):
            logging.info(f"[DEBUG PROMPT]\n{prompt}\n[END PROMPT]")

        api_cfg = config["api"]
        errors = []

        # Try OpenWebUI (OpenAI-compatible)
        try:
            raw = APIClient._call_openwebui(prompt, extraction, api_cfg)
            return APIClient._parse_response(raw)
        except Exception as e:
            errors.append(f"OpenWebUI: {e}")
            logging.warning(f"OpenWebUI API call failed: {e}")

        # Fallback: direct Ollama
        try:
            schema = APIClient._STRUCTURED_SCHEMA if structured else None
            raw = APIClient._call_ollama(prompt, extraction, api_cfg, schema=schema)
            return APIClient._parse_response(raw)
        except Exception as e:
            errors.append(f"Ollama: {e}")
            logging.error(f"Ollama fallback also failed: {e}")

        raise ConnectionError("Both API endpoints failed.\n" + "\n".join(errors))

    @staticmethod
    def _call_openwebui(prompt: str, extraction: ExtractionResult, api_cfg: dict) -> str:
        url = api_cfg["openwebui_url"].rstrip("/") + "/api/chat/completions"
        headers = {"Content-Type": "application/json"}
        if api_cfg.get("api_key"):
            headers["Authorization"] = f"Bearer {api_cfg['api_key']}"

        if extraction.content_type == "image":
            # OpenAI vision format: content is an array. Send every page
            # the extractor produced (vision mode may return multiple).
            image_list = extraction.images or [extraction.content]
            content_parts = [
                {"type": "text", "text": prompt.replace("[See attached image]", "").strip()}
            ]
            for b64 in image_list:
                content_parts.append({
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{extraction.mime_type};base64,{b64}"
                    },
                })
            messages = [{"role": "user", "content": content_parts}]
        else:
            messages = [{"role": "user", "content": prompt}]

        payload = {
            "model": api_cfg["model"],
            "messages": messages,
            "stream": False,
        }
        resp = requests.post(
            url, json=payload, headers=headers,
            timeout=(api_cfg["timeout_connect"], api_cfg["timeout_read"])
        )
        resp.raise_for_status()
        return resp.json()["choices"][0]["message"]["content"]

    @staticmethod
    def _call_ollama(prompt: str, extraction: ExtractionResult, api_cfg: dict,
                      schema: Optional[dict] = None) -> str:
        url = api_cfg["ollama_url"].rstrip("/") + "/api/generate"
        payload = {
            "model": api_cfg["model"],
            "prompt": prompt,
            # Ollama's `format` accepts either the string "json" (loose,
            # legacy behavior) or a JSON-schema object constraining the
            # exact shape of the response. Use the schema when the
            # structured-output prompt is in play so malformed responses
            # become rare; otherwise keep the legacy "json" behavior
            # unchanged.
            "format": schema if schema else "json",
            "stream": False,
        }
        if extraction.content_type == "image":
            payload["images"] = extraction.images or [extraction.content]

        resp = requests.post(
            url, json=payload,
            timeout=(api_cfg["timeout_connect"], api_cfg["timeout_read"])
        )
        resp.raise_for_status()
        body = resp.json()
        # Ollama can return HTTP 200 with an error payload instead of a
        # completion — e.g. {"error": "model 'X' not found, try pulling it
        # first"} when the configured model isn't pulled, or a schema
        # rejection when `format` is a JSON-schema object the installed
        # Ollama version doesn't support. raise_for_status() doesn't catch
        # this (200 is still 200), and silently falling back to "" here
        # made every such failure look identical to "the model returned
        # unparseable text" with no way to tell them apart in the log.
        if "response" not in body:
            err = body.get("error") or json.dumps(body)[:300]
            raise ValueError(f"Ollama returned no completion — {err}")
        return body["response"]

    @staticmethod
    def _find_balanced_json_object(raw: str) -> Optional[str]:
        """Find the first balanced {...} substring in `raw`, tracking string
        literals/escapes so quoted braces don't confuse the depth count.

        Replaces the old `r'\\{[^{}]*\\}'` regex, which only matched a FLAT
        object — the moment the model's answer contained any nested object
        (e.g. wrapped as {"result": {...}}, or a stray metadata sub-object)
        that regex would grab the innermost `{...}` instead of the real
        answer, or fail outright. This scans forward from the first `{` and
        returns the substring once bracket depth returns to zero.
        """
        start = raw.find("{")
        if start == -1:
            return None
        depth = 0
        in_string = False
        escape = False
        for i in range(start, len(raw)):
            ch = raw[i]
            if in_string:
                if escape:
                    escape = False
                elif ch == "\\":
                    escape = True
                elif ch == '"':
                    in_string = False
                continue
            if ch == '"':
                in_string = True
            elif ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    return raw[start:i + 1]
        return None

    @staticmethod
    def _parse_response(raw: str) -> dict:
        """Parse the model's raw answer into a dict. Handles both the
        legacy {client, desc, confidence} shape and the structured
        {client, doc_type, recipient, direction, doc_date, confidence}
        shape — callers use .get() with defaults either way.

        Three attempts, in order:
          1. The whole response is valid JSON.
          2. A balanced {...} substring somewhere inside it (handles prose
             wrapping the JSON, and — unlike the old flat regex — a
             genuinely nested object inside the answer).
          3. Give up.

        On failure (3), the returned dict carries `_parse_failed: True` so
        callers can tell "we couldn't read the model's answer" apart from
        "the model said it doesn't know" (raw_client == NEEDS_REVIEW). The
        raw response is logged at WARNING either way.
        """
        # 1. direct parse
        try:
            data = json.loads(raw)
            if isinstance(data, dict) and "client" in data:
                return data
        except json.JSONDecodeError:
            pass

        # 2. balanced-brace scan
        candidate = APIClient._find_balanced_json_object(raw)
        if candidate:
            try:
                data = json.loads(candidate)
                if isinstance(data, dict):
                    return data
            except json.JSONDecodeError:
                pass

        # 3. give up
        logging.warning(f"Could not parse API response as JSON. Raw: {raw[:300]}")
        return {
            "client": "A-NEEDS REVIEW",
            "desc": "Unknown Document",
            "confidence": "low",
            "_parse_failed": True,
        }

    @staticmethod
    def test_connection(config: dict) -> tuple:
        api_cfg = config["api"]
        # Try OpenWebUI
        try:
            url = api_cfg["openwebui_url"].rstrip("/") + "/api/chat/completions"
            headers = {"Content-Type": "application/json"}
            if api_cfg.get("api_key"):
                headers["Authorization"] = f"Bearer {api_cfg['api_key']}"
            payload = {
                "model": api_cfg["model"],
                "messages": [{"role": "user", "content": "Reply with just the word: ready"}],
                "stream": False,
            }
            resp = requests.post(url, json=payload, headers=headers,
                                 timeout=(api_cfg["timeout_connect"], 30))
            resp.raise_for_status()
            return True, f"Connected via OpenWebUI  ({api_cfg['model']})"
        except Exception as e1:
            # Capture the text NOW. Python unbinds the `as` name at the end
            # of an except block, so referencing `e1` in the second handler
            # below raised UnboundLocalError — meaning the one case this
            # message exists for (both endpoints down, i.e. every
            # misconfigured setup) crashed instead of explaining itself.
            openwebui_error = str(e1) or e1.__class__.__name__
        # Try Ollama direct
        try:
            url = api_cfg["ollama_url"].rstrip("/") + "/api/generate"
            payload = {
                "model": api_cfg["model"],
                "prompt": "Reply with just the word: ready",
                "stream": False,
            }
            resp = requests.post(url, json=payload,
                                 timeout=(api_cfg["timeout_connect"], 30))
            resp.raise_for_status()
            return True, f"Connected via Ollama direct  ({api_cfg['model']})"
        except Exception as e2:
            ollama_error = str(e2) or e2.__class__.__name__
            return False, (
                f"Could not connect.\nOpenWebUI: {openwebui_error}\nOllama: {ollama_error}"
            )


# ─────────────────────────────────────────────────────────────
# NameTemplate — PASS 5 filename template engine
# ─────────────────────────────────────────────────────────────
#
# Renders a filename stem from structured fields (client, doc_type,
# recipient, doc_date, claim_number, direction) using a small template
# language:
#
#   {placeholder}       plain substitution — missing/empty fields render ""
#   [ ...{placeholder}...]   an optional segment: dropped in its entirety
#                            if ANY placeholder inside it is empty
#
# Example: "{client} - {doc_type}[ to {recipient}]" renders as
# "VALADEZ, Secilia - Reduction Request to Chiropractic Works" when a
# recipient is known, and "VALADEZ, Secilia - Reduction Request" (no
# dangling " to") when it isn't.
#
# Bracketed groups are a single flat level — nesting ("[a[b]c]") is not
# supported and is not needed by any template this tool ships.
#
# Only used when naming.use_templates is True (FileProcessor.process_file);
# with that flag off, filename construction is untouched.

class NameTemplate:
    MAX_LEN = 150  # excluding extension

    _PLACEHOLDER_RE = re.compile(r"\{(\w+)\}")
    _GROUP_RE = re.compile(r"\[([^\[\]]*)\]")

    @staticmethod
    def format_date(value: str, date_format: str) -> str:
        """Reformat an ISO 'YYYY-MM-DD' date string per `date_format`
        (e.g. "%m-%d-%y" -> "07-15-26"). If `value` isn't a parseable ISO
        date, it's returned verbatim — never dropped just because it
        doesn't match the expected shape. Empty input returns ""."""
        value = (value or "").strip()
        if not value:
            return ""
        try:
            dt = datetime.datetime.strptime(value, "%Y-%m-%d")
            return dt.strftime(date_format)
        except (ValueError, TypeError):
            return value

    @classmethod
    def _render_group(cls, body: str, fields: dict) -> str:
        placeholders = cls._PLACEHOLDER_RE.findall(body)
        if placeholders and any(not fields.get(p) for p in placeholders):
            return ""
        return cls._PLACEHOLDER_RE.sub(lambda m: fields.get(m.group(1), ""), body)

    @classmethod
    def render(cls, template: str, fields: dict) -> str:
        """Render `template` against `fields` (any key not present is
        treated as an empty string). See module-level docstring for the
        template language. Collapses whitespace and strips separators
        ("-", spaces) a dropped group can leave dangling at either end or
        doubled in the middle."""
        safe_fields = {k: (v or "") for k, v in (fields or {}).items()}

        rendered = cls._GROUP_RE.sub(lambda m: cls._render_group(m.group(1), safe_fields), template)
        rendered = cls._PLACEHOLDER_RE.sub(lambda m: safe_fields.get(m.group(1), ""), rendered)

        rendered = re.sub(r"\s+", " ", rendered).strip()
        rendered = re.sub(r"\s*-\s*-\s*", " - ", rendered)   # a dropped middle group can leave "X -  - Y"
        rendered = re.sub(r"^[\s\-]+", "", rendered)
        rendered = re.sub(r"[\s\-.]+$", "", rendered)        # trailing "-"/"." is unsafe on Windows anyway
        return rendered

    @classmethod
    def build(cls, template: str, fields: dict, max_len: int = MAX_LEN) -> str:
        """render() plus illegal-character sanitation and a length guard.

        Sanitizes for illegal filesystem characters the same way
        FileProcessor._safe_subject does, but deliberately does NOT
        title-case the result — the {client} segment is already
        "LAST, First" and must keep its exact capitalization and comma.

        If the rendered name would exceed `max_len` characters (excluding
        extension), the doc_type field is shortened first, then recipient,
        and the truncation is logged. A hard character-count truncation is
        the last resort if that still isn't enough, so this never returns
        something that could fail to write on Windows.
        """
        def _sanitize(text: str) -> str:
            text = ILLEGAL_CHARS_RE.sub("", text)
            return re.sub(r"\s+", " ", text).strip()

        rendered = _sanitize(cls.render(template, fields))
        if len(rendered) <= max_len:
            return rendered

        work_fields = dict(fields or {})
        truncated = False
        for key in ("doc_type", "recipient"):
            if len(rendered) <= max_len:
                break
            value = work_fields.get(key) or ""
            if not value:
                continue
            excess = len(rendered) - max_len
            new_len = max(0, len(value) - excess)
            work_fields[key] = value[:new_len].rstrip()
            truncated = True
            rendered = _sanitize(cls.render(template, work_fields))

        if len(rendered) > max_len:
            rendered = rendered[:max_len].rstrip(" -.")
            truncated = True

        if truncated:
            logging.info(
                f"Filename template rendered over {max_len} chars — truncated to fit: '{rendered}'"
            )
        return rendered


# ─────────────────────────────────────────────────────────────
# FileProcessor
# ─────────────────────────────────────────────────────────────

class FileProcessor:

    @staticmethod
    def process_file(file_path: str, config: dict, client_list: list,
                      batch_id: str = "", reserved: Optional[set] = None) -> ProcessResult:
        filename = os.path.basename(file_path)
        ext = os.path.splitext(filename)[1].lower()
        proc_cfg = config["processing"]
        safety_cfg = config.get("safety", {})
        naming_cfg = config.get("naming", {})
        dry_run = config.get("automation", {}).get("dry_run", False)
        if reserved is None:
            reserved = set()

        # Skip already-processed files
        if proc_cfg.get("skip_already_processed") and \
                FileProcessor._already_processed(filename, client_list, naming_cfg):
            return ProcessResult(
                original_name=filename,
                final_name=filename,
                status="skipped",
                skip_reason="already processed",
            )

        try:
            if not os.path.isfile(file_path):
                # Not an error: in a Dropbox-synced folder with more than
                # one instance running, this is exactly what it looks like
                # when another pass got to this file first between our
                # directory listing and now. Nothing to fix, nothing to
                # alarm the user about.
                return ProcessResult(
                    original_name=filename,
                    final_name=filename,
                    status="skipped",
                    skip_reason="already handled by another run",
                )

            # Dropbox may still be mid-sync. Sample the size twice ~0.6s
            # apart (repeating up to ~5s total) before reading the file —
            # a file that's still growing/shrinking isn't safe to OCR yet.
            settle_skip = FileProcessor._wait_for_settled(file_path)
            if settle_skip is not None:
                return ProcessResult(
                    original_name=filename,
                    final_name=filename,
                    status="skipped",
                    skip_reason=settle_skip,
                )

            doc_hash = FileProcessor._file_hash(file_path)

            # Decide extraction method. Vision mode is only honored when:
            #   1. The user selected it in Settings, AND
            #   2. The selected model is on the local vision allowlist.
            # Any other case silently falls back to OCR so the pipeline
            # never breaks if the user switches to a non-vision model.
            use_vision = (
                proc_cfg.get("extraction_method", "ocr") == "vision"
                and model_supports_vision(config.get("api", {}).get("model", ""))
            )

            reading_cfg = config.get("reading", {})
            skip_fax_cover_pages = reading_cfg.get("skip_fax_cover_pages", False)
            deskew_photos = reading_cfg.get("deskew_photos", False)

            # Extract content
            extraction = DocumentExtractor.extract(
                file_path,
                proc_cfg["max_ocr_chars"],
                proc_cfg.get("max_pages", 5),
                vision_mode=use_vision,
                max_vision_pages=proc_cfg.get("max_vision_pages", 2),
                ocr_preprocess=proc_cfg.get("ocr_preprocess", True),
                skip_fax_cover_pages=skip_fax_cover_pages,
                deskew_photos=deskew_photos,
            )

            # ── OCR -> vision escalation (reading.vision_escalation) ──
            # Only fires when we did NOT already use vision, only ever one
            # attempt, and only when the currently selected model can
            # actually see images — otherwise there's nothing to escalate
            # to and we just say so once and keep the OCR result.
            if reading_cfg.get("vision_escalation", False) and not use_vision \
                    and extraction.content_type == "text":
                model_name = config.get("api", {}).get("model", "")
                if model_supports_vision(model_name):
                    escalation_threshold = reading_cfg.get("escalation_threshold", 0.35)
                    quality = DocumentExtractor.assess_text_quality(extraction.content)
                    if quality < escalation_threshold:
                        logging.info(
                            f"{filename}: text quality {quality:.2f} below "
                            f"{escalation_threshold} — escalating to vision"
                        )
                        try:
                            vision_extraction = DocumentExtractor.extract(
                                file_path,
                                proc_cfg["max_ocr_chars"],
                                proc_cfg.get("max_pages", 5),
                                vision_mode=True,
                                max_vision_pages=proc_cfg.get("max_vision_pages", 2),
                                ocr_preprocess=proc_cfg.get("ocr_preprocess", True),
                                skip_fax_cover_pages=skip_fax_cover_pages,
                                deskew_photos=deskew_photos,
                            )
                            prior_method = extraction.method
                            vision_extraction.method = f"{prior_method}->vision"
                            extraction = vision_extraction
                        except Exception as e:
                            logging.warning(
                                f"{filename}: vision escalation failed, keeping OCR "
                                f"result: {e}"
                            )
                else:
                    logging.info(
                        f"{filename}: vision_escalation is on but model "
                        f"'{model_name}' is not vision-capable — keeping OCR result"
                    )

            # ── Claim number / DOI / DOB extraction (reading.extract_claim_numbers) ──
            # Runs on whatever text ended up being used for classification —
            # if escalation above switched to vision, there's no text left to
            # mine and identifiers stay empty for this document.
            identifiers = {"claim_number": "", "date_of_injury": "", "dob": ""}
            if reading_cfg.get("extract_claim_numbers", False) \
                    and extraction.content_type == "text" and extraction.content:
                identifiers = DocumentExtractor.extract_identifiers(extraction.content)

            # ── PASS 4 classification vocabularies (all off by default) ──
            cls_cfg = config.get("classification", {})
            paths_cfg = config.get("paths", {})

            document_types_for_prompt = None
            doc_type_alias_map: dict = {}
            if cls_cfg.get("use_document_types", False):
                doc_types_path = paths_cfg.get("document_types_file", "")
                document_types_for_prompt = DocumentTypeManager.load(doc_types_path)
                doc_type_alias_map = DocumentTypeManager.load_alias_map(doc_types_path)

            providers_list: list = []
            if cls_cfg.get("use_providers", False):
                providers_list = ProviderManager.load(paths_cfg.get("providers_file", ""))

            # Dead-code cleanup (PASS 4 item 6): ClientListManager.filter_candidates
            # existed but was never called. Wired in here behind its own flag —
            # only offered to the prompt when it actually narrowed the list
            # (its own fallback is the unfiltered full list, which isn't a
            # useful "candidates" hint).
            candidates_for_prompt = None
            if cls_cfg.get("use_candidate_shortlist", False) \
                    and extraction.content_type == "text" and extraction.content:
                shortlist_size = proc_cfg.get("candidate_list_size", 10)
                shortlist = ClientListManager.filter_candidates(
                    extraction.content, client_list, top_n=shortlist_size
                )
                if len(shortlist) < len(client_list):
                    candidates_for_prompt = shortlist

            # Classify via AI
            result = APIClient.classify(
                extraction, client_list, config,
                document_types=document_types_for_prompt,
                candidates=candidates_for_prompt,
            )
            raw_client = result.get("client", "NEEDS_REVIEW").strip().strip("\"'")
            raw_desc = result.get("desc", "Unknown Document")
            raw_doc_type = (result.get("doc_type") or "").strip()
            raw_recipient = (result.get("recipient") or "").strip()
            direction = (result.get("direction") or "").strip().lower()
            if direction not in ("incoming", "outgoing"):
                direction = ""
            raw_doc_date = (result.get("doc_date") or "").strip()
            confidence = result.get("confidence", "low")
            raw_confidence = confidence
            parse_failed = bool(result.get("_parse_failed", False))

            # First 2000 chars of the actual extracted text, for later
            # learning/dedupe features. Vision mode reads images, not text,
            # so there's nothing textual to capture there.
            extracted_text = (
                extraction.content[:2000]
                if extraction.content_type == "text" and extraction.content
                else ""
            )

            # Log what the AI returned so failures are easy to diagnose
            logging.info(f"{filename}: AI returned client='{raw_client}' confidence={confidence}")

            # ── Grounding self-check (classification.grounding_check) ──
            # Text extractions only — vision mode never produced OCR text to
            # check against. Also feeds evidence_confidence's rubric below,
            # so it runs whenever either flag wants it, but only FORCES a
            # confidence downgrade when grounding_check itself is on.
            grounded = True
            grounding_evaluated = False
            if extraction.content_type == "text" and extraction.content \
                    and raw_client not in ("NEEDS_REVIEW", "A-NEEDS REVIEW", "") \
                    and (cls_cfg.get("grounding_check", False)
                         or cls_cfg.get("evidence_confidence", False)):
                grounded = FileProcessor._grounding_check(raw_client, extraction.content)
                grounding_evaluated = True
                if not grounded and cls_cfg.get("grounding_check", False):
                    logging.warning(
                        f"{filename}: model returned client '{raw_client}' but "
                        "that name does not appear in the document — "
                        "downgrading confidence"
                    )
                    confidence = "low"

            # ── Document type normalization (classification.use_document_types) ──
            # Keeps the model's raw printed title when nothing in the
            # vocabulary is close enough — never discards it.
            doc_type_matched = False
            final_doc_type = raw_doc_type
            if cls_cfg.get("use_document_types", False) and raw_doc_type:
                normalized_type = DocumentTypeManager.normalize(raw_doc_type, doc_type_alias_map)
                if normalized_type:
                    final_doc_type = normalized_type
                    doc_type_matched = True

            # ── Recipient normalization (classification.use_providers) ──
            final_recipient = raw_recipient
            if cls_cfg.get("use_providers", False) and raw_recipient:
                normalized_recipient = ProviderManager.normalize(raw_recipient, providers_list)
                if normalized_recipient:
                    final_recipient = normalized_recipient
                else:
                    final_recipient = ProviderManager.strip_address(raw_recipient)

            # ── Evidence-based confidence (classification.evidence_confidence) ──
            # Overrides the model's self-reported confidence with a score
            # computed from independently-checkable signals. See
            # _compute_confidence's docstring for the rubric.
            if cls_cfg.get("evidence_confidence", False):
                confidence = FileProcessor._compute_confidence(
                    raw_client, extraction, client_list,
                    doc_type_matched=doc_type_matched,
                    grounded=grounded,
                    grounding_evaluated=grounding_evaluated,
                    parse_failed=parse_failed,
                )

            # Confidence gate — configurable in Settings.
            # require_high_confidence=True (default): only "high" proceeds to rename.
            # require_high_confidence=False: "medium" also proceeds (more matches, more risk).
            # "low" always goes to NEEDS_REVIEW regardless of this setting.
            require_high = proc_cfg.get("require_high_confidence", True)
            _confidence_ok = (
                confidence == "high"
                or (confidence == "medium" and not require_high)
            )
            _skip_fuzzy = (
                raw_client in ("NEEDS_REVIEW", "A-NEEDS REVIEW", "")
                or not _confidence_ok
            )
            if not _confidence_ok and raw_client not in ("NEEDS_REVIEW", "A-NEEDS REVIEW", ""):
                logging.info(
                    f"{filename}: confidence={confidence} "
                    f"({'below high threshold' if require_high else 'low'}) "
                    "— sending to NEEDS_REVIEW without matching"
                )

            threshold = proc_cfg.get("fuzzy_threshold", 0.82)
            matched = None if _skip_fuzzy else \
                ClientListManager.fuzzy_match(raw_client, client_list, threshold)

            if matched:
                logging.info(f"{filename}: fuzzy matched '{raw_client}' → '{matched}'")

            if matched:
                final_client = matched
                status = "renamed"
                match_source = "fuzzy"
            else:
                # PASS 6 (learning): before falling back to an unknown-client
                # label, see if a claim number or a raw-name relationship
                # already learned from past corrections resolves this file
                # (the "George was a passenger in Mary's accident" case).
                final_client, status, match_source = FileProcessor._resolve_unmatched_client(
                    raw_client=raw_client,
                    claim_number=identifiers.get("claim_number", ""),
                    doc_hash=doc_hash,
                    filename=filename,
                    raw_desc=raw_desc,
                    extracted_text=extracted_text,
                    config=config,
                    naming_cfg=naming_cfg,
                )

            # Which text feeds the description / {doc_type} placeholder: the
            # model's own printed doc_type when structured_output is on, the
            # free-form desc otherwise (today's exact behavior either way).
            if cls_cfg.get("structured_output", False) and final_doc_type:
                safe_desc = FileProcessor._safe_subject(final_doc_type) or "Document"
            else:
                safe_desc = FileProcessor._safe_subject(raw_desc) or "Document"

            # Safety net: if the description is still fax-related despite the prompt rule,
            # substitute a neutral fallback rather than letting it become the filename.
            if FileProcessor._desc_is_fax(safe_desc):
                logging.info(
                    f"{filename}: AI returned fax-related desc '{safe_desc}' — "
                    "substituting 'Incoming Document'"
                )
                safe_desc = "Incoming Document"

            # ── Filename construction (naming.use_templates) ──────────────
            # Off (default): byte-for-byte today's construction. On: render
            # via the per-doc-type template — recipient/date segments are
            # further gated by naming.include_recipient/include_doc_date so
            # a document's fields can be extracted without appearing in the
            # name unless the office has actually turned that on.
            if naming_cfg.get("use_templates", False):
                doc_type_display = FileProcessor._safe_subject(final_doc_type) if final_doc_type else safe_desc
                if not doc_type_display:
                    doc_type_display = "Document"
                if FileProcessor._desc_is_fax(doc_type_display):
                    doc_type_display = "Incoming Document"

                recipient_display = ""
                if naming_cfg.get("include_recipient", False) and final_recipient:
                    recipient_display = FileProcessor._safe_subject(final_recipient)

                doc_date_source = raw_doc_date or identifiers.get("date_of_injury", "")
                doc_date_display = ""
                if naming_cfg.get("include_doc_date", False) and doc_date_source:
                    doc_date_display = NameTemplate.format_date(
                        doc_date_source, naming_cfg.get("date_format", "%m-%d-%y")
                    )

                template_fields = {
                    "client": final_client,
                    "doc_type": doc_type_display,
                    "recipient": recipient_display,
                    "doc_date": doc_date_display,
                    "claim_number": identifiers.get("claim_number", ""),
                    "direction": direction,
                }
                templates_map = naming_cfg.get("templates") or {}
                template_str = (
                    templates_map.get(final_doc_type)
                    or naming_cfg.get("default_template")
                    or "{client} - {doc_type}"
                )
                stem = NameTemplate.build(template_str, template_fields)
                if not stem:
                    stem = f"{final_client} - {doc_type_display}".strip(" -")
                new_name = f"{stem}{ext}"
            else:
                new_name = f"{final_client} - {safe_desc}{ext}"

            # Collision avoidance. `reserved` also carries names already
            # claimed earlier in this batch — needed in dry-run, where
            # nothing actually lands on disk, so two documents that would
            # both become the same name must still preview as distinct.
            # naming.date_disambiguation: before falling back to a bare
            # "(1)"/"(2)" counter, try appending the document date, then the
            # claim number — a far more informative disambiguator than a
            # meaningless index for the office's alphabetical listing.
            dest_dir = os.path.dirname(file_path)
            collision_extra = None
            if naming_cfg.get("date_disambiguation", False):
                collision_extra = {
                    "date_disambiguation": True,
                    "doc_date": raw_doc_date or identifiers.get("date_of_injury", ""),
                    "claim_number": identifiers.get("claim_number", ""),
                    "date_format": naming_cfg.get("date_format", "%m-%d-%y"),
                }
            new_name = FileProcessor._resolve_collision(
                dest_dir, new_name, filename, reserved=reserved, extra=collision_extra
            )
            reserved.add(new_name)

            renamed_at = None
            if new_name != filename and not dry_run:
                # Re-validate immediately before touching disk: another
                # instance (or a stray earlier pass) may have already
                # claimed this exact file since we started reading it.
                if not os.path.isfile(file_path):
                    return ProcessResult(
                        original_name=filename,
                        final_name=filename,
                        status="skipped",
                        skip_reason="already handled by another run",
                        doc_hash=doc_hash,
                    )
                if safety_cfg.get("recheck_before_rename", True):
                    current_hash = FileProcessor._file_hash(file_path)
                    if current_hash != doc_hash:
                        return ProcessResult(
                            original_name=filename,
                            final_name=filename,
                            status="skipped",
                            skip_reason="file changed during processing",
                            doc_hash=doc_hash,
                        )
                # NOTE: there used to be a second _already_processed(filename)
                # check here, meant to catch "another instance renamed this
                # file while we were reading it". It could never fire —
                # `filename` is captured once at the top of this function and
                # never reassigned, so it re-evaluated the identical check
                # already made at entry. That case is genuinely covered by
                # the os.path.isfile and hash re-checks immediately above,
                # which look at the actual state on disk.
                dest_path = os.path.join(dest_dir, new_name)
                os.rename(file_path, dest_path)
                renamed_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if safety_cfg.get("undo_log", True):
                    log_rename(batch_id, "rename", file_path, dest_path, "auto")

            return ProcessResult(
                original_name=filename,
                final_name=new_name,
                status=status,
                client=final_client,
                description=safe_desc,
                confidence=confidence,
                renamed_at=renamed_at,
                extraction_method=extraction.method,
                raw_client=raw_client,
                raw_confidence=raw_confidence,
                extracted_text=extracted_text,
                doc_hash=doc_hash,
                claim_number=identifiers.get("claim_number", ""),
                # doc_date: prefer the date the model read directly off the
                # document itself (letterhead/signature date, exam/visit
                # date — only populated when classification.structured_output
                # is on) and fall back to date_of_injury only when the model
                # didn't return one, since a document is often dated well
                # after the injury. This is the "later pass" PASS 3 left a
                # comment for.
                doc_date=raw_doc_date or identifiers.get("date_of_injury", ""),
                doc_type=final_doc_type,
                recipient=final_recipient,
                direction=direction,
                was_dry_run=dry_run,
                match_source=match_source,
            )

        except Exception as e:
            logging.error(f"Error processing {filename}: {e}", exc_info=True)
            return ProcessResult(
                original_name=filename,
                final_name=filename,
                status="error",
                error_message=str(e),
            )

    # Phrases that indicate the AI described the fax wrapper rather than the real document
    _FAX_DESC_PATTERNS = [
        "fax", "facsimile", "telecopy", "send result", "transmission report",
        "activity report", "communication journal",
    ]

    @staticmethod
    def _desc_is_fax(desc: str) -> bool:
        lower = desc.lower()
        return any(p in lower for p in FileProcessor._FAX_DESC_PATTERNS)

    # ── Grounding self-check (classification.grounding_check) ──────────────

    @staticmethod
    def _grounding_check(client_name: str, text: str) -> bool:
        """For text extractions only: verify a model-returned client name is
        actually present in the extracted document text, catching
        hallucinated names cheaply.

        Compares the normalized surname against the normalized document
        text two ways:
          1. Plain containment — fast path, covers the vast majority of
             real documents.
          2. A difflib ratio over a sliding window the width of the
             surname, to tolerate OCR noise (a dropped or substituted
             letter) that would defeat plain containment.

        Returns True when there's nothing meaningful to check (empty name,
        empty text, or a surname too short to check reliably) so a caller
        never downgrades a document there was no way to verify.
        """
        if not client_name or not text:
            return True
        name = client_name.strip()
        surname = name.split(",")[0].strip() if "," in name else (name.split() or [name])[-1]
        surname_norm = re.sub(r"[^a-z]", "", surname.lower())
        if len(surname_norm) < 3:
            return True

        text_norm = re.sub(r"[^a-z]", "", text.lower())
        if not text_norm:
            return True

        # 1. plain containment
        if surname_norm in text_norm:
            return True

        # 2. sliding-window difflib ratio, to tolerate OCR noise
        window = len(surname_norm)
        step = max(1, window // 2)
        best = 0.0
        for i in range(0, max(1, len(text_norm) - window + 1), step):
            chunk = text_norm[i:i + window]
            score = difflib.SequenceMatcher(None, surname_norm, chunk).ratio()
            if score > best:
                best = score
                if best >= 0.9:
                    break
        return best >= 0.82

    # ── Evidence-based confidence (classification.evidence_confidence) ─────

    @staticmethod
    def _compute_confidence(raw_client: str, extraction: "ExtractionResult",
                             client_list: list, doc_type_matched: bool,
                             grounded: bool, grounding_evaluated: bool,
                             parse_failed: bool) -> str:
        """Compute a confidence level from independently-checkable evidence
        instead of trusting the model's own self-reported confidence field.

        Rubric — each signal nudges a running score, then the score is
        bucketed into high/medium/low:

            +2 strong     the client name was found under a recognized
                          label (DocumentExtractor._extract_labeled_snippets
                          returns a snippet that overlaps the returned name)
            +2 strong     fuzzy match score against the client list >= 0.95
            +1 moderate   doc_type matched a known canonical type
            +1 moderate   the grounding check ran and passed
            -3 strong     response parsing failed (APIClient._parse_response
                          had to fall back — `_parse_failed`)
            -3 strong     the grounding check ran and failed
            -2 strong     extraction text quality (assess_text_quality) is
                          below 0.35 — the OCR itself looked unreliable

            score >= 3  -> "high"
            score >= 1  -> "medium"
            otherwise   -> "low"

        `grounding_evaluated` distinguishes "the check ran and passed" from
        "the check never ran" (vision mode, empty client, or both grounding
        flags off) — the latter is neutral, not evidence of anything, so it
        contributes neither the + nor the - grounding signal.

        Vision extractions (no OCR text) skip the text-dependent signals
        (labeled snippet, text quality) but still get the fuzzy-match,
        doc_type, and parse-failure signals.

        Only called when classification.evidence_confidence is on; the
        model's self-reported confidence is used unchanged otherwise.
        """
        score = 0.0
        text = extraction.content if extraction.content_type == "text" else ""

        if text and raw_client:
            client_last = raw_client.split(",")[0].strip().lower()
            for snippet in DocumentExtractor._extract_labeled_snippets(text):
                snippet_lower = snippet.lower()
                if client_last and (client_last in snippet_lower or snippet_lower in raw_client.lower()):
                    score += 2
                    break

        if raw_client and client_list:
            candidate_norm = ClientListManager._normalize(raw_client)
            best = 0.0
            for entry in client_list:
                r = difflib.SequenceMatcher(None, candidate_norm, ClientListManager._normalize(entry)).ratio()
                if r > best:
                    best = r
            if best >= 0.95:
                score += 2

        if doc_type_matched:
            score += 1

        if grounding_evaluated:
            if grounded:
                score += 1
            else:
                score -= 3

        if parse_failed:
            score -= 3
        if text and DocumentExtractor.assess_text_quality(text) < 0.35:
            score -= 2

        if score >= 3:
            return "high"
        if score >= 1:
            return "medium"
        return "low"

    # Small joining words that stay lowercase when they aren't the first
    # token of the subject (e.g. "Reduction Request to Chiropractic Works").
    _LOWERCASE_JOINERS = {
        "to", "of", "and", "for", "in", "on", "at", "by", "the", "a", "an",
        "with", "from", "vs",
    }

    @staticmethod
    def _is_acronym(token: str) -> bool:
        """True if `token` is an ALL-CAPS 2-5 char run of letters (PPR, TTD,
        IME, MRI, EMG, NCV, PPD, TTP, DOI, DOB, PT, OT, ...). Trailing
        punctuation (e.g. a stray period) is ignored when checking."""
        core = token.strip(".,;:!?")
        return 2 <= len(core) <= 5 and core.isalpha() and core.isupper()

    @staticmethod
    def _safe_subject(text: str, preserve_acronyms: bool = True) -> str:
        # Replace slashes and underscores with spaces BEFORE stripping illegal
        # characters, so "EMG/NCV" -> "EMG NCV" and "Incoming_Document" ->
        # "Incoming Document" instead of the words being silently fused.
        text = text.replace("/", " ").replace("_", " ")
        text = ILLEGAL_CHARS_RE.sub("", text)
        text = re.sub(r"\s+", " ", text).strip()

        if not preserve_acronyms:
            # Legacy behavior: no acronym awareness at all.
            text = re.sub(r"([a-z])([A-Z])", r"\1 \2", text)
            text = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1 \2", text)
            return text.title()

        # Split run-on words: insert space before uppercase letters that follow
        # a lowercase letter (e.g. "RetainerAgreement" → "Retainer Agreement").
        # Skip this when the whole token is already a recognised acronym so
        # "EMG" or "IME" are never split apart.
        def _split_run_on(token: str) -> str:
            if FileProcessor._is_acronym(token):
                return token
            token = re.sub(r"([a-z])([A-Z])", r"\1 \2", token)
            token = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1 \2", token)
            return token

        tokens = text.split(" ")
        split_tokens: List[str] = []
        for tok in tokens:
            if not tok:
                continue
            split_tokens.extend(_split_run_on(tok).split(" "))

        out_tokens = []
        for i, tok in enumerate(split_tokens):
            if FileProcessor._is_acronym(tok):
                out_tokens.append(tok)
            elif i > 0 and tok.lower() in FileProcessor._LOWERCASE_JOINERS:
                out_tokens.append(tok.lower())
            else:
                out_tokens.append(tok[:1].upper() + tok[1:].lower() if tok else tok)

        return " ".join(out_tokens).strip()

    @staticmethod
    def _resolve_collision(directory: str, filename: str, source_name: str,
                            reserved: Optional[set] = None,
                            extra: Optional[dict] = None) -> str:
        """If `filename` already exists in `directory` (and isn't the source
        file) OR has already been claimed via `reserved`, find a free name.

        `reserved` lets a batch track names it has already handed out to
        earlier files without relying on the filesystem — essential in
        dry-run mode, where nothing is actually written to disk, so two
        documents that would both become the same name must still preview
        as distinct.

        `extra` (naming.date_disambiguation) — when given with
        `extra["date_disambiguation"]` truthy, informative disambiguators
        are tried before the bare "(1)", "(2)", … counter: first the
        document date (`extra["doc_date"]`, formatted per
        `extra["date_format"]`), then the claim number
        (`extra["claim_number"]`). Either may be absent or fail to help —
        this always falls through to the counter as a last resort. `extra`
        is optional and defaults to None so call sites without these
        fields (e.g. the audit-mode rename) are unaffected."""
        if filename == source_name:
            return filename

        def _taken(name: str) -> bool:
            return (
                os.path.exists(os.path.join(directory, name))
                or (reserved is not None and name in reserved)
            )

        if not _taken(filename):
            return filename
        base, ext = os.path.splitext(filename)

        if extra and extra.get("date_disambiguation"):
            doc_date = (extra.get("doc_date") or "").strip()
            if doc_date:
                formatted = NameTemplate.format_date(doc_date, extra.get("date_format") or "%m-%d-%y")
                formatted = ILLEGAL_CHARS_RE.sub("", formatted).strip()
                if formatted:
                    candidate = f"{base} {formatted}{ext}"
                    if not _taken(candidate):
                        return candidate
            claim_number = (extra.get("claim_number") or "").strip()
            if claim_number:
                claim_clean = ILLEGAL_CHARS_RE.sub("", claim_number).strip()
                if claim_clean:
                    candidate = f"{base} {claim_clean}{ext}"
                    if not _taken(candidate):
                        return candidate

        counter = 1
        while True:
            candidate = f"{base} ({counter}){ext}"
            if not _taken(candidate):
                return candidate
            counter += 1

    @staticmethod
    def _wait_for_settled(file_path: str, max_wait: float = 5.0,
                           sample_interval: float = 0.6) -> Optional[str]:
        """Guard against reading a file mid-sync (Dropbox). Samples the file
        size twice ~`sample_interval` seconds apart, repeating until the
        size stops changing or `max_wait` seconds have elapsed. Returns
        None when the file looks settled and non-empty; otherwise a
        skip_reason string explaining why the file should be skipped
        for now rather than processed."""
        try:
            size = os.path.getsize(file_path)
        except OSError:
            return "file still being written"
        if size == 0:
            return "empty file (0 bytes)"

        waited = 0.0
        while waited < max_wait:
            time.sleep(sample_interval)
            waited += sample_interval
            try:
                new_size = os.path.getsize(file_path)
            except OSError:
                return "file still being written"
            if new_size == size:
                return "empty file (0 bytes)" if new_size == 0 else None
            size = new_size

        return "file still being written"

    @staticmethod
    def _file_hash(path: str) -> str:
        """Return the sha256 hex digest of a file's bytes, read in chunks.
        Returns "" on any error (missing file, permission error, etc.) rather
        than raising — this is a best-effort helper for dedupe/learning
        features, not something that should ever break processing."""
        try:
            h = hashlib.sha256()
            with open(path, "rb") as f:
                for chunk in iter(lambda: f.read(65536), b""):
                    h.update(chunk)
            return h.hexdigest()
        except Exception:
            return ""

    # Sentinel values the model can return (or _parse_response can fall back
    # to) meaning "no usable person name" — never a plausible client name.
    _CLIENT_SENTINELS = frozenset({"", "NEEDS_REVIEW", "A-NEEDS REVIEW", "A-UNKNOWN CLIENT"})

    @staticmethod
    def is_sentinel_client(name: str, extra_labels: Optional[list] = None) -> bool:
        """True if `name` is one of the tool's own "unresolved" placeholders
        rather than a real client — the built-in sentinels above, plus any
        labels a firm customized in Settings (naming.unknown_client_label /
        naming.no_client_label), which callers pass via `extra_labels`.

        This is the single gate that keeps a placeholder from ever being
        treated as a client name. It matters most in LearningStore: a
        no-change commit on a still-unresolved row records
        corrected_client == "A-NEEDS REVIEW", and without this check that
        would accumulate as alias/claim evidence and eventually be offered
        as 'file George Martinez under A-NEEDS REVIEW'. Accepting that
        suggestion would rename real documents to the placeholder while
        marking them "renamed" — filed, to all appearances, but lost.
        """
        candidate = (name or "").strip()
        if candidate.upper() in FileProcessor._CLIENT_SENTINELS:
            return True
        for label in (extra_labels or []):
            if label and candidate.upper() == str(label).strip().upper():
                return True
        return False

    @staticmethod
    def _looks_like_person_name(raw_client: str) -> bool:
        """True if `raw_client` plausibly holds a person's name — non-empty,
        not one of the model's own NEEDS_REVIEW-style sentinels, and made up
        of at least two name-like tokens (e.g. "SMITH, John" or "John
        Smith"). Used by naming.split_unknown_states to tell apart
        A-UNKNOWN CLIENT (a name was read but isn't on the client list —
        the "George was a passenger in Mary's accident" case) from
        A-NEEDS REVIEW (no name could be read at all)."""
        if not raw_client:
            return False
        stripped = raw_client.strip()
        if stripped.upper() in FileProcessor._CLIENT_SENTINELS:
            return False
        tokens = [t for t in re.split(r"[,\s]+", stripped) if t]
        name_tokens = [t for t in tokens if re.match(r"^[A-Za-z][A-Za-z'\-]*\.?$", t)]
        return len(name_tokens) >= 2

    @staticmethod
    def _resolve_unmatched_client(raw_client: str, claim_number: str, doc_hash: str,
                                   filename: str, raw_desc: str, extracted_text: str,
                                   config: dict, naming_cfg: dict) -> tuple:
        """PASS 6 payoff: what to do once fuzzy_match has failed, tried in
        this order before falling back to an unknown-client label:

          1. learning.claim_linking == "auto" and a claim number was read:
             try LearningStore.lookup_claim. On a hit, use that client.
          2. learning.client_relationships == "auto": try
             LearningStore.lookup_alias(raw_client) (the "George is a
             passenger in Mary's accident" case). On a hit, use that client.
          3. Either flag set to "suggest" (not "auto"): apply nothing, just
             record the raw guess as an observation so a later confirmed
             alias can retroactively sweep this file (see
             LearningStore.plan_retroactive_renames).
          4. Both flags "off" (the default): do nothing — byte-for-byte the
             pre-PASS-6 unknown-label fallback.

        Returns (final_client, status, match_source)."""
        learning_cfg = config.get("learning", {})
        claim_mode = learning_cfg.get("claim_linking", "off")
        alias_mode = learning_cfg.get("client_relationships", "off")

        if claim_mode != "off" or alias_mode != "off":
            try:
                store = _get_learning_store(config)

                if claim_mode == "auto" and claim_number:
                    claim_client = store.lookup_claim(claim_number)
                    if claim_client:
                        logging.info(
                            f"{filename}: claim number '{claim_number}' resolved "
                            f"to learned client '{claim_client}'"
                        )
                        return claim_client, "renamed", "claim"

                if alias_mode == "auto" and FileProcessor._looks_like_person_name(raw_client):
                    alias_client = store.lookup_alias(raw_client)
                    if alias_client:
                        logging.info(
                            f"{filename}: raw name '{raw_client}' resolved to "
                            f"learned client '{alias_client}' (alias)"
                        )
                        return alias_client, "renamed", "alias"

                if FileProcessor._looks_like_person_name(raw_client) and (
                        claim_mode == "suggest" or alias_mode == "suggest"):
                    store.log_observation(
                        doc_hash=doc_hash, raw_client=raw_client,
                        original_name=filename, claim_number=claim_number,
                        predicted_desc=raw_desc, text_excerpt=extracted_text,
                    )
            except Exception as e:
                logging.warning(f"{filename}: learning lookup failed: {e}")

        # Two distinct unresolved states (naming.split_unknown_states):
        #   A-UNKNOWN CLIENT — a person's name WAS read but doesn't
        #     match the client list (the "George was a passenger in
        #     Mary's accident" case; now backlogged for PASS 6 to learn).
        #   A-NEEDS REVIEW — no person name could be read at all, or
        #     the model returned NEEDS_REVIEW / parsing failed.
        # Both keep the "A-" prefix so they sort to the top of an
        # alphabetical folder listing. When the flag is off, both
        # cases collapse to no_client_label exactly as today.
        if naming_cfg.get("split_unknown_states", False) \
                and FileProcessor._looks_like_person_name(raw_client):
            final_client = naming_cfg.get("unknown_client_label", "A-UNKNOWN CLIENT")
        else:
            final_client = naming_cfg.get("no_client_label", "A-NEEDS REVIEW")
        return final_client, "needs_review", ""

    @staticmethod
    def _already_processed(filename: str, client_list: list,
                            naming_cfg: Optional[dict] = None) -> bool:
        """True if the file already looks like 'LAST, First - Subject.ext'
        with a recognised client name at the front.

        Short-circuits to False for either of the two "unresolved" labels
        this tool can produce — A-NEEDS REVIEW (no name could be read) and
        A-UNKNOWN CLIENT (a name was read but isn't on the client list,
        naming.split_unknown_states) — so files carrying either one are
        never mistaken for already-named and silently skipped forever.
        `naming_cfg` (config["naming"]) lets a caller honor customized
        label text; the two default labels are always checked either way."""
        labels = {"A-NEEDS REVIEW", "A-UNKNOWN CLIENT"}
        if naming_cfg:
            labels.add(naming_cfg.get("no_client_label", "A-NEEDS REVIEW"))
            labels.add(naming_cfg.get("unknown_client_label", "A-UNKNOWN CLIENT"))
        if any(lbl and filename.startswith(lbl) for lbl in labels):
            return False
        match = re.match(r"^(.+?) - .+\.(pdf|jpg|jpeg)$", filename, re.IGNORECASE)
        if not match:
            return False
        name_part = match.group(1)
        return ClientListManager.fuzzy_match(name_part, client_list, threshold=0.90) is not None


# ─────────────────────────────────────────────────────────────
# ProcessingEngine  (runs in a background thread)
# ─────────────────────────────────────────────────────────────

class ProcessingEngine:
    def __init__(self):
        self._stop_event = threading.Event()
        self.batch_id: str = ""   # set at the start of each run_batch(); "" before any batch has run

    def stop(self):
        self._stop_event.set()

    @staticmethod
    def _list_supported_files(scandocs: str) -> List[str]:
        return [
            f for f in os.listdir(scandocs)
            if os.path.isfile(os.path.join(scandocs, f))
            and os.path.splitext(f)[1].lower() in SUPPORTED_EXTENSIONS
        ]

    def run_batch(self, config: dict, result_queue: queue.Queue):
        self._stop_event.clear()

        # Work from a private snapshot. The caller hands us the live
        # ConfigManager.config dict, and nothing stops the user opening
        # Settings and hitting Save while a batch runs — without this, that
        # would change the rules mid-batch. automation.dry_run is the
        # alarming one: flip it halfway and the first half of the run is a
        # preview while the second half really renames files.
        try:
            config = ConfigManager._deep_copy(config)
        except Exception as e:
            logging.warning(f"Could not snapshot config for batch ({e}) — using it live.")

        scandocs = config["paths"]["scandocs_folder"]
        client_list_path = config["paths"]["client_list_file"]

        self.batch_id = uuid.uuid4().hex[:8]
        safety_cfg = config.get("safety", {})
        dry_run = config.get("automation", {}).get("dry_run", False)
        use_lock = safety_cfg.get("instance_lock", True) and bool(scandocs) and os.path.isdir(scandocs)

        lock = FolderLock(scandocs) if use_lock else None
        if lock is not None:
            try:
                lock.acquire()
            except FolderLockHeld as e:
                result_queue.put({"type": "error", "message": str(e)})
                result_queue.put({"type": "done"})
                return

        try:
            try:
                client_list = ClientListManager.load(client_list_path)
                if not client_list:
                    result_queue.put({
                        "type": "error",
                        "message": (
                            "The client list is empty.\n\n"
                            "Go to the Client List tab, add your clients, and save."
                        ),
                    })
                    return

                try:
                    files = self._list_supported_files(scandocs)
                except Exception as e:
                    result_queue.put({
                        "type": "error",
                        "message": f"Cannot read scandocs folder:\n{e}",
                    })
                    return

                if not files:
                    result_queue.put({
                        "type": "error",
                        "message": "No PDF or JPG files found in the scandocs folder.",
                    })
                    return

                result_queue.put({"type": "total", "count": len(files)})

                # Names claimed so far this batch — carried across files so
                # dry-run collision resolution doesn't depend on the
                # filesystem (nothing is actually written to disk then).
                reserved: set = set()

                # Every filename this batch has touched, in EITHER direction:
                # what we read, and what we renamed it to. The straggler
                # sweep below subtracts this from a fresh listing, so a file
                # this batch produced can never be mistaken for a new
                # arrival. Inferring that from _already_processed() does not
                # work — it deliberately returns False for the
                # A-NEEDS REVIEW / A-UNKNOWN CLIENT labels (so those files
                # are never skipped as "already named"), which meant every
                # needs-review file the batch produced looked brand new and
                # was read, classified and billed for a second time.
                seen_names: set = set(files)
                processed_count = 0

                def _handle(filename: str, label: str = "") -> None:
                    nonlocal processed_count
                    processed_count += 1
                    result_queue.put({
                        "type": "progress",
                        "current": processed_count,
                        "filename": f"{filename}{label}",
                    })
                    result = FileProcessor.process_file(
                        os.path.join(scandocs, filename), config, client_list,
                        batch_id=self.batch_id, reserved=reserved,
                    )
                    seen_names.add(filename)
                    if result.final_name:
                        seen_names.add(result.final_name)
                    result_queue.put({"type": "result", "result": result})
                    if lock is not None:
                        lock.heartbeat()

                for filename in files:
                    if self._stop_event.is_set():
                        result_queue.put({"type": "stopped"})
                        break
                    _handle(filename)

                # Straggler pickup: a Dropbox-synced folder can gain new
                # files while this batch was running (another workstation
                # scanning in, or Dropbox finishing a sync). One extra
                # sweep — never more — for anything that showed up and
                # isn't already accounted for.
                if not self._stop_event.is_set() and not dry_run:
                    try:
                        current = set(self._list_supported_files(scandocs))
                        stragglers = sorted(current - seen_names)
                        stragglers = [
                            f for f in stragglers
                            if not (config["processing"].get("skip_already_processed")
                                    and FileProcessor._already_processed(f, client_list, config.get("naming", {})))
                        ]
                    except Exception as e:
                        logging.warning(f"Straggler sweep could not list folder: {e}")
                        stragglers = []

                    logging.info(f"Straggler sweep: {len(stragglers)} newly-arrived file(s) found.")
                    if stragglers:
                        result_queue.put({
                            "type": "total",
                            "count": processed_count + len(stragglers),
                        })

                    for filename in stragglers:
                        if self._stop_event.is_set():
                            result_queue.put({"type": "stopped"})
                            break
                        _handle(filename, label=" (newly arrived)")

            except Exception as e:
                logging.error(f"Unhandled batch error: {e}", exc_info=True)
                result_queue.put({"type": "error", "message": str(e)})
            finally:
                result_queue.put({"type": "done"})
        finally:
            if lock is not None:
                lock.release()


# ─────────────────────────────────────────────────────────────
# SplashScreen
# ─────────────────────────────────────────────────────────────

class SplashScreen:
    """Fun splash window shown while the main app initialises.
    Shows the logo, an animated spinner, and a stream of dad jokes."""

    DAD_JOKES = [
        ("Why don't scientists trust atoms?",
         "Because they make up everything!"),
        ("I told my wife she was drawing her eyebrows too high.",
         "She looked surprised."),
        ("I'm reading a book about anti-gravity.",
         "It's impossible to put down!"),
        ("Did you hear about the claustrophobic astronaut?",
         "He just needed a little space."),
        ("Why don't eggs tell jokes?",
         "They'd crack each other up!"),
        ("I asked the librarian if they had books about paranoia.",
         "She whispered: 'They're right behind you!'"),
        ("What do you call a fish without eyes?",
         "A fsh."),
        ("Why can't a leopard hide?",
         "Because he's always spotted!"),
        ("I used to hate facial hair…",
         "…but then it grew on me."),
        ("What do you call a factory that makes okay products?",
         "A satisfactory."),
    ]

    def __init__(self, parent: tk.Misc, on_done, primary_color: str = "#1565c0",
                 show_duration_ms: int = 7500):
        self._on_done = on_done
        self._primary = primary_color
        import random as _random
        self._joke_idx = _random.randint(0, len(self.DAD_JOKES) - 1)
        self._spinner_angle = 0
        self._dot_count = 0
        self._done_called = False

        self.win = tk.Toplevel(parent)
        self.win.overrideredirect(True)
        self.win.resizable(False, False)
        self.win.configure(bg="#ffffff")

        W, H = 520, 360
        sw = self.win.winfo_screenwidth()
        sh = self.win.winfo_screenheight()
        self.win.geometry(f"{W}x{H}+{(sw - W)//2}+{(sh - H)//2}")
        self.win.lift()
        self.win.attributes("-topmost", True)

        # DWM rounded window corners (Windows 11)
        try:
            import ctypes
            hwnd = self.win.winfo_id()
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, 33, ctypes.byref(ctypes.c_int(2)), ctypes.sizeof(ctypes.c_int))
        except Exception:
            pass

        self._build(W, H)
        self._animate_spinner()
        self.win.after(400, self._show_question)
        self.win.after(show_duration_ms, self._finish)

    # ── Layout ────────────────────────────────────────────────

    def _build(self, W: int, H: int):
        primary = self._primary

        # Coloured top banner
        banner = tk.Frame(self.win, bg=primary, height=8)
        banner.pack(fill=tk.X, side=tk.TOP)

        # Coloured bottom banner
        tk.Frame(self.win, bg=primary, height=8).pack(fill=tk.X, side=tk.BOTTOM)

        # Logo
        self._logo_img = None
        png_path = os.path.join(ASSETS_DIR, "Speedy Scandocs Logo.png")
        if PILImage is not None and os.path.isfile(png_path):
            try:
                img = PILImage.open(png_path).convert("RGBA")
                bbox = img.getbbox()
                if bbox:
                    img = img.crop(bbox)
                target_h = 90
                ratio = target_h / img.height
                new_w = max(1, int(img.width * ratio))
                img = img.resize((new_w, target_h), PILImage.LANCZOS)
                bg_img = PILImage.new("RGB", (new_w, target_h), (255, 255, 255))
                bg_img.paste(img, mask=img.split()[3])
                self._logo_img = PILImageTk.PhotoImage(bg_img)
            except Exception:
                pass

        if self._logo_img:
            tk.Label(self.win, image=self._logo_img,
                     bg="#ffffff", bd=0).place(relx=0.5, y=70, anchor="center")
        else:
            tk.Label(self.win, text="Speedy Scandocs", bg="#ffffff",
                     fg=primary, font=(APP_FONT, 20, "bold")).place(
                relx=0.5, y=70, anchor="center")

        # Spinner canvas
        self._spin_canvas = tk.Canvas(
            self.win, width=46, height=46, bg="#ffffff", highlightthickness=0)
        self._spin_canvas.place(relx=0.5, y=168, anchor="center")

        # Joke label
        self._joke_var = tk.StringVar(value="")
        self._joke_lbl = tk.Label(
            self.win, textvariable=self._joke_var, bg="#ffffff",
            fg="#555555", font=(APP_FONT, 10, "italic"),
            wraplength=460, justify="center",
        )
        self._joke_lbl.place(relx=0.5, y=240, anchor="center", width=480)

        # Footer hint
        tk.Label(self.win, text="Loading, please wait…", bg="#ffffff",
                 fg="#bbbbbb", font=(APP_FONT, 8)).place(
            relx=0.5, y=320, anchor="center")

    # ── Spinner animation ─────────────────────────────────────

    def _animate_spinner(self):
        if not self.win.winfo_exists():
            return
        c = self._spin_canvas
        c.delete("all")
        a = self._spinner_angle
        p = self._primary
        c.create_arc(3, 3, 43, 43, start=a,       extent=270,
                     style="arc", outline=p,       width=5)
        c.create_arc(3, 3, 43, 43, start=a + 270, extent=90,
                     style="arc", outline="#dddddd", width=5)
        self._spinner_angle = (self._spinner_angle + 14) % 360
        self.win.after(40, self._animate_spinner)

    # ── Dad joke cycle ────────────────────────────────────────

    def _show_question(self):
        if not self.win.winfo_exists():
            return
        joke = self.DAD_JOKES[self._joke_idx % len(self.DAD_JOKES)]
        self._joke_var.set(f"{joke[0]}")
        self._dot_count = 0
        self.win.after(2200, self._animate_dots)

    def _animate_dots(self):
        if not self.win.winfo_exists():
            return
        dots = "  .  " * ((self._dot_count % 3) + 1)
        self._joke_var.set(dots)
        self._dot_count += 1
        if self._dot_count <= 5:
            self.win.after(350, self._animate_dots)
        else:
            self.win.after(200, self._show_answer)

    def _show_answer(self):
        if not self.win.winfo_exists():
            return
        joke = self.DAD_JOKES[self._joke_idx % len(self.DAD_JOKES)]
        self._joke_var.set(f"{joke[1]}")
        self._joke_idx += 1
        self.win.after(2600, self._show_question)

    # ── Close ─────────────────────────────────────────────────

    def _finish(self):
        if self._done_called:
            return
        self._done_called = True
        try:
            if self.win.winfo_exists():
                self.win.destroy()
        except Exception:
            pass
        try:
            self._on_done()
        except Exception as exc:
            logging.warning(f"SplashScreen on_done error: {exc}")


# ─────────────────────────────────────────────────────────────
# ScandocsApp  (tkinter GUI)
# ─────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────
# Auto-update (GitHub Releases)
# ─────────────────────────────────────────────────────────────

def _parse_version(tag: str) -> tuple:
    """'v1.8.0' or '1.8.0' -> (1, 8, 0). Unknown suffixes are dropped.
    Returns (0,) on a tag we can't parse so we never offer a bogus update."""
    if not tag:
        return (0,)
    s = tag.strip().lstrip("vV")
    parts = []
    for chunk in s.split("."):
        num = ""
        for ch in chunk:
            if ch.isdigit():
                num += ch
            else:
                break
        if not num:
            break
        parts.append(int(num))
    return tuple(parts) if parts else (0,)


def _pick_release_asset(release: dict) -> Optional[dict]:
    """Pick the right platform asset from a GitHub release JSON payload.
    Windows -> first .exe, Mac -> first .dmg. Returns asset dict or None."""
    assets = release.get("assets") or []
    if sys.platform == "win32":
        want = ".exe"
    elif sys.platform == "darwin":
        want = ".dmg"
    else:
        return None
    for a in assets:
        name = (a.get("name") or "").lower()
        if name.endswith(want):
            return a
    return None


def fetch_latest_release(timeout: float = 8.0) -> Optional[dict]:
    """Hit the GitHub API for the latest release. Returns parsed JSON or None.
    Silent on network failure — we don't want offline users to see errors."""
    if requests is None:
        return None
    try:
        r = requests.get(
            UPDATE_API_URL,
            headers={"Accept": "application/vnd.github+json"},
            timeout=timeout,
        )
        if r.status_code != 200:
            logging.info(f"Update check: HTTP {r.status_code}")
            return None
        return r.json()
    except Exception as e:
        logging.info(f"Update check failed: {e}")
        return None


def download_file(url: str, dest_path: str,
                  progress_cb=None, cancel_flag=None) -> bool:
    """Stream a file to dest_path. progress_cb(downloaded, total) called
    periodically. cancel_flag is a callable returning True to abort.
    Returns True on success."""
    if requests is None:
        return False
    try:
        with requests.get(url, stream=True, timeout=30) as r:
            r.raise_for_status()
            total = int(r.headers.get("Content-Length") or 0)
            downloaded = 0
            tmp = dest_path + ".part"
            with open(tmp, "wb") as f:
                for chunk in r.iter_content(chunk_size=65536):
                    if cancel_flag and cancel_flag():
                        try: os.remove(tmp)
                        except OSError: pass
                        return False
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        if progress_cb:
                            progress_cb(downloaded, total)
            os.replace(tmp, dest_path)
            return True
    except Exception as e:
        logging.warning(f"Download failed: {e}")
        return False


class ScandocsApp(ttk.Window):

    def report_callback_exception(self, exc, val, tb):
        """Tkinter normally prints callback exceptions to stderr, which is
        invisible when the app is launched without a console — silently
        swallowing bugs in button commands etc. Log them instead."""
        logging.error("Unhandled exception in a UI callback", exc_info=(exc, val, tb))

    def __init__(self):
        super().__init__(themename="litera")
        self.withdraw()               # hidden until splash finishes
        self.title(APP_TITLE)
        self.geometry("1500x1000")
        self.minsize(1100, 800)

        self._apply_app_font()

        self.config_mgr = ConfigManager()
        self.engine = ProcessingEngine()
        self._queue: queue.Queue = queue.Queue()
        self._results: list = []
        self._total_files: int = 0
        self._iid_to_result: dict = {}   # treeview item id → ProcessResult
        self._correction_buttons: dict = {}  # treeview item id → overlaid "Manual Correction" button
        self._audit_updating: bool = False  # prevent recursive checkbox callbacks
        self.fo_dest_var = tk.StringVar()        # file-mode destination folder
        self.s_file_mode_auto_var      = tk.BooleanVar(value=True)
        self.s_file_mode_manual_var    = tk.BooleanVar(value=True)
        self.s_suggest_loc_var         = tk.BooleanVar(value=False)
        self.s_suggest_parent_var      = tk.StringVar(value="")
        self.s_require_high_conf_var   = tk.BooleanVar(value=True)
        self.s_show_manual_tab_var     = tk.BooleanVar(value=False)
        self._all_clients: list = []        # full client list for combo filtering
        self._correction_iid: str = ""      # results_tree row id currently under Manual Correction
        self._pre_correction_height: Optional[int] = None  # window height before growing for the panel
        self._review_selected_file: str = ""  # last file chosen in the legacy Manual Entry list
        self._sort_col: str = ""            # treeview column currently sorted
        self._sort_reverse: bool = False    # ascending=False, descending=True
        self._file_popup = None             # currently open document preview Toplevel
        self._file_popup_canvas = None      # canvas inside the popup for refreshing

        os.makedirs(DEFAULT_REPORTS_FOLDER, exist_ok=True)

        # Show splash FIRST so the user sees something immediately while the
        # heavy UI build runs. Force a redraw so the splash paints before
        # _build_ui blocks the main thread.
        self._splash = SplashScreen(
            self, on_done=self.deiconify, primary_color=_APP_PRIMARY,
            show_duration_ms=3000,
        )
        self.update_idletasks()
        self.update()

        # Defer heavy initialisation to after_idle so the splash's event loop
        # has a chance to render and animate before we block on UI build.
        self.after_idle(self._deferred_init)

    def _deferred_init(self):
        """Run the heavy UI build after the splash has rendered."""
        self._build_ui()
        # Must run AFTER _build_ui — ttkbootstrap builds its per-color button
        # styles lazily the first time a button with that bootstyle is
        # constructed, so an earlier override would be overwritten.
        self._apply_rounded_buttons()
        self._load_settings_to_ui()
        self._refresh_client_list_tab()
        self._refresh_unnamed_count()
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        # Set icon after ttkbootstrap finishes its own setup to prevent it being overridden
        self.after(100, self._set_window_icon)
        self.after(200, self._check_first_run)
        # Auto-update: check on startup if enabled and 24h have passed.
        # Delay past splash so we don't race the initial UI paint.
        self.after(4000, self._maybe_check_for_updates_async)

    def _apply_app_font(self):
        """Point every named tk font, ttk style, and option-db default at
        APP_FONT so widgets built before explicit font= kwargs still pick it up."""
        import tkinter.font as tkfont
        for name in ("TkDefaultFont", "TkTextFont", "TkMenuFont", "TkHeadingFont",
                     "TkCaptionFont", "TkSmallCaptionFont", "TkIconFont",
                     "TkTooltipFont", "TkFixedFont"):
            try:
                tkfont.nametofont(name).configure(family=APP_FONT)
            except tk.TclError:
                pass
        try:
            self.style.configure(".", font=(APP_FONT, 10))
            self.style.configure("Treeview", font=(APP_FONT, 10))
            self.style.configure("Treeview.Heading", font=(APP_FONT, 10, "bold"))
            self.style.configure("TNotebook.Tab", font=(APP_FONT, 10))
            self.style.configure("TButton", font=(APP_FONT, 10))
            self.style.configure("TLabel", font=(APP_FONT, 10))
            self.style.configure("TEntry", font=(APP_FONT, 10))
            self.style.configure("TCombobox", font=(APP_FONT, 10))
            self.style.configure("TCheckbutton", font=(APP_FONT, 10))
            self.style.configure("TRadiobutton", font=(APP_FONT, 10))
            self.style.configure("TLabelframe.Label", font=(APP_FONT, 10, "bold"))
            self.style.configure("TMenubutton", font=(APP_FONT, 10))
        except Exception as e:
            logging.info(f"Could not apply ttk font: {e}")
        # Option db — covers classic tk widgets (tk.Label, tk.Button, Listbox,
        # Menu, Text, Entry) created without an explicit font= kwarg.
        self.option_add("*Font", (APP_FONT, 10))
        self.option_add("*TCombobox*Listbox.font", (APP_FONT, 10))

    def _apply_rounded_buttons(self):
        """Replace ttk button layouts with 9-slice rounded-rectangle images
        for every bootstyle color the app uses. Preserves the ttkbootstrap
        color scheme — only changes corner shape."""
        if PILImage is None:
            return
        try:
            from PIL import ImageDraw, ImageTk
        except ImportError:
            return

        radius = 10
        w, h = 220, 46
        theme = self.style.colors
        color_map = {}
        for cname in ("primary", "secondary", "success", "danger",
                      "warning", "info", "dark", "light"):
            try:
                color_map[cname] = getattr(theme, cname)
            except AttributeError:
                pass

        self._rounded_btn_imgs = []  # keep PhotoImage refs alive

        def _shade(hex_c: str, amount: float) -> str:
            hex_c = hex_c.lstrip("#")
            r = int(hex_c[0:2], 16); g = int(hex_c[2:4], 16); b = int(hex_c[4:6], 16)
            if amount >= 0:
                r = max(0, int(r * (1 - amount)))
                g = max(0, int(g * (1 - amount)))
                b = max(0, int(b * (1 - amount)))
            else:
                a = -amount
                r = min(255, int(r + (255 - r) * a))
                g = min(255, int(g + (255 - g) * a))
                b = min(255, int(b + (255 - b) * a))
            return f"#{r:02x}{g:02x}{b:02x}"

        def _make(fill=None, border=None, bw=0):
            img = PILImage.new("RGBA", (w, h), (0, 0, 0, 0))
            d = ImageDraw.Draw(img)
            kw = {}
            if fill: kw["fill"] = fill
            if border and bw:
                kw["outline"] = border
                kw["width"] = bw
            d.rounded_rectangle((0, 0, w - 1, h - 1), radius=radius, **kw)
            ph = ImageTk.PhotoImage(img)
            self._rounded_btn_imgs.append(ph)
            return ph

        def _register(style_name, el_name, normal, active, pressed, disabled,
                      fg, hover_fg=None):
            try:
                self.style.element_create(
                    el_name, "image", normal,
                    ("pressed", pressed),
                    ("active", active),
                    ("disabled", disabled),
                    border=radius, sticky="nsew",
                )
            except tk.TclError:
                return  # element already exists — safe to ignore
            self.style.layout(style_name, [
                (el_name, {"sticky": "nsew", "children": [
                    ("Button.padding", {"sticky": "nsew", "children": [
                        ("Button.label", {"sticky": "nsew"}),
                    ]}),
                ]}),
            ])
            self.style.configure(style_name, foreground=fg,
                                 font=(APP_FONT, 10),
                                 padding=(14, 7),
                                 borderwidth=0, focuscolor="",
                                 relief="flat")
            if hover_fg:
                self.style.map(style_name, foreground=[
                    ("active", hover_fg), ("pressed", hover_fg),
                ])

        # Solid + outline per color
        for name, color in color_map.items():
            fg_on_color = "#ffffff" if name != "light" else "#212529"
            _register(
                f"{name}.TButton",
                f"Rounded.{name}.button",
                _make(fill=color),
                _make(fill=_shade(color, 0.08)),
                _make(fill=_shade(color, 0.18)),
                _make(fill=_shade(color, -0.5)),
                fg=fg_on_color,
            )
            _register(
                f"{name}.Outline.TButton",
                f"Rounded.{name}.Outline.button",
                _make(fill="#ffffff", border=color, bw=2),
                _make(fill=color, border=color, bw=2),
                _make(fill=_shade(color, 0.15), border=_shade(color, 0.15), bw=2),
                _make(fill="#ffffff", border=_shade(color, -0.5), bw=2),
                fg=color,
                hover_fg="#ffffff" if name != "light" else "#212529",
            )

        # Default (un-bootstyled) button
        default_bg = color_map.get("light", "#f1f3f5")
        border_c = "#ced4da"
        _register(
            "TButton",
            "Rounded.default.button",
            _make(fill=default_bg, border=border_c, bw=1),
            _make(fill=_shade(default_bg, 0.05), border=border_c, bw=1),
            _make(fill=_shade(default_bg, 0.12), border=border_c, bw=1),
            _make(fill=default_bg, border=_shade(border_c, -0.3), bw=1),
            fg="#212529",
        )

    # ── UI Construction ───────────────────────────────────────

    def _set_window_icon(self):
        """Set the title bar and taskbar icon to the GDJ logo."""
        png_path = os.path.join(ASSETS_DIR, "GDJ Logo.png")
        ico_path = os.path.join(ASSETS_DIR, "GDJ Logo.ico")

        # Generate the .ico from PNG if it doesn't exist yet (Windows)
        if not os.path.isfile(ico_path) and PILImage is not None and os.path.isfile(png_path):
            try:
                src = PILImage.open(png_path).convert("RGBA")
                sizes = [256, 128, 64, 48, 32, 16]
                frames = [src.resize((s, s), PILImage.LANCZOS) for s in sizes]
                frames[0].save(ico_path, format="ICO",
                               append_images=frames[1:],
                               sizes=[(s, s) for s in sizes])
            except Exception as e:
                logging.warning(f"Could not create icon file: {e}")

        if sys.platform == "darwin":
            # macOS: use iconphoto with a PNG — iconbitmap(.ico) does not work
            if PILImage is not None and os.path.isfile(png_path):
                try:
                    img = PILImage.open(png_path).convert("RGBA")
                    self._app_icon_photo = PILImageTk.PhotoImage(img)
                    self.iconphoto(True, self._app_icon_photo)
                except Exception as e:
                    logging.warning(f"iconphoto (macOS) failed: {e}")
        else:
            # Windows/Linux: use iconbitmap with .ico
            if os.path.isfile(ico_path):
                try:
                    self.iconbitmap(ico_path)
                except Exception as e:
                    logging.warning(f"iconbitmap failed: {e}")

    def _build_ui(self):
        self._build_header()
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))
        self._build_process_tab()
        self._build_review_tab()   # legacy Manual Entry tab — attached/detached via Settings toggle
        self._build_clients_tab()
        self._build_document_types_tab()
        self._build_providers_tab()
        self._build_suggestions_tab()
        self._build_settings_tab()
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)
        # The Suggestions tab's label carries a pending-count badge, e.g.
        # "  Suggestions (3)  " — refresh it once at startup too, not only
        # on tab-change, so a badge is visible before it's ever opened.
        self.after(500, self._refresh_suggestions_badge)

    def _on_tab_changed(self, _evt=None):
        # Refresh each tab when it becomes active so it always reflects the
        # current scandocs folder / client list on disk.
        try:
            current = self.notebook.nametowidget(self.notebook.select())
        except Exception:
            return
        if getattr(self, "_process_tab", None) is current:
            self._refresh_unnamed_count()
        elif getattr(self, "_review_tab", None) is current:
            self._refresh_review_tab()
        elif getattr(self, "_clients_tab", None) is current:
            self._refresh_client_list_tab()
        elif getattr(self, "_doc_types_tab", None) is current:
            self._refresh_document_types_tab()
        elif getattr(self, "_providers_tab", None) is current:
            self._refresh_providers_tab()
        elif getattr(self, "_suggestions_tab", None) is current:
            self._refresh_suggestions_tab()

    def _refresh_unnamed_count(self):
        """Count files in the scandocs folder that don't yet match the
        'LAST, First - Subject.ext' format, and show it on the Auto-Process tab."""
        scandocs = self.config_mgr.config["paths"]["scandocs_folder"]
        if not scandocs or not os.path.isdir(scandocs):
            self.unnamed_count_var.set("Scandocs folder not configured or not found.")
            return
        client_list_path = self.config_mgr.config["paths"]["client_list_file"]
        client_list = ClientListManager.load(client_list_path)
        try:
            unnamed = [
                f for f in os.listdir(scandocs)
                if os.path.isfile(os.path.join(scandocs, f))
                and os.path.splitext(f)[1].lower() in SUPPORTED_EXTENSIONS
                and not FileProcessor._already_processed(f, client_list, self.config_mgr.config.get("naming", {}))
            ]
        except OSError as e:
            self.unnamed_count_var.set(f"Could not read scandocs folder: {e}")
            return
        n = len(unnamed)
        if n == 0:
            self.unnamed_count_var.set("No unnamed files in scandocs folder.")
        elif n == 1:
            self.unnamed_count_var.set("1 unnamed file ready to process.")
        else:
            self.unnamed_count_var.set(f"{n} unnamed files ready to process.")

    def _build_header(self):
        """Header bar: Speedy Scandocs logo PNG, right-aligned."""
        header = tk.Frame(self, bg="#ffffff", height=80)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        # Hairline separator beneath the header (palette-coloured)
        self._header_sep = tk.Frame(self, bg="#dee2e6", height=2)
        self._header_sep.pack(fill=tk.X)

        self._header_logo = None
        png_path = os.path.join(ASSETS_DIR, "Speedy Scandocs Logo.png")
        if PILImage is not None and os.path.isfile(png_path):
            try:
                img = PILImage.open(png_path).convert("RGBA")
                # Trim surrounding whitespace so the logo fills the space cleanly
                bbox = img.getbbox()
                if bbox:
                    img = img.crop(bbox)
                # Scale to fit inside the header with a small top/bottom margin
                target_h = 64
                ratio = target_h / img.height
                new_w = max(1, int(img.width * ratio))
                img = img.resize((new_w, target_h), PILImage.LANCZOS)
                # Composite onto white so RGBA looks clean on the white header
                bg = PILImage.new("RGB", (new_w, target_h), (255, 255, 255))
                bg.paste(img, mask=img.split()[3])
                self._header_logo = PILImageTk.PhotoImage(bg)
            except Exception as e:
                logging.warning(f"Could not load header logo: {e}")

        if self._header_logo:
            lbl = tk.Label(header, image=self._header_logo, bg="#ffffff", bd=0)
            lbl.place(relx=1.0, rely=0.5, anchor="e", x=-16)
        else:
            # Fallback text if image unavailable
            tk.Label(header, text="Speedy Scandocs", bg="#ffffff",
                     fg="#212529", font=(APP_FONT, 15, "bold")).place(
                relx=1.0, rely=0.5, anchor="e", x=-16)

    # ── Tab 1: Process ────────────────────────────────────────

    def _build_process_tab(self):
        self.style.configure("AutoTab.TFrame", background="#e3f2fd")
        tab = ttk.Frame(self.notebook, style="AutoTab.TFrame")
        self._process_tab = tab
        self._process_tab_frame = tab
        self.notebook.add(tab, text="  Process  ")
        # Accent bar
        tk.Frame(tab, bg=_APP_PRIMARY, height=6).pack(fill=tk.X)

        # Preview-mode banner — hidden unless automation.dry_run is on.
        # Not packed here; _apply_dry_run_banner() shows/hides it (with
        # before=btn_row) so a preview run can never be mistaken for a real
        # one, on top of the window-title indicator _start_processing sets.
        self._preview_banner = tk.Frame(tab, bg="#fff3cd")
        tk.Label(
            self._preview_banner,
            text="⚠  PREVIEW MODE — Settings has \"Preview mode\" turned on. "
                 "No files will actually be renamed or moved.",
            bg="#fff3cd", fg="#7a5b00", font=(APP_FONT, 10, "bold"),
        ).pack(pady=5)

        # Button row
        btn_row = ttk.Frame(tab)
        btn_row.pack(fill=tk.X, padx=10, pady=(10, 4))
        self._process_btn_row = btn_row
        self.btn_process = ttk.Button(
            btn_row, text="Auto-Process Documents", command=self._start_processing,
            bootstyle="primary",
        )
        self.btn_process.pack(side=tk.LEFT, padx=(0, 6))
        self.btn_stop = ttk.Button(
            btn_row, text="Stop", command=self._stop_processing,
            state=tk.DISABLED, bootstyle="danger-outline",
        )
        self.btn_stop.pack(side=tk.LEFT)
        self.btn_undo_batch = ttk.Button(
            btn_row, text="Undo Last Rename Batch…", command=self._show_undo_dialog,
            bootstyle="warning-outline",
        )
        self.btn_undo_batch.pack(side=tk.LEFT, padx=(6, 0))
        right_frame = ttk.Frame(btn_row)
        right_frame.pack(side=tk.RIGHT)
        self.btn_open_report = ttk.Button(
            right_frame, text="Open Report", command=self._open_report,
            bootstyle="primary-outline",
        )
        self.btn_open_report.pack(anchor="e")
        self.btn_submit_audit = ttk.Button(
            right_frame, text="Submit Audit", command=self._submit_audit,
            bootstyle="primary",
        )
        # btn_submit_audit visibility is toggled by _apply_audit_mode
        ttk.Label(
            right_frame, text="Reports Saved Automatically",
            font=(APP_FONT, 7), foreground="gray",
        ).pack(anchor="e")

        # Unnamed-files indicator (refreshes when this tab is selected)
        self.unnamed_count_var = tk.StringVar(value="")
        ttk.Label(
            tab, textvariable=self.unnamed_count_var,
            font=(APP_FONT, 10, "bold"), foreground="#1565c0",
            anchor="w",
        ).pack(fill=tk.X, padx=10, pady=(2, 0))

        # Progress
        prog_frame = ttk.Frame(tab)
        prog_frame.pack(fill=tk.X, padx=10, pady=(0, 4))
        self.progress_var = tk.DoubleVar(value=0)
        ttk.Progressbar(
            prog_frame, variable=self.progress_var, maximum=100
        ).pack(fill=tk.X)
        self.status_var = tk.StringVar(
            value="Ready — configure Settings then press Auto-Process Documents."
        )
        ttk.Label(prog_frame, textvariable=self.status_var, anchor="w").pack(
            fill=tk.X, pady=(2, 0)
        )

        # Results table
        cols = ("audited", "original", "new_name", "status", "client",
                "new_location", "correction")
        col_cfg = {
            "audited":      ("✓",              32),
            "original":     ("Original File",  180),
            "new_name":     ("New Name",        240),
            "status":       ("Status",           75),
            "client":       ("Client",          170),
            "new_location": ("New Location",    130),
            "correction":   ("",                 160),
        }
        tree_frame = ttk.Frame(tab)
        self._tree_frame = tree_frame   # stable pack anchor — always packed, unlike audit_panel
        # tree_frame is packed AFTER the bottom panels so it fills remaining space

        self.results_tree = ttk.Treeview(
            tree_frame, columns=cols, show="headings", selectmode="extended"
        )
        # Columns that support click-to-sort
        _sortable = {"original", "new_name", "status", "client"}
        for col in cols:
            label, width = col_cfg[col]
            if col in _sortable:
                self.results_tree.heading(
                    col, text=label,
                    command=lambda c=col: self._sort_treeview(c))
            else:
                self.results_tree.heading(col, text=label)
            self.results_tree.column(
                col, width=width,
                minwidth=32 if col == "audited" else 50,
                anchor="center" if col in ("audited", "correction") else "w",
            )
        self.results_tree.tag_configure("renamed",      background="#d4edda")
        self.results_tree.tag_configure("needs_review", background="#fff3cd")
        self.results_tree.tag_configure("error",        background="#f8d7da")
        self.results_tree.tag_configure("skipped",      background="#e2e3e5")
        self.results_tree.tag_configure("audited",      background="#fddcb0")
        self.results_tree.tag_configure("moved",        background="#cce5ff")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",   command=self.results_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.results_tree.xview)

        def _yscroll_set(*args):
            vsb.set(*args)
            self._reposition_all_correction_buttons()

        def _xscroll_set(*args):
            hsb.set(*args)
            self._reposition_all_correction_buttons()

        self.results_tree.configure(yscrollcommand=_yscroll_set, xscrollcommand=_xscroll_set)
        vsb.pack(side=tk.RIGHT,  fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self.results_tree.pack(fill=tk.BOTH, expand=True)
        self.results_tree.bind("<<TreeviewSelect>>", self._on_result_select)
        self.results_tree.bind("<space>",           self._on_tree_return)
        self.results_tree.bind("<Double-Button-1>", self._on_tree_double_click)
        self.results_tree.bind("<Left>",   self._audit_prev)
        self.results_tree.bind("<Right>",  self._audit_next)
        self.results_tree.bind("<Configure>", lambda e: self._reposition_all_correction_buttons())
        # Mousewheel scroll on the treeview
        _tree_scroll = lambda e: self.results_tree.yview_scroll(
            -1 * (e.delta // 120) if e.delta else (-1 if e.num == 4 else 1), "units")
        self._bind_mousewheel(self.results_tree, _tree_scroll)

        # Pack tree now — bottom panels pack after (side=BOTTOM) so they always show
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(4, 0))

        # ── Audit panel ──────────────────────────────────────
        self.audit_panel = ttk.LabelFrame(tab, text="Audit Mode")
        self.audit_panel.pack(fill=tk.X, padx=10, pady=(6, 0))
        audit_outer = self.audit_panel

        # Row 1: filename + Open File button
        file_row = ttk.Frame(audit_outer)
        file_row.pack(fill=tk.X, padx=8, pady=(6, 2))
        self.audit_file_label = ttk.Label(
            file_row, text="Select a row above to audit it.",
            foreground="gray", anchor="w",
        )
        self.audit_file_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.audit_open_btn = ttk.Button(
            file_row, text="Open File", bootstyle="dark-outline",
            command=self._audit_open_file, state=tk.DISABLED,
        )
        self.audit_open_btn.pack(side=tk.RIGHT, padx=(8, 0))
        self.audit_next_btn = ttk.Button(
            file_row, text="Next →", bootstyle="dark-outline",
            command=self._audit_next, state=tk.DISABLED,
        )
        self.audit_next_btn.pack(side=tk.RIGHT, padx=(4, 0))
        self.audit_prev_btn = ttk.Button(
            file_row, text="← Prev", bootstyle="dark-outline",
            command=self._audit_prev, state=tk.DISABLED,
        )
        self.audit_prev_btn.pack(side=tk.RIGHT, padx=(4, 0))

        # Checkboxes — two rows so all are always visible
        self.audit_correct_var       = tk.BooleanVar()
        self.audit_wrong_client_var  = tk.BooleanVar()
        self.audit_bad_desc_var      = tk.BooleanVar()
        self.audit_failed_client_var = tk.BooleanVar()
        self.audit_should_flag_var   = tk.BooleanVar()

        def _make_chk(parent, text, var, flag, fg="#212529"):
            return tk.Checkbutton(
                parent, text=text, variable=var,
                command=lambda: self._on_audit_check(flag),
                fg=fg, bg="#f8f9fa",
                activeforeground=fg, activebackground="#f8f9fa",
                selectcolor="#ffffff",
                disabledforeground="#adb5bd",
                font=(APP_FONT, 11),
                padx=8, pady=6,
                bd=2,
                state=tk.DISABLED,
            )

        # Row 2a: positive check
        chk_row1 = ttk.Frame(audit_outer)
        chk_row1.pack(fill=tk.X, padx=8, pady=(2, 0))
        self.audit_correct_chk = _make_chk(
            chk_row1, "✓  Correct", self.audit_correct_var, "correct", fg="#1a6e31")
        self.audit_correct_chk.pack(side=tk.LEFT)

        # Row 2b: problem flags
        chk_row2 = ttk.Frame(audit_outer)
        chk_row2.pack(fill=tk.X, padx=8, pady=(2, 8))

        self.audit_wrong_client_chk = _make_chk(
            chk_row2, "Wrong client name", self.audit_wrong_client_var, "wrong_client")
        self.audit_wrong_client_chk.pack(side=tk.LEFT, padx=(0, 24))

        self.audit_bad_desc_chk = _make_chk(
            chk_row2, "Bad description", self.audit_bad_desc_var, "bad_description")
        self.audit_bad_desc_chk.pack(side=tk.LEFT, padx=(0, 24))

        self.audit_failed_client_chk = _make_chk(
            chk_row2, "Failed to identify client", self.audit_failed_client_var, "failed_client")
        self.audit_failed_client_chk.pack(side=tk.LEFT, padx=(0, 24))

        self.audit_should_flag_chk = _make_chk(
            chk_row2, "Should have been flagged for review", self.audit_should_flag_var, "should_review")
        self.audit_should_flag_chk.pack(side=tk.LEFT)

        # Row 2c: rename hint — shown when any problem flag is active
        self.audit_rename_hint_var = tk.StringVar(value="")
        self.audit_rename_hint_lbl = tk.Label(
            audit_outer,
            textvariable=self.audit_rename_hint_var,
            font=(APP_FONT, 8, "italic"),
            fg="#777777",
            bg="#f8f9fa",
            anchor="w",
            wraplength=0,   # no wrapping — single line
        )
        self.audit_rename_hint_lbl.pack(fill=tk.X, padx=8, pady=(0, 6))

        # ── File Operations panel ─────────────────────────────
        self.file_ops_panel = ttk.LabelFrame(tab, text="File Mode — Move Files")
        self.file_ops_panel.pack(fill=tk.X, padx=10, pady=(6, 10))

        # Row 1: destination folder + browse + apply-to-selected
        dest_row = ttk.Frame(self.file_ops_panel)
        dest_row.pack(fill=tk.X, padx=8, pady=(8, 4))
        ttk.Label(dest_row, text="Destination Folder:").pack(side=tk.LEFT)
        ttk.Entry(dest_row, textvariable=self.fo_dest_var, width=40).pack(
            side=tk.LEFT, padx=(8, 4), fill=tk.X, expand=True)
        ttk.Button(
            dest_row, text="Browse",
            command=self._fo_browse_dest,
            bootstyle="dark-outline",
        ).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(
            dest_row, text="Apply to Selected",
            command=self._fo_apply_to_selected,
            bootstyle="dark-outline",
        ).pack(side=tk.LEFT)

        # Row 2: commit button + status
        action_row = ttk.Frame(self.file_ops_panel)
        action_row.pack(fill=tk.X, padx=8, pady=(0, 8))
        ttk.Button(
            action_row, text="Move Files",
            command=self._fo_move_all,
            bootstyle="primary",
        ).pack(side=tk.LEFT)
        self.fo_status_var = tk.StringVar(value="")
        ttk.Label(action_row, textvariable=self.fo_status_var,
                  foreground="gray").pack(side=tk.LEFT, padx=14)

        # ── Manual Correction panel (hidden until a row's "Manual Correction"
        #    cell is clicked) ─────────────────────────────────
        self.correction_panel = ttk.LabelFrame(tab, text="Manual Correction")
        # Not packed here — _open_manual_correction() shows it, _hide_manual_correction() hides it.

        corr = self.correction_panel

        corr_file_row = ttk.Frame(corr)
        corr_file_row.pack(fill=tk.X, padx=8, pady=(8, 2))
        self.corr_file_var = tk.StringVar(value="")
        ttk.Label(corr_file_row, textvariable=self.corr_file_var,
                  font=(APP_FONT, 10, "bold")).pack(side=tk.LEFT)

        # Large, unmistakable preview button — spacebar types a space in this
        # panel's fields instead of toggling the viewer, so this button is the
        # clear way to see the document while correcting it.
        ttk.Button(
            corr, text="🔍  Open File Preview",
            command=self._corr_open_preview,
            bootstyle="info", padding=(16, 12),
        ).pack(fill=tk.X, padx=8, pady=(2, 10))

        corr_fields = ttk.Frame(corr)
        corr_fields.pack(fill=tk.X, padx=8, pady=(0, 4))
        corr_fields.columnconfigure(1, weight=1)

        # Row 0: Client typing entry
        ttk.Label(corr_fields, text="Client:").grid(
            row=0, column=0, padx=(0, 8), pady=(0, 2), sticky="nw"
        )
        self.corr_client_var = tk.StringVar()
        corr_client_entry = ttk.Entry(
            corr_fields, textvariable=self.corr_client_var, width=40
        )
        corr_client_entry.grid(row=0, column=1, pady=(0, 2), sticky="w")
        corr_client_entry.bind("<KeyRelease>", self._filter_client_combo)
        corr_client_entry.bind("<Return>", lambda e: self._corr_next())

        # Row 1: Always-visible scrollable client list — shrinks as user types
        client_lb_frame = ttk.Frame(corr_fields)
        client_lb_frame.grid(row=1, column=1, pady=(0, 4), sticky="ew")
        self.corr_client_listbox = tk.Listbox(
            client_lb_frame, height=6,
            font=(APP_FONT, 9),
            selectmode=tk.SINGLE,
            exportselection=False,
            activestyle="none",
        )
        cl_lb_vsb = ttk.Scrollbar(client_lb_frame, orient="vertical",
                                   command=self.corr_client_listbox.yview)
        self.corr_client_listbox.configure(yscrollcommand=cl_lb_vsb.set)
        cl_lb_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.corr_client_listbox.pack(fill=tk.X, expand=True)
        self.corr_client_listbox.bind(
            "<<ListboxSelect>>", self._on_client_listbox_select)
        self.corr_client_listbox.bind(
            "<Return>", lambda e: self._corr_next())

        # Row 2: Subject
        ttk.Label(corr_fields, text="Subject:").grid(
            row=2, column=0, padx=(0, 8), pady=6, sticky="w"
        )
        self.corr_subject_var = tk.StringVar()
        corr_subject_entry = ttk.Entry(
            corr_fields, textvariable=self.corr_subject_var, width=40)
        corr_subject_entry.grid(row=2, column=1, pady=6, sticky="w")
        corr_subject_entry.bind("<Return>", lambda e: self._corr_next())

        # Row 3: Destination folder (shown only when File Mode is on for Manual Correction)
        self.corr_dest_label = ttk.Label(corr_fields, text="Move to folder:")
        self.corr_dest_label.grid(row=3, column=0, padx=(0, 8), pady=6, sticky="w")
        self.corr_dest_label.grid_remove()
        corr_dest_inner = ttk.Frame(corr_fields)
        corr_dest_inner.grid(row=3, column=1, pady=6, sticky="ew")
        corr_dest_inner.grid_remove()
        ttk.Entry(corr_dest_inner, textvariable=self.fo_dest_var).pack(
            side=tk.LEFT, padx=(0, 4), fill=tk.X, expand=True)
        ttk.Button(corr_dest_inner, text="Browse",
                   command=self._fo_browse_dest,
                   bootstyle="dark-outline").pack(side=tk.LEFT)
        self._corr_dest_widgets = [self.corr_dest_label, corr_dest_inner]

        # Row 4: Close / Next buttons
        corr_btn_row = ttk.Frame(corr)
        corr_btn_row.pack(fill=tk.X, padx=8, pady=(4, 10))
        ttk.Button(corr_btn_row, text="Next →",
                   command=self._corr_next,
                   bootstyle="primary").pack(side=tk.RIGHT)
        ttk.Button(corr_btn_row, text="Close",
                   command=self._corr_close,
                   bootstyle="dark-outline").pack(side=tk.RIGHT, padx=(0, 6))

    # ── Legacy Manual Entry tab (Settings: "Show Manual Entry tab") ──
    # Built once at startup but only attached to the notebook (via
    # _apply_manual_tab_visibility) when the Settings toggle is on.

    def _build_review_tab(self):
        tab = ttk.Frame(self.notebook)
        self._review_tab = tab
        self._review_tab_frame = tab
        tk.Frame(tab, bg=_APP_PRIMARY, height=5).pack(fill=tk.X)

        top = ttk.Frame(tab)
        top.pack(fill=tk.X, padx=10, pady=(10, 0))
        self.review_count_var = tk.StringVar(value="Files awaiting review: 0")
        ttk.Label(top, textvariable=self.review_count_var,
                  font=(APP_FONT, 10, "bold")).pack(side=tk.LEFT)
        ttk.Button(top, text="Refresh", command=self._refresh_review_tab,
                   bootstyle="dark-outline").pack(side=tk.RIGHT)
        ttk.Button(top, text="Next →", command=self._review_next,
                   bootstyle="dark-outline").pack(side=tk.RIGHT, padx=(4, 6))
        ttk.Button(top, text="← Prev", command=self._review_prev,
                   bootstyle="dark-outline").pack(side=tk.RIGHT, padx=(0, 4))

        # File list
        list_lf = ttk.LabelFrame(tab, text="Files Awaiting Review")
        list_lf.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)
        self.review_listbox = tk.Listbox(list_lf, selectmode=tk.SINGLE,
                                         font=(APP_FONT, 10), exportselection=False)
        rv_sb = ttk.Scrollbar(list_lf, orient="vertical", command=self.review_listbox.yview)
        self.review_listbox.configure(yscrollcommand=rv_sb.set)
        rv_sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.review_listbox.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        self.review_listbox.bind("<<ListboxSelect>>", self._on_review_select)
        self.review_listbox.bind("<space>",           self._on_review_return)
        self.review_listbox.bind("<Double-Button-1>", self._on_review_double_click)
        self.review_listbox.bind("<Left>",   self._review_prev)
        self.review_listbox.bind("<Right>",  self._review_next)
        # Up/Down already navigate the Listbox natively and fire <<ListboxSelect>>.
        _rv_scroll = lambda e: self.review_listbox.yview_scroll(
            -1 * (e.delta // 120) if e.delta else (-1 if e.num == 4 else 1), "units")
        self._bind_mousewheel(self.review_listbox, _rv_scroll)

        # Assignment panel
        assign_lf = ttk.LabelFrame(tab, text="Assign Selected File")
        assign_lf.pack(fill=tk.X, padx=10, pady=(0, 10))
        assign_lf.columnconfigure(1, weight=1)

        ttk.Label(assign_lf, text="Open file:").grid(
            row=0, column=0, padx=8, pady=6, sticky="w"
        )
        ttk.Button(assign_lf, text="Open in Viewer",
                   command=self._open_review_file,
                   bootstyle="dark-outline").grid(
            row=0, column=1, padx=4, pady=6, sticky="w"
        )

        # Row 1: Client typing entry
        ttk.Label(assign_lf, text="Client:").grid(
            row=1, column=0, padx=8, pady=(6, 2), sticky="nw"
        )
        self.review_client_var = tk.StringVar()
        review_client_entry = ttk.Entry(
            assign_lf, textvariable=self.review_client_var, width=40
        )
        review_client_entry.grid(row=1, column=1, padx=4, pady=(6, 2), sticky="w")
        review_client_entry.bind("<KeyRelease>", self._filter_review_client_combo)
        review_client_entry.bind("<Return>", lambda e: self._assign_review_file())

        # Row 2: Always-visible scrollable client list — shrinks as user types
        client_lb_frame = ttk.Frame(assign_lf)
        client_lb_frame.grid(row=2, column=1, padx=4, pady=(0, 4), sticky="ew")
        self.review_client_listbox = tk.Listbox(
            client_lb_frame, height=7,
            font=(APP_FONT, 9),
            selectmode=tk.SINGLE,
            exportselection=False,
            activestyle="none",
        )
        cl_lb_vsb = ttk.Scrollbar(client_lb_frame, orient="vertical",
                                   command=self.review_client_listbox.yview)
        self.review_client_listbox.configure(yscrollcommand=cl_lb_vsb.set)
        cl_lb_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.review_client_listbox.pack(fill=tk.X, expand=True)
        self.review_client_listbox.bind(
            "<<ListboxSelect>>", self._on_review_client_listbox_select)
        self.review_client_listbox.bind(
            "<Return>", lambda e: self._assign_review_file())

        # Row 3: Subject
        ttk.Label(assign_lf, text="Subject:").grid(
            row=3, column=0, padx=8, pady=6, sticky="w"
        )
        self.review_subject_var = tk.StringVar()
        review_subject_entry = ttk.Entry(
            assign_lf, textvariable=self.review_subject_var, width=40)
        review_subject_entry.grid(row=3, column=1, padx=4, pady=6, sticky="w")
        review_subject_entry.bind("<Return>", lambda e: self._assign_review_file())

        # Row 4: Destination folder (shown only when File Mode is on)
        self.review_dest_label = ttk.Label(assign_lf, text="Move to folder:")
        self.review_dest_label.grid(row=4, column=0, padx=8, pady=6, sticky="w")
        self.review_dest_label.grid_remove()
        review_dest_inner = ttk.Frame(assign_lf)
        review_dest_inner.grid(row=4, column=1, padx=4, pady=6, sticky="ew")
        review_dest_inner.grid_remove()
        ttk.Entry(review_dest_inner, textvariable=self.fo_dest_var).pack(
            side=tk.LEFT, padx=(0, 4), fill=tk.X, expand=True)
        ttk.Button(review_dest_inner, text="Browse",
                   command=self._fo_browse_dest,
                   bootstyle="dark-outline").pack(side=tk.LEFT)
        self._review_dest_widgets = [self.review_dest_label, review_dest_inner]

        # Row 5: Submit button
        _btn_assign = ttk.Button(assign_lf, text="Assign, Rename, & Move",
                   command=self._assign_review_file,
                   bootstyle="primary")
        _btn_assign.grid(
            row=5, column=1, padx=4, pady=(2, 8), sticky="w"
        )

    # ── Tab 2: Client List ────────────────────────────────────

    def _build_clients_tab(self):
        tab = ttk.Frame(self.notebook)
        self._clients_tab = tab
        self.notebook.add(tab, text="  Client List  ")

        ttk.Label(
            tab,
            text="Format: LAST, First   (e.g.  GARCIA, Maria  |  REYES, Carlos A)",
            foreground="gray",
        ).pack(anchor="w", padx=10, pady=(10, 2))

        # List + scrollbar
        list_frame = ttk.Frame(tab)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)
        self.client_listbox = tk.Listbox(
            list_frame, selectmode=tk.SINGLE, font=(APP_FONT, 10)
        )
        cl_sb = ttk.Scrollbar(list_frame, orient="vertical",
                               command=self.client_listbox.yview)
        self.client_listbox.configure(yscrollcommand=cl_sb.set)
        cl_sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.client_listbox.pack(fill=tk.BOTH, expand=True)
        _cl_scroll = lambda e: self.client_listbox.yview_scroll(
            -1 * (e.delta // 120) if e.delta else (-1 if e.num == 4 else 1), "units")
        self._bind_mousewheel(self.client_listbox, _cl_scroll)

        # Add / Remove
        edit_frame = ttk.Frame(tab)
        edit_frame.pack(fill=tk.X, padx=10, pady=(0, 2))
        self.new_client_var = tk.StringVar()
        entry = ttk.Entry(edit_frame, textvariable=self.new_client_var, width=32)
        entry.pack(side=tk.LEFT, padx=(0, 6))
        entry.bind("<Return>", lambda _: self._add_client())
        entry.bind("<KeyRelease>", lambda _: self._on_client_entry_key())
        _btn_add_cl = ttk.Button(edit_frame, text="Add Client",
                   command=self._add_client,
                   bootstyle="primary-outline")
        _btn_add_cl.pack(side=tk.LEFT, padx=(0, 6))
        _btn_rm_cl = ttk.Button(edit_frame, text="Remove Selected",
                   command=self._remove_client,
                   bootstyle="danger-outline")
        _btn_rm_cl.pack(side=tk.LEFT)

        # Dynamic "Add X to list" suggestion — shown when typed name is not yet in the list
        self._add_suggestion_lbl = tk.Label(
            tab, text="", fg="#1565c0", bg="#ffffff",
            font=(APP_FONT, 9, "underline"), cursor="hand2", anchor="w",
        )
        self._add_suggestion_lbl.pack(fill=tk.X, padx=12, pady=(0, 2))
        self._add_suggestion_lbl.bind("<Button-1>", lambda _: self._add_client())

        # Save row
        save_frame = ttk.Frame(tab)
        save_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        _btn_save_cl = ttk.Button(save_frame, text="Save Client List",
                   command=self._save_client_list,
                   bootstyle="primary")
        _btn_save_cl.pack(side=tk.LEFT)
        self.client_status_var = tk.StringVar(value="")
        ttk.Label(save_frame, textvariable=self.client_status_var,
                  foreground="green").pack(side=tk.LEFT, padx=10)

    # ── Tab: Document Types ────────────────────────────────────

    def _build_document_types_tab(self):
        tab = ttk.Frame(self.notebook)
        self._doc_types_tab = tab
        self.notebook.add(tab, text="  Document Types  ")

        ttk.Label(
            tab,
            text="Format on disk: Canonical Name | alias1 | alias2 …  "
                 "(this list edits canonical names only — add or edit aliases directly "
                 "in document_types.txt). Used when Settings → Understanding Documents → "
                 "'Use the document type list' is on.",
            foreground="gray", wraplength=760, justify="left",
        ).pack(anchor="w", padx=10, pady=(10, 2))

        list_frame = ttk.Frame(tab)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)
        self.doc_type_listbox = tk.Listbox(
            list_frame, selectmode=tk.SINGLE, font=(APP_FONT, 10)
        )
        dt_sb = ttk.Scrollbar(list_frame, orient="vertical",
                               command=self.doc_type_listbox.yview)
        self.doc_type_listbox.configure(yscrollcommand=dt_sb.set)
        dt_sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.doc_type_listbox.pack(fill=tk.BOTH, expand=True)
        _dt_scroll = lambda e: self.doc_type_listbox.yview_scroll(
            -1 * (e.delta // 120) if e.delta else (-1 if e.num == 4 else 1), "units")
        self._bind_mousewheel(self.doc_type_listbox, _dt_scroll)

        edit_frame = ttk.Frame(tab)
        edit_frame.pack(fill=tk.X, padx=10, pady=(0, 2))
        self.new_doc_type_var = tk.StringVar()
        entry = ttk.Entry(edit_frame, textvariable=self.new_doc_type_var, width=32)
        entry.pack(side=tk.LEFT, padx=(0, 6))
        entry.bind("<Return>", lambda _: self._add_doc_type())
        ttk.Button(edit_frame, text="Add Document Type", command=self._add_doc_type,
                   bootstyle="primary-outline").pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(edit_frame, text="Remove Selected", command=self._remove_doc_type,
                   bootstyle="danger-outline").pack(side=tk.LEFT)

        save_frame = ttk.Frame(tab)
        save_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        ttk.Button(save_frame, text="Save Document Types", command=self._save_document_types,
                   bootstyle="primary").pack(side=tk.LEFT)
        self.doc_type_status_var = tk.StringVar(value="")
        ttk.Label(save_frame, textvariable=self.doc_type_status_var,
                  foreground="green").pack(side=tk.LEFT, padx=10)

    def _refresh_document_types_tab(self):
        path = self.config_mgr.config["paths"].get("document_types_file", "")
        names = DocumentTypeManager.load(path)
        self.doc_type_listbox.delete(0, tk.END)
        for n in sorted(names):
            self.doc_type_listbox.insert(tk.END, n)

    def _add_doc_type(self):
        name = self.new_doc_type_var.get().strip()
        if not name:
            return
        existing = list(self.doc_type_listbox.get(0, tk.END))
        if name in existing:
            messagebox.showinfo("Duplicate", f'"{name}" is already in the list.')
            return
        existing.append(name)
        existing.sort()
        self.doc_type_listbox.delete(0, tk.END)
        for item in existing:
            self.doc_type_listbox.insert(tk.END, item)
        self.new_doc_type_var.set("")
        self.doc_type_status_var.set("Unsaved changes")

    def _remove_doc_type(self):
        sel = self.doc_type_listbox.curselection()
        if not sel:
            return
        name = self.doc_type_listbox.get(sel[0])
        if messagebox.askyesno(
            "Remove", f'Remove "{name}" from the list?\n\n'
            "Any alternate wordings recorded for it are removed as well. "
            "Every other document type keeps its alternate wordings.",
        ):
            self.doc_type_listbox.delete(sel[0])
            self.doc_type_status_var.set("Unsaved changes")

    def _save_document_types(self):
        path = self.config_mgr.config["paths"].get("document_types_file", "")
        if not path:
            messagebox.showerror("No File", "No document types file is configured.")
            return
        names = list(self.doc_type_listbox.get(0, tk.END))
        try:
            DocumentTypeManager.save(path, names)
            self.doc_type_status_var.set(f"Saved {len(names)} document types.")
        except Exception as e:
            messagebox.showerror("Save Failed", str(e))

    # ── Tab: Providers ───────────────────────────────────────

    def _build_providers_tab(self):
        tab = ttk.Frame(self.notebook)
        self._providers_tab = tab
        self.notebook.add(tab, text="  Providers  ")

        ttk.Label(
            tab,
            text="Facilities, clinics, hospitals, insurers — one per line in providers.txt. "
                 "Used when Settings → Understanding Documents → 'Use the provider list' is on.",
            foreground="gray", wraplength=760, justify="left",
        ).pack(anchor="w", padx=10, pady=(10, 2))

        list_frame = ttk.Frame(tab)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)
        self.provider_listbox = tk.Listbox(
            list_frame, selectmode=tk.SINGLE, font=(APP_FONT, 10)
        )
        pr_sb = ttk.Scrollbar(list_frame, orient="vertical",
                               command=self.provider_listbox.yview)
        self.provider_listbox.configure(yscrollcommand=pr_sb.set)
        pr_sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.provider_listbox.pack(fill=tk.BOTH, expand=True)
        _pr_scroll = lambda e: self.provider_listbox.yview_scroll(
            -1 * (e.delta // 120) if e.delta else (-1 if e.num == 4 else 1), "units")
        self._bind_mousewheel(self.provider_listbox, _pr_scroll)

        edit_frame = ttk.Frame(tab)
        edit_frame.pack(fill=tk.X, padx=10, pady=(0, 2))
        self.new_provider_var = tk.StringVar()
        entry = ttk.Entry(edit_frame, textvariable=self.new_provider_var, width=32)
        entry.pack(side=tk.LEFT, padx=(0, 6))
        entry.bind("<Return>", lambda _: self._add_provider())
        ttk.Button(edit_frame, text="Add Provider", command=self._add_provider,
                   bootstyle="primary-outline").pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(edit_frame, text="Remove Selected", command=self._remove_provider,
                   bootstyle="danger-outline").pack(side=tk.LEFT)

        save_frame = ttk.Frame(tab)
        save_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        ttk.Button(save_frame, text="Save Providers", command=self._save_providers,
                   bootstyle="primary").pack(side=tk.LEFT)
        self.provider_status_var = tk.StringVar(value="")
        ttk.Label(save_frame, textvariable=self.provider_status_var,
                  foreground="green").pack(side=tk.LEFT, padx=10)

    def _refresh_providers_tab(self):
        path = self.config_mgr.config["paths"].get("providers_file", "")
        names = ProviderManager.load(path)
        self.provider_listbox.delete(0, tk.END)
        for n in sorted(names):
            self.provider_listbox.insert(tk.END, n)

    def _add_provider(self):
        name = self.new_provider_var.get().strip()
        if not name:
            return
        existing = list(self.provider_listbox.get(0, tk.END))
        if name in existing:
            messagebox.showinfo("Duplicate", f'"{name}" is already in the list.')
            return
        existing.append(name)
        existing.sort()
        self.provider_listbox.delete(0, tk.END)
        for item in existing:
            self.provider_listbox.insert(tk.END, item)
        self.new_provider_var.set("")
        self.provider_status_var.set("Unsaved changes")

    def _remove_provider(self):
        sel = self.provider_listbox.curselection()
        if not sel:
            return
        name = self.provider_listbox.get(sel[0])
        if messagebox.askyesno("Remove", f'Remove "{name}" from the list?'):
            self.provider_listbox.delete(sel[0])
            self.provider_status_var.set("Unsaved changes")

    def _save_providers(self):
        path = self.config_mgr.config["paths"].get("providers_file", "")
        if not path:
            messagebox.showerror("No File", "No providers file is configured.")
            return
        names = list(self.provider_listbox.get(0, tk.END))
        try:
            ProviderManager.save(path, names)
            self.provider_status_var.set(f"Saved {len(names)} providers.")
        except Exception as e:
            messagebox.showerror("Save Failed", str(e))

    # ── Tab: Suggestions ─────────────────────────────────────
    # Reads LearningStore.pending_suggestions() — client-relationship
    # (alias), document-type, and provider candidates that have cleared
    # their observation-count guard but haven't been confirmed or
    # dismissed by a human yet. Accept/Ignore call the corresponding
    # LearningStore methods (see LearningStore §"suggestions / decisions").

    def _build_suggestions_tab(self):
        tab = ttk.Frame(self.notebook)
        self._suggestions_tab = tab
        self.notebook.add(tab, text="  Suggestions  ")

        ttk.Label(
            tab,
            text="Things the tool has noticed while processing documents and would like "
                 "you to confirm. Nothing here has been applied yet.",
            foreground="gray", wraplength=760, justify="left",
        ).pack(anchor="w", padx=10, pady=(10, 4))

        _scroll_canvas = tk.Canvas(tab, highlightthickness=0)
        _vsb = ttk.Scrollbar(tab, orient="vertical", command=_scroll_canvas.yview)
        _scroll_canvas.configure(yscrollcommand=_vsb.set)
        _vsb.pack(side=tk.RIGHT, fill=tk.Y)
        _scroll_canvas.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        inner = ttk.Frame(_scroll_canvas)
        _win_id = _scroll_canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_inner_configure(e):
            _scroll_canvas.configure(scrollregion=_scroll_canvas.bbox("all"))
        inner.bind("<Configure>", _on_inner_configure)

        def _on_canvas_configure(e):
            _scroll_canvas.itemconfig(_win_id, width=e.width)
        _scroll_canvas.bind("<Configure>", _on_canvas_configure)

        def _on_mousewheel(e):
            _scroll_canvas.yview_scroll(
                -1 * (e.delta // 120) if e.delta else (-1 if e.num == 4 else 1), "units")
        self.after(200, lambda: ScandocsApp._bind_mousewheel(tab, _on_mousewheel))

        self._suggestions_inner = inner

    def _refresh_suggestions_tab(self):
        if not hasattr(self, "_suggestions_inner"):
            return
        for child in self._suggestions_inner.winfo_children():
            child.destroy()
        try:
            store = _get_learning_store(self.config_mgr.config)
            suggestions = store.pending_suggestions()
        except Exception as e:
            logging.warning(f"Could not load suggestions: {e}")
            suggestions = []
        self._refresh_suggestions_badge(len(suggestions))
        if not suggestions:
            ttk.Label(
                self._suggestions_inner, text="No pending suggestions right now.",
                foreground="gray",
            ).pack(anchor="w", padx=8, pady=20)
            return
        for s in suggestions:
            self._build_suggestion_row(self._suggestions_inner, s, store)

    def _refresh_suggestions_badge(self, count: Optional[int] = None):
        """Put a pending-item count on the Suggestions tab label, e.g.
        '  Suggestions (3)  '. Called on tab-change (via _refresh_suggestions_tab)
        and once at startup so the badge is visible before the tab is ever opened."""
        if not hasattr(self, "_suggestions_tab"):
            return
        if count is None:
            try:
                store = _get_learning_store(self.config_mgr.config)
                count = len(store.pending_suggestions())
            except Exception:
                count = 0
        label = f"  Suggestions ({count})  " if count else "  Suggestions  "
        try:
            self.notebook.tab(self._suggestions_tab, text=label)
        except tk.TclError:
            pass

    def _build_suggestion_row(self, parent, s: dict, store: "LearningStore"):
        kind = s.get("kind")
        row = ttk.Frame(parent, padding=(8, 10))
        row.pack(fill=tk.X)
        ttk.Separator(parent, orient="horizontal").pack(fill=tk.X)

        if kind == "alias":
            text = (
                f'File documents for "{s.get("raw_name", "")}" under '
                f'"{s.get("resolved_client", "")}"?  '
                f'(seen in {s.get("observations", 0)} documents)'
            )
            accept_cmd = lambda: self._accept_alias_suggestion(s, store)
            reject_cmd = lambda: self._reject_alias_suggestion(s, store)
        elif kind == "doc_type":
            text = (f'Add "{s.get("name", "")}" to the Document Types list?  '
                    f'(seen {s.get("count", 0)} times)')
            accept_cmd = lambda: self._accept_doc_type_suggestion(s, store)
            reject_cmd = lambda: self._reject_doc_type_suggestion(s, store)
        elif kind == "provider":
            text = (f'Add "{s.get("name", "")}" to the Providers list?  '
                    f'(seen {s.get("count", 0)} times)')
            accept_cmd = lambda: self._accept_provider_suggestion(s, store)
            reject_cmd = lambda: self._reject_provider_suggestion(s, store)
        else:
            return

        ttk.Label(row, text=text, wraplength=720, justify="left",
                  font=(APP_FONT, 10)).pack(anchor="w")
        btns = ttk.Frame(row)
        btns.pack(anchor="w", pady=(6, 0))
        ttk.Button(btns, text="Accept", bootstyle="success",
                   command=accept_cmd).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(btns, text="Ignore", bootstyle="secondary-outline",
                   command=reject_cmd).pack(side=tk.LEFT)

    def _accept_alias_suggestion(self, s: dict, store: "LearningStore"):
        raw_name = s.get("raw_name", "")
        client = s.get("resolved_client", "")
        store.confirm_alias(raw_name, client)
        if self.config_mgr.config.get("learning", {}).get("retroactive_rename") == "preview":
            scandocs = self.config_mgr.config["paths"].get("scandocs_folder", "")
            folders = [scandocs] if scandocs else []
            try:
                proposals = store.plan_retroactive_renames(raw_name, client, folders)
            except Exception as e:
                logging.warning(f"plan_retroactive_renames failed: {e}")
                proposals = []
            if proposals:
                self._show_retroactive_rename_preview(proposals)
        self._refresh_suggestions_tab()

    def _reject_alias_suggestion(self, s: dict, store: "LearningStore"):
        store.reject_alias(s.get("raw_name", ""))
        self._refresh_suggestions_tab()

    def _accept_doc_type_suggestion(self, s: dict, store: "LearningStore"):
        name = s.get("name", "")
        path = self.config_mgr.config["paths"].get("document_types_file", "")
        if name and path:
            names = DocumentTypeManager.load(path)
            if name not in names:
                names.append(name)
                DocumentTypeManager.save(path, names)
        store.accept_doc_type_candidate(name)
        self._refresh_suggestions_tab()
        if getattr(self, "_doc_types_tab", None) is not None:
            self._refresh_document_types_tab()

    def _reject_doc_type_suggestion(self, s: dict, store: "LearningStore"):
        store.reject_doc_type_candidate(s.get("name", ""))
        self._refresh_suggestions_tab()

    def _accept_provider_suggestion(self, s: dict, store: "LearningStore"):
        name = s.get("name", "")
        path = self.config_mgr.config["paths"].get("providers_file", "")
        if name and path:
            names = ProviderManager.load(path)
            if name not in names:
                names.append(name)
                ProviderManager.save(path, names)
        store.accept_provider_candidate(name)
        self._refresh_suggestions_tab()
        if getattr(self, "_providers_tab", None) is not None:
            self._refresh_providers_tab()

    def _reject_provider_suggestion(self, s: dict, store: "LearningStore"):
        store.reject_provider_candidate(s.get("name", ""))
        self._refresh_suggestions_tab()

    # ── Retroactive rename preview (Pass 7 §3) ────────────────

    def _show_retroactive_rename_preview(self, proposals: list):
        """Scrollable checklist of current->proposed renames for files that
        used a raw guess now resolved by a confirmed client relationship.
        Everything is checked by default. Applying routes every rename
        through log_rename so it's undoable via 'Undo Last Rename Batch…',
        and honors automation.dry_run like every other rename in the app."""
        dlg = tk.Toplevel(self)
        dlg.title("Retroactive Rename Preview")
        dlg.resizable(True, True)
        dlg.transient(self)
        dlg.grab_set()
        self.update_idletasks()
        w, h = 640, 480
        x = self.winfo_x() + (self.winfo_width() - w) // 2
        y = self.winfo_y() + (self.winfo_height() - h) // 2
        dlg.geometry(f"{w}x{h}+{x}+{y}")

        ttk.Label(
            dlg,
            text=f"{len(proposals)} file(s) matched this raw name. Everything is checked "
                 "by default — uncheck any you don't want renamed.",
            font=(APP_FONT, 10), wraplength=600, justify="left",
        ).pack(anchor="w", padx=16, pady=(16, 8))

        list_frame = ttk.Frame(dlg)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 8))
        canvas = tk.Canvas(list_frame, highlightthickness=0)
        sb = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(fill=tk.BOTH, expand=True)
        inner = ttk.Frame(canvas)
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))
        self._bind_mousewheel(canvas, lambda e: canvas.yview_scroll(
            -1 * (e.delta // 120) if e.delta else (-1 if e.num == 4 else 1), "units"))

        check_vars = []
        for p in proposals:
            var = tk.BooleanVar(value=True)
            check_vars.append(var)
            ttk.Checkbutton(
                inner, variable=var,
                text=f'{p["current_name"]}  →  {p["proposed_name"]}',
            ).pack(anchor="w", padx=4, pady=2)

        def _apply_selected():
            dry_run = self.config_mgr.config.get("automation", {}).get("dry_run", False)
            batch_id = "retro-" + datetime.datetime.now().strftime("%Y%m%d%H%M%S")
            applied, skipped, errors = 0, 0, []
            for p, var in zip(proposals, check_vars):
                if not var.get():
                    continue
                src = p["path"]
                dst = os.path.join(os.path.dirname(src), p["proposed_name"])
                if not os.path.isfile(src):
                    skipped += 1
                    continue
                if os.path.exists(dst):
                    skipped += 1
                    errors.append(f'{p["current_name"]}: a file already exists at the new name')
                    continue
                if dry_run:
                    applied += 1
                    continue
                try:
                    os.rename(src, dst)
                    log_rename(batch_id, "rename", src, dst, "correction")
                    applied += 1
                except Exception as ex:
                    errors.append(f'{p["current_name"]}: {ex}')
            msg = f"Applied: {applied}\nSkipped: {skipped}"
            if dry_run:
                msg += "\n\n(Preview mode is on — no files were actually renamed.)"
            if errors:
                shown = errors[:10]
                msg += "\n\nErrors:\n" + "\n".join(shown)
                if len(errors) > 10:
                    msg += f"\n… and {len(errors) - 10} more"
            messagebox.showinfo("Retroactive Rename", msg)
            self._refresh_unnamed_count()
            dlg.destroy()

        btn_row = ttk.Frame(dlg)
        btn_row.pack(pady=(0, 16))
        ttk.Button(btn_row, text="Apply Selected", bootstyle="primary",
                   command=_apply_selected).pack(side=tk.LEFT, padx=6)
        ttk.Button(btn_row, text="Cancel", bootstyle="secondary-outline",
                   command=dlg.destroy).pack(side=tk.LEFT, padx=6)

    # ── Tab 3: Settings ───────────────────────────────────────


    def _build_settings_tab(self):
        tab = ttk.Frame(self.notebook)
        self._settings_tab_frame = tab
        self.notebook.add(tab, text="  Settings  ")

        # Registries the generic load/save/reset/dependency logic below
        # walks — every setting built through add_row+register / add_bool /
        # add_mode3 / add_mode2 ends up in _settings_registry, so a setting
        # that's on this tab always round-trips (see _load_settings_to_ui,
        # _save_settings). _bool_setting_specs is the smaller subset of
        # checkboxes that have a prerequisite (dep_key) — the only ones
        # _refresh_settings_dependencies ever touches.
        self._settings_registry = {}
        self._bool_setting_specs = []

        # ── Pinned footer (always visible at the bottom) ──────
        # Pack BEFORE the scroll canvas so pack reserves its space first.
        _footer_frame = ttk.Frame(tab, padding=(20, 8, 20, 10))
        _footer_frame.pack(side=tk.BOTTOM, fill=tk.X)
        ttk.Separator(tab, orient="horizontal").pack(side=tk.BOTTOM, fill=tk.X)

        # ── Scrollable wrapper ────────────────────────────────
        _scroll_canvas = tk.Canvas(tab, highlightthickness=0)
        _vsb = ttk.Scrollbar(tab, orient="vertical", command=_scroll_canvas.yview)
        _scroll_canvas.configure(yscrollcommand=_vsb.set)
        _vsb.pack(side=tk.RIGHT, fill=tk.Y)
        _scroll_canvas.pack(fill=tk.BOTH, expand=True)

        outer = ttk.Frame(_scroll_canvas)
        _win_id = _scroll_canvas.create_window((0, 0), window=outer, anchor="nw")

        def _on_inner_configure(e):
            _scroll_canvas.configure(scrollregion=_scroll_canvas.bbox("all"))
        outer.bind("<Configure>", _on_inner_configure)

        def _on_canvas_configure(e):
            _scroll_canvas.itemconfig(_win_id, width=e.width)
        _scroll_canvas.bind("<Configure>", _on_canvas_configure)

        def _on_mousewheel(e):
            _scroll_canvas.yview_scroll(
                -1 * (e.delta // 120) if e.delta else (-1 if e.num == 4 else 1), "units")
        # Bind scroll recursively after the tab is fully built (so all children exist)
        self.after(200, lambda: ScandocsApp._bind_mousewheel(tab, _on_mousewheel))

        # ── Inner padding frame (same variable name 'outer' the rest of
        #    _build_settings_tab uses, so nothing below changes) ──────
        outer = ttk.Frame(outer, padding=(20, 15, 20, 15))
        outer.pack(fill=tk.BOTH, expand=True)

        # ── Shared builder helpers ─────────────────────────────
        def register(path, var, kind, **meta):
            """Record one setting so _load_settings_to_ui / _save_settings /
            reset-to-defaults / export-import all handle it automatically."""
            self._settings_registry[path] = dict(var=var, kind=kind, **meta)
            return var

        def add_row(parent, row_idx, label_text, str_var,
                    browse_dir=False, browse_file=False, masked=False, info_msg=None):
            ttk.Label(parent, text=label_text).grid(
                row=row_idx, column=0, sticky="w", pady=5
            )
            kw = {"textvariable": str_var, "width": 46}
            if masked:
                kw["show"] = "*"
            entry = ttk.Entry(parent, **kw)
            entry.grid(row=row_idx, column=1, sticky="ew", padx=(8, 4))
            if browse_dir:
                _b = ttk.Button(
                    parent, text="Browse",
                    command=lambda v=str_var: self._browse_dir(v),
                    bootstyle="primary-outline",
                )
                _b.grid(row=row_idx, column=2, padx=(0, 4))
            elif browse_file:
                _b = ttk.Button(
                    parent, text="Browse",
                    command=lambda v=str_var: self._browse_file(v),
                    bootstyle="primary-outline",
                )
                _b.grid(row=row_idx, column=2, padx=(0, 4))
            if info_msg:
                _q = ttk.Button(
                    parent, text="?", width=2,
                    command=lambda msg=info_msg: messagebox.showinfo("Help", msg),
                    bootstyle="primary-outline",
                )
                _q.grid(row=row_idx, column=3, padx=(0, 4))

        def add_reset_button(labelframe, title, paths):
            """Small per-section reset button, corner-anchored so it works
            regardless of whether the frame's children use grid or pack."""
            ttk.Button(
                labelframe, text="Reset", width=7, bootstyle="link",
                command=lambda: self._reset_settings_section(title, paths),
            ).place(relx=1.0, x=-2, y=0, anchor="ne")

        def section_header(text, subtitle=None):
            ttk.Label(outer, text=text, font=(APP_FONT, 13, "bold")).pack(
                anchor="w", pady=(16, 2))
            if subtitle:
                ttk.Label(outer, text=subtitle, font=(APP_FONT, 8),
                          foreground="gray").pack(anchor="w", pady=(0, 8))

        def add_bool(parent, path, label_text, help_text, dep_key=None):
            """One checkbox + a plain-English help line beneath it, wired
            into the settings registry (and, if dep_key is given, into the
            dependency-greying registry — see _refresh_settings_dependencies)."""
            section, key = path.split(".", 1)
            attr = f"s_{section}_{key}_var"
            var = tk.BooleanVar(value=False)
            setattr(self, attr, var)
            row = ttk.Frame(parent)
            row.pack(fill=tk.X, padx=8, pady=(4, 0))
            chk = ttk.Checkbutton(row, text=label_text, variable=var)
            chk.pack(anchor="w")
            ttk.Label(row, text="    " + help_text, font=(APP_FONT, 8),
                      foreground="gray", wraplength=640, justify="left").pack(anchor="w")
            reason_lbl = None
            if dep_key:
                reason_lbl = ttk.Label(row, text="", font=(APP_FONT, 8, "italic"),
                                        foreground="#b35c00", wraplength=640, justify="left")
                reason_lbl.pack(anchor="w")
                self._bool_setting_specs.append({
                    "checkbutton": chk, "dep_key": dep_key, "reason_label": reason_lbl,
                })
            register(path, var, "bool")
            return chk

        def add_mode3(parent, key, label_text, help_text):
            """A 3-state Off / Suggest only / Automatic combobox for one of
            the learning.* settings — see learning_mode_to_display/_to_config."""
            var = tk.StringVar(value="Off")
            attr = f"s_learning_{key}_var"
            setattr(self, attr, var)
            row = ttk.Frame(parent)
            row.pack(fill=tk.X, padx=8, pady=(4, 0))
            ttk.Label(row, text=label_text).pack(side=tk.LEFT)
            combo = ttk.Combobox(row, textvariable=var, values=LEARNING_MODE_VALUES,
                                  state="readonly", width=16)
            combo.pack(side=tk.LEFT, padx=(8, 0))
            _disable_combobox_scroll(combo)
            ttk.Label(parent, text="    " + help_text, font=(APP_FONT, 8),
                      foreground="gray", wraplength=640, justify="left").pack(
                anchor="w", padx=8, pady=(0, 6))
            register(f"learning.{key}", var, "mode3")
            return combo

        def add_mode2(parent, key, label_text, help_text):
            """Retroactive rename's 2-state Off / Preview only combobox —
            deliberately no 'Automatic'. See retroactive_mode_to_display/_to_config."""
            var = tk.StringVar(value="Off")
            attr = f"s_learning_{key}_var"
            setattr(self, attr, var)
            row = ttk.Frame(parent)
            row.pack(fill=tk.X, padx=8, pady=(4, 0))
            ttk.Label(row, text=label_text).pack(side=tk.LEFT)
            combo = ttk.Combobox(row, textvariable=var, values=RETROACTIVE_MODE_VALUES,
                                  state="readonly", width=16)
            combo.pack(side=tk.LEFT, padx=(8, 0))
            _disable_combobox_scroll(combo)
            ttk.Label(parent, text="    " + help_text, font=(APP_FONT, 8),
                      foreground="gray", wraplength=640, justify="left").pack(
                anchor="w", padx=8, pady=(0, 6))
            register(f"learning.{key}", var, "mode2")
            return combo

        # ══════════════════════════════════════════════════════
        # TIER 1 — Everyday
        # ══════════════════════════════════════════════════════
        section_header("Everyday")

        # Paths
        paths_lf = ttk.LabelFrame(outer, text="Paths")
        paths_lf.pack(fill=tk.X, pady=(0, 10))
        paths_lf.columnconfigure(1, weight=1)
        self.s_scandocs_var = tk.StringVar()
        self.s_client_list_var = tk.StringVar()
        add_row(paths_lf, 0, "Scandocs Folder:", self.s_scandocs_var, browse_dir=True)
        add_row(paths_lf, 1, "Client List File:", self.s_client_list_var, browse_file=True)
        register("paths.scandocs_folder", self.s_scandocs_var, "str")
        register("paths.client_list_file", self.s_client_list_var, "str")

        # API / Model
        api_lf = ttk.LabelFrame(outer, text="API & Model")
        api_lf.pack(fill=tk.X, pady=(0, 10))
        api_lf.columnconfigure(1, weight=1)
        self.s_owui_url_var = tk.StringVar()
        self.s_ollama_url_var = tk.StringVar()
        self.s_model_var = tk.StringVar()
        self.s_api_key_var = tk.StringVar()
        add_row(api_lf, 0, "OpenWebUI URL:", self.s_owui_url_var)
        add_row(api_lf, 1, "Ollama URL:", self.s_ollama_url_var)

        # Model row — editable Combobox + Refresh button
        _FALLBACK_MODELS = [
            "llama3.2-vision", "llama3.2", "llama3.1", "llama3",
            "mistral", "mixtral", "phi3", "gemma2", "qwen2",
            "deepseek-r1", "llava", "bakllava",
        ]
        ttk.Label(api_lf, text="Model:").grid(row=2, column=0, sticky="w", pady=5)
        self.s_model_combo = ttk.Combobox(
            api_lf, textvariable=self.s_model_var,
            values=_FALLBACK_MODELS, width=44,
            state="readonly",
        )
        self.s_model_combo.grid(row=2, column=1, sticky="ew", padx=(8, 4))
        _disable_combobox_scroll(self.s_model_combo)
        ttk.Button(
            api_lf, text="⟳", width=3,
            command=self._refresh_models,
            bootstyle="primary-outline",
        ).grid(row=2, column=2, padx=(0, 4))

        add_row(api_lf, 3, "API Key (optional):", self.s_api_key_var, masked=True,
                info_msg=(
                    "API Key\n\n"
                    "Only required if your OpenWebUI instance has authentication enabled.\n\n"
                    "To find your key: in OpenWebUI go to Settings → Account → "
                    "API Keys, then generate or copy an existing key.\n\n"
                    "Leave blank if your server does not require authentication "
                    "(typical for local setups)."
                ))
        register("api.openwebui_url", self.s_owui_url_var, "str")
        register("api.ollama_url", self.s_ollama_url_var, "str")
        register("api.model", self.s_model_var, "str")
        register("api.api_key", self.s_api_key_var, "str")

        # Behavior — the two switches with the most visible consequences
        behavior_lf = ttk.LabelFrame(outer, text="Behavior")
        behavior_lf.pack(fill=tk.X, pady=(0, 10))
        add_reset_button(behavior_lf, "Behavior", ["automation.dry_run"])

        self.s_automation_dry_run_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            behavior_lf,
            text="Preview mode — process files but don't rename or move anything",
            variable=self.s_automation_dry_run_var,
        ).pack(anchor="w", padx=8, pady=(8, 0))
        ttk.Label(
            behavior_lf,
            text="    Runs the whole batch and shows what WOULD happen, without touching "
                 "any files. A big yellow banner appears on the Process tab whenever this "
                 "is on, so a preview run is never mistaken for a real one.",
            font=(APP_FONT, 8), foreground="gray", wraplength=640, justify="left",
        ).pack(anchor="w", padx=8, pady=(0, 8))
        register("automation.dry_run", self.s_automation_dry_run_var, "bool")

        self.s_automation_watch_folder_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            behavior_lf,
            text="Process automatically when files arrive",
            variable=self.s_automation_watch_folder_var,
            state=tk.DISABLED,
        ).pack(anchor="w", padx=8, pady=(0, 0))
        ttk.Label(
            behavior_lf,
            text="    Not yet active — this only saves the setting for now. Nothing "
                 "watches the folder in the background yet; you still need to click "
                 "Auto-Process Documents yourself. Coming in a future update.",
            font=(APP_FONT, 8, "italic"), foreground="#b35c00",
            wraplength=640, justify="left",
        ).pack(anchor="w", padx=8, pady=(0, 8))
        register("automation.watch_folder", self.s_automation_watch_folder_var, "bool")

        # ══════════════════════════════════════════════════════
        # TIER 2 — Features (grouped by what can go wrong)
        # ══════════════════════════════════════════════════════
        section_header("Features",
                        "Turn individual behaviors on or off. Each one explains itself.")

        # ── Naming ──────────────────────────────────────────
        naming_lf = ttk.LabelFrame(outer, text="Naming")
        naming_lf.pack(fill=tk.X, pady=(0, 10))
        add_reset_button(naming_lf, "Naming", [
            "naming.preserve_acronyms", "naming.use_templates", "naming.include_recipient",
            "naming.include_doc_date", "naming.date_disambiguation", "naming.split_unknown_states",
        ])
        add_bool(naming_lf, "naming.preserve_acronyms",
                 "Preserve acronyms in names (PPR, TTD, IME)",
                 "Keeps known acronyms in full capitals instead of Title Case when building a filename.")
        add_bool(naming_lf, "naming.use_templates",
                 "Use document-type naming templates",
                 "Builds each filename using a template chosen by document type (e.g. adds the date "
                 "to Progress Reports) instead of the plain 'Client - Description' format.")
        add_bool(naming_lf, "naming.include_recipient",
                 'Add recipient to name ("to Chiropractic Works")',
                 "Appends who the document was sent to or received from, when known.",
                 dep_key="naming.include_recipient")
        add_bool(naming_lf, "naming.include_doc_date",
                 "Add document date to progress reports",
                 "Includes the date found on the document itself, not just today's date.")
        add_bool(naming_lf, "naming.date_disambiguation",
                 "Number duplicates by date instead of (1)",
                 'When the same client/description appears twice, use the document date to tell '
                 'them apart instead of appending "(1)", "(2)", etc.')
        add_bool(naming_lf, "naming.split_unknown_states",
                 'Separate "unknown client" from "needs review"',
                 "Uses two different labels: one when a name was read but isn't on the client list, "
                 "and another when no name could be read at all.")

        # ── Reading documents ────────────────────────────────
        reading_lf = ttk.LabelFrame(outer, text="Reading Documents")
        reading_lf.pack(fill=tk.X, pady=(0, 10))
        add_reset_button(reading_lf, "Reading Documents", [
            "reading.skip_fax_cover_pages", "reading.deskew_photos",
            "reading.vision_escalation", "reading.extract_claim_numbers",
        ])
        add_bool(reading_lf, "reading.skip_fax_cover_pages",
                 "Skip fax cover and transmission report pages",
                 "Ignores the cover sheet and confirmation page so the AI reads the actual document.")
        add_bool(reading_lf, "reading.deskew_photos",
                 "Straighten and crop photographed pages",
                 "Auto-rotates and trims photos of documents (as opposed to flatbed scans) before reading them.",
                 dep_key="reading.deskew_photos")
        add_bool(reading_lf, "reading.vision_escalation",
                 "Re-read with vision model when text is unclear",
                 "If the extracted text looks too poor to trust, automatically re-reads the page as "
                 "an image with the vision model instead.",
                 dep_key="reading.vision_escalation")
        add_bool(reading_lf, "reading.extract_claim_numbers",
                 "Find claim numbers and injury dates",
                 "Pulls out workers'-comp claim numbers and injury dates when present, for claim linking.")

        # ── Understanding documents ───────────────────────────
        classification_lf = ttk.LabelFrame(outer, text="Understanding Documents")
        classification_lf.pack(fill=tk.X, pady=(0, 10))
        add_reset_button(classification_lf, "Understanding Documents", [
            "classification.structured_output", "classification.use_document_types",
            "classification.use_providers", "classification.extract_recipient",
            "classification.grounding_check", "classification.evidence_confidence",
            "classification.use_candidate_shortlist",
        ])
        add_bool(classification_lf, "classification.structured_output",
                 "Ask the model for structured details",
                 "Requests a structured answer (client, document type, recipient, date, etc.) instead "
                 "of free-form text. Several other settings below require this to be on.")
        add_bool(classification_lf, "classification.use_document_types",
                 "Use the document type list",
                 "Matches the AI's answer against your Document Types list so wording stays consistent.",
                 dep_key="classification.use_document_types")
        add_bool(classification_lf, "classification.use_providers",
                 "Use the provider list",
                 "Matches recipients/senders against your Providers list so names stay consistent.",
                 dep_key="classification.use_providers")
        add_bool(classification_lf, "classification.extract_recipient",
                 "Identify who the document is going to",
                 "Has the AI identify the recipient (e.g. a clinic or insurer) separately from the client.",
                 dep_key="classification.extract_recipient")
        add_bool(classification_lf, "classification.grounding_check",
                 "Verify the client name appears in the document",
                 "Double-checks that the client name the AI returned is actually present in the text "
                 "before trusting it.")
        add_bool(classification_lf, "classification.evidence_confidence",
                 "Judge confidence from evidence instead of the model's own guess",
                 "Calculates how confident to be from what was actually found in the document, rather "
                 "than trusting the AI's self-reported confidence.")
        add_bool(classification_lf, "classification.use_candidate_shortlist",
                 "Narrow the client list before asking",
                 "Sends the AI a shortlist of the most likely clients instead of the full list, which "
                 "can improve accuracy on large client lists.")

        # Re-evaluate dependent Naming/Understanding controls whenever any
        # of their prerequisites change.
        self.s_classification_structured_output_var.trace_add(
            "write", lambda *_: self._refresh_settings_dependencies())
        self.s_classification_extract_recipient_var.trace_add(
            "write", lambda *_: self._refresh_settings_dependencies())
        self.s_model_var.trace_add(
            "write", lambda *_: self._refresh_settings_dependencies())

        # ── Learning ──────────────────────────────────────────
        learn_lf = ttk.LabelFrame(outer, text="Learning")
        learn_lf.pack(fill=tk.X, pady=(0, 10))
        add_reset_button(learn_lf, "Learning", [
            "learning.document_types", "learning.client_relationships", "learning.claim_linking",
            "learning.log_corrections", "learning.observations_required",
            "learning.few_shot_examples", "learning.retroactive_rename",
        ])
        add_mode3(learn_lf, "document_types", "Learn document types:",
                  "Off: never suggest new document types. Suggest only: shows candidates on the "
                  "Suggestions tab for approval. Automatic: adds them without asking.")
        add_mode3(learn_lf, "client_relationships", "Learn client relationships:",
                  "Learns who else's mail belongs under a given client (e.g. a passenger on the "
                  "client's accident). Suggest only surfaces a suggestion; Automatic files future "
                  "documents under the client once confirmed.")
        add_mode3(learn_lf, "claim_linking", "Link documents by claim number:",
                  "Uses a workers'-comp claim number to identify the client on future documents that "
                  "share the same claim number.")
        add_bool(learn_lf, "learning.log_corrections",
                 "Remember corrections",
                 "Keeps a private log of every correction an employee makes. This is the foundation "
                 "the three settings above and 'Use past corrections as examples' below all read "
                 "from — turning it off stops all learning, even 'Suggest only' modes.")
        _obs_row = ttk.Frame(learn_lf)
        _obs_row.pack(fill=tk.X, padx=8, pady=(4, 0))
        ttk.Label(_obs_row, text="Times seen before suggesting:").pack(side=tk.LEFT)
        self.s_learning_observations_required_var = tk.StringVar(value="3")
        _obs_spin = ttk.Spinbox(
            _obs_row, from_=2, to=10, textvariable=self.s_learning_observations_required_var,
            width=6, state="readonly",
        )
        _obs_spin.pack(side=tk.LEFT, padx=(8, 0))
        _disable_combobox_scroll(_obs_spin)
        ttk.Label(
            learn_lf,
            text="    How many distinct documents must agree on the same guess before it's suggested "
                 "for approval. Higher = fewer, more confident suggestions.",
            font=(APP_FONT, 8), foreground="gray", wraplength=640, justify="left",
        ).pack(anchor="w", padx=8, pady=(0, 6))
        register("learning.observations_required", self.s_learning_observations_required_var,
                  "int_str", min=2, max=10, label="Times seen before suggesting")
        add_bool(learn_lf, "learning.few_shot_examples",
                 "Use past corrections as examples",
                 "Shows the AI a few similar past corrections to help it get today's document right.")
        add_mode2(learn_lf, "retroactive_rename", "Retroactively rename old files:",
                  "When a new client relationship is confirmed, offers to rename past files that used "
                  "the same guess. Deliberately has no 'Automatic' option — a person always reviews "
                  "the list on-screen before anything is renamed.")

        # ══════════════════════════════════════════════════════
        # TIER 3 — Advanced (collapsed by default)
        # ══════════════════════════════════════════════════════
        self._advanced_expanded = False
        adv_header = ttk.Frame(outer)
        adv_header.pack(fill=tk.X, pady=(16, 0))
        self._advanced_toggle_btn = ttk.Button(
            adv_header, text="▶  Advanced Settings", bootstyle="link",
            command=self._toggle_advanced_settings,
        )
        self._advanced_toggle_btn.pack(anchor="w")
        ttk.Label(
            adv_header,
            text="Fine-tuning for someone comfortable with the technical details. "
                 "Most people never need to open this.",
            font=(APP_FONT, 8), foreground="gray",
        ).pack(anchor="w", padx=(4, 0))
        self._advanced_frame = ttk.Frame(outer)
        # Not packed yet — _toggle_advanced_settings shows/hides it.
        advanced_frame = self._advanced_frame

        # Processing
        proc_lf = ttk.LabelFrame(advanced_frame, text="Processing")
        proc_lf.pack(fill=tk.X, pady=(0, 10))
        add_reset_button(proc_lf, "Processing", [
            "processing.fuzzy_threshold", "processing.max_ocr_chars", "processing.max_pages",
            "processing.max_vision_pages", "processing.candidate_list_size",
            "processing.ocr_preprocess", "processing.require_high_confidence",
        ])
        proc_lf.columnconfigure(1, weight=1)
        self.s_threshold_var = tk.StringVar()
        self.s_max_chars_var = tk.StringVar()
        self.s_max_pages_var = tk.StringVar()
        add_row(proc_lf, 0, "Fuzzy Match Threshold (0.0 – 1.0):", self.s_threshold_var,
                info_msg=(
                    "Fuzzy Match Threshold\n\n"
                    "Controls how closely a client name found in a document must match "
                    "a name in your client list before it is accepted.\n\n"
                    "1.0 = exact match only\n"
                    "0.85 = allows small differences (recommended)\n"
                    "0.70 = more lenient — may cause false matches\n\n"
                    "If the tool is failing to recognise clients whose names appear "
                    "slightly differently in documents (e.g. missing accents, "
                    "abbreviations), try lowering this value slightly."
                ))
        add_row(proc_lf, 1, "Max OCR Characters:", self.s_max_chars_var,
                info_msg=(
                    "Max OCR Characters\n\n"
                    "The maximum number of characters of extracted text that will be "
                    "sent to the AI for classification.\n\n"
                    "Higher values give the AI more context but increase the time each "
                    "file takes to process.\n\n"
                    "2000–4000 is usually sufficient for identifying the client and "
                    "document type from the first page or two of a document."
                ))
        add_row(proc_lf, 2, "Max Pages Per Document:", self.s_max_pages_var,
                info_msg=(
                    "Max Pages Per Document\n\n"
                    "The maximum number of pages that will be read from each PDF.\n\n"
                    "Keeping this low (5–10) prevents very large documents from "
                    "slowing down the batch. Client names and document types are "
                    "almost always found within the first few pages."
                ))
        register("processing.fuzzy_threshold", self.s_threshold_var, "float_str",
                  min=0.0, max=1.0, label="Fuzzy Match Threshold")
        register("processing.max_ocr_chars", self.s_max_chars_var, "int_str",
                  min=100, label="Max OCR Characters")
        register("processing.max_pages", self.s_max_pages_var, "int_str",
                  min=1, label="Max Pages Per Document")

        ttk.Checkbutton(
            proc_lf,
            text="Require high confidence — only rename when AI is confident (recommended)",
            variable=self.s_require_high_conf_var,
        ).grid(row=3, column=0, columnspan=3, sticky="w", padx=8, pady=(4, 0))
        ttk.Label(
            proc_lf,
            text="  When unchecked, medium-confidence results are also renamed (more matches, higher false-positive risk)",
            font=(APP_FONT, 8), foreground="gray",
        ).grid(row=4, column=0, columnspan=3, sticky="w", padx=8, pady=(0, 6))
        register("processing.require_high_confidence", self.s_require_high_conf_var, "bool")

        self.s_ocr_preprocess_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            proc_lf,
            text="Preprocess scans before OCR — improves accuracy on messy/low-DPI scans (recommended)",
            variable=self.s_ocr_preprocess_var,
        ).grid(row=8, column=0, columnspan=3, sticky="w", padx=8, pady=(4, 0))
        ttk.Button(
            proc_lf, text="?", width=2,
            command=lambda: messagebox.showinfo("Preprocess Scans Before OCR",
                "Preprocess Scans Before OCR\n\n"
                "When enabled, each page image is cleaned up before it reaches "
                "Tesseract:\n\n"
                "  • Upscale — small/low-DPI scans are enlarged toward 300 DPI\n"
                "  • Autocontrast — normalizes faded or shadowed pages\n"
                "  • Binarize — converts the image to pure black and white so "
                "ink stands out from paper texture\n\n"
                "This usually improves OCR accuracy on real-world scans but adds "
                "a small amount of time per page. Turn it off if you notice "
                "worse results on already-clean digital PDFs."),
            bootstyle="primary-outline",
        ).grid(row=8, column=3, padx=(0, 4), pady=(4, 0))
        ttk.Label(
            proc_lf,
            text="  Upscales, normalizes contrast, and binarizes each page image — adds ~100–300ms per page",
            font=(APP_FONT, 8), foreground="gray",
        ).grid(row=9, column=0, columnspan=3, sticky="w", padx=8, pady=(0, 6))
        register("processing.ocr_preprocess", self.s_ocr_preprocess_var, "bool")

        # ── Text Extraction Method (OCR vs Vision) ────────────────
        self.s_extraction_method_var = tk.StringVar()
        self.s_max_vision_pages_var = tk.StringVar()

        ttk.Label(proc_lf, text="Text Extraction Method:").grid(
            row=5, column=0, sticky="w", pady=5
        )
        self.s_method_combo = ttk.Combobox(
            proc_lf,
            textvariable=self.s_extraction_method_var,
            values=["Use OCR", "Use Vision Model"],
            state="readonly",
            width=44,
        )
        self.s_method_combo.grid(row=5, column=1, sticky="ew", padx=(8, 4))
        _disable_combobox_scroll(self.s_method_combo)
        self.s_method_combo.bind(
            "<<ComboboxSelected>>", lambda _e: self._apply_extraction_method_ui()
        )
        ttk.Button(
            proc_lf, text="?", width=2,
            command=lambda: messagebox.showinfo("Text Extraction Method",
                "Text Extraction Method\n\n"
                "Use OCR (default): extracts text from documents with Tesseract OCR "
                "and sends the text to the AI model. Fast and reliable for clean scans.\n\n"
                "Use Vision Model: sends the first few page images directly to a "
                "vision-capable model (Llama 3.2 Vision, Gemma 4). Can help on "
                "low-quality scans, stamps, and letterheads, but is slower.\n\n"
                "Vision Model is only selectable when the currently chosen model "
                "supports vision. Cloud variants are disabled to keep client "
                "documents on-device."),
            bootstyle="primary-outline",
        ).grid(row=5, column=3, padx=(0, 4))

        ttk.Label(proc_lf, text="Max Vision Pages:").grid(
            row=6, column=0, sticky="w", pady=5
        )
        self.s_vision_pages_entry = ttk.Entry(
            proc_lf, textvariable=self.s_max_vision_pages_var, width=46
        )
        self.s_vision_pages_entry.grid(row=6, column=1, sticky="ew", padx=(8, 4))
        ttk.Button(
            proc_lf, text="?", width=2,
            command=lambda: messagebox.showinfo("Max Vision Pages",
                "Max Vision Pages\n\n"
                "How many pages from each PDF are sent to the vision model. "
                "Each page is an image, so this is much more expensive than OCR "
                "— keep it low (1–2 is usually plenty to find the client name).\n\n"
                "Only applies when 'Use Vision Model' is selected."),
            bootstyle="primary-outline",
        ).grid(row=6, column=3, padx=(0, 4))

        self._method_hint_lbl = ttk.Label(
            proc_lf, text="", font=(APP_FONT, 8), foreground="gray",
        )
        self._method_hint_lbl.grid(row=7, column=0, columnspan=4,
                                    sticky="w", padx=8, pady=(0, 6))

        # Re-evaluate vision availability when the user changes the model
        self.s_model_combo.bind(
            "<<ComboboxSelected>>", lambda _e: self._apply_extraction_method_ui()
        )
        self.s_model_var.trace_add(
            "write", lambda *_: self._apply_extraction_method_ui()
        )
        register("processing.max_vision_pages", self.s_max_vision_pages_var, "int_str",
                  min=1, label="Max Vision Pages")

        # candidate list size — how many clients get shortlisted (classification.use_candidate_shortlist)
        self.s_candidate_list_size_var = tk.StringVar()
        add_row(proc_lf, 10, "Candidate List Size:", self.s_candidate_list_size_var,
                info_msg=(
                    "Candidate List Size\n\n"
                    "How many likely clients are shortlisted for the AI when "
                    "'Narrow the client list before asking' is turned on.\n\n"
                    "A smaller list is faster and can improve accuracy on very "
                    "large client lists; too small a list risks leaving the "
                    "correct client out."
                ))
        register("processing.candidate_list_size", self.s_candidate_list_size_var,
                  "int_str", min=1, label="Candidate List Size")

        # Reading details
        reading_adv_lf = ttk.LabelFrame(advanced_frame, text="Reading Details")
        reading_adv_lf.pack(fill=tk.X, pady=(0, 10))
        reading_adv_lf.columnconfigure(1, weight=1)
        self.s_escalation_threshold_var = tk.StringVar()
        add_row(reading_adv_lf, 0, "Vision Escalation Threshold (0.0 – 1.0):",
                self.s_escalation_threshold_var,
                info_msg=(
                    "Vision Escalation Threshold\n\n"
                    "When 'Re-read with vision model when text is unclear' is on, this is "
                    "the text-quality score below which a page is re-read with the vision "
                    "model instead of trusted as OCR text.\n\n"
                    "Lower = only the worst pages get escalated (faster). "
                    "Higher = more pages get the extra vision pass (slower, more thorough)."
                ))
        register("reading.escalation_threshold", self.s_escalation_threshold_var,
                  "float_str", min=0.0, max=1.0, label="Vision Escalation Threshold")

        # API timeouts
        timeout_lf = ttk.LabelFrame(advanced_frame, text="API Timeouts")
        timeout_lf.pack(fill=tk.X, pady=(0, 10))
        timeout_lf.columnconfigure(1, weight=1)
        self.s_timeout_connect_var = tk.StringVar()
        self.s_timeout_read_var = tk.StringVar()
        add_row(timeout_lf, 0, "Connect Timeout (seconds):", self.s_timeout_connect_var)
        add_row(timeout_lf, 1, "Read Timeout (seconds):", self.s_timeout_read_var)
        register("api.timeout_connect", self.s_timeout_connect_var, "int_str",
                  min=1, label="Connect Timeout")
        register("api.timeout_read", self.s_timeout_read_var, "int_str",
                  min=1, label="Read Timeout")

        # Reports
        rep_lf = ttk.LabelFrame(advanced_frame, text="Reports")
        rep_lf.pack(fill=tk.X, pady=(0, 10))
        add_reset_button(rep_lf, "Reports", [
            "reports.report_folder", "reports.auto_save",
            "processing.audit_mode", "processing.show_manual_entry_tab",
        ])
        rep_lf.columnconfigure(1, weight=1)
        self.s_report_folder_var = tk.StringVar()
        add_row(rep_lf, 0, "Report Folder:", self.s_report_folder_var, browse_dir=True)
        register("reports.report_folder", self.s_report_folder_var, "str")
        self.s_auto_save_var = tk.BooleanVar()
        ttk.Checkbutton(
            rep_lf, text="Auto-save report when batch completes",
            variable=self.s_auto_save_var,
        ).grid(row=1, column=0, columnspan=3, sticky="w", padx=8, pady=(0, 6))
        register("reports.auto_save", self.s_auto_save_var, "bool")

        self.s_audit_mode_var = tk.BooleanVar()
        ttk.Checkbutton(
            rep_lf,
            text="Audit Mode — prompt employee to review each result before closing",
            variable=self.s_audit_mode_var,
            command=self._apply_audit_mode,
        ).grid(row=2, column=0, columnspan=3, sticky="w", padx=8, pady=(0, 4))
        _q_audit = ttk.Button(
            rep_lf, text="?", width=2,
            command=lambda: messagebox.showinfo("Audit Mode",
                "Audit Mode\n\n"
                "When enabled, an Audit Mode panel appears below the results table.\n\n"
                "For each document, mark one of:\n"
                "  ✓ Correct — the rename was accurate\n"
                "  Wrong Client Name — Submit Audit will rename it to A-NEEDS REVIEW\n"
                "  Bad Description — Submit Audit will change the description to 'Scanned Document'\n"
                "  Failed to Identify Client — tool could not find the client\n"
                "  Should Have Been Flagged — should have been sent to Manual Correction\n\n"
                "Click 'Submit Audit' to apply all corrections and save the report."),
            bootstyle="primary-outline",
        )
        _q_audit.grid(row=2, column=3, padx=(0, 4), pady=(0, 4), sticky="w")
        register("processing.audit_mode", self.s_audit_mode_var, "bool")

        self.s_file_mode_var = tk.BooleanVar()
        ttk.Checkbutton(
            rep_lf,
            text="File Mode — allow moving renamed files to a destination folder",
            variable=self.s_file_mode_var,
            command=self._apply_file_mode,
        ).grid(row=3, column=0, columnspan=3, sticky="w", padx=8, pady=(0, 2))
        _q_file = ttk.Button(
            rep_lf, text="?", width=2,
            command=lambda: messagebox.showinfo("File Mode",
                "File Mode\n\n"
                "When enabled, a 'File Mode — Move Files' panel appears on the "
                "selected tabs.\n\n"
                "Use 'Apply to Selected' to stage destination folders for one or more "
                "files, then 'Move Files' to commit all staged moves.\n\n"
                "You can enable File Mode independently for the Process tab and "
                "the Manual Correction panel."),
            bootstyle="primary-outline",
        )
        _q_file.grid(row=3, column=3, padx=(0, 4), pady=(0, 2), sticky="w")
        register("processing.file_mode", self.s_file_mode_var, "bool")

        # Sub-checkboxes: per-tab enable
        sub_frame = ttk.Frame(rep_lf)
        sub_frame.grid(row=4, column=0, columnspan=4, sticky="w", padx=28, pady=(0, 8))
        self._fm_auto_chk = ttk.Checkbutton(
            sub_frame, text="Enable in Process tab",
            variable=self.s_file_mode_auto_var,
            command=self._apply_file_mode,
        )
        self._fm_auto_chk.pack(side=tk.LEFT, padx=(0, 16))
        self._fm_manual_chk = ttk.Checkbutton(
            sub_frame, text="Enable in Manual Correction panel",
            variable=self.s_file_mode_manual_var,
            command=self._apply_file_mode,
        )
        self._fm_manual_chk.pack(side=tk.LEFT)
        self._file_mode_sub_frame = sub_frame
        register("processing.file_mode_auto", self.s_file_mode_auto_var, "bool")
        register("processing.file_mode_manual", self.s_file_mode_manual_var, "bool")

        ttk.Checkbutton(
            rep_lf,
            text="Show Manual Entry tab (legacy) — a separate tab for assigning "
                 "files without going through Auto-Process",
            variable=self.s_show_manual_tab_var,
            command=self._apply_manual_tab_visibility,
        ).grid(row=6, column=0, columnspan=3, sticky="w", padx=8, pady=(0, 6))
        register("processing.show_manual_entry_tab", self.s_show_manual_tab_var, "bool")

        # ── Suggest Location sub-option ───────────────────────────
        self._fm_suggest_frame = ttk.Frame(rep_lf)
        self._fm_suggest_frame.grid(row=5, column=0, columnspan=4, sticky="w",
                                    padx=28, pady=(0, 4))

        self._fm_suggest_chk = ttk.Checkbutton(
            self._fm_suggest_frame,
            text="Suggest Location — automatically find the client's folder",
            state=tk.DISABLED,
        )
        self._fm_suggest_chk.pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(
            self._fm_suggest_frame,
            text="Coming Soon",
            foreground="#999999",
            font=(APP_FONT, 8, "italic"),
        ).pack(side=tk.LEFT)

        # ── Auto-commit stub (Coming Soon) ────────────────────────
        self._fm_autocommit_frame = ttk.Frame(rep_lf)
        self._fm_autocommit_frame.grid(row=7, column=0, columnspan=4, sticky="w",
                                       padx=28, pady=(0, 8))
        _ac_chk = ttk.Checkbutton(
            self._fm_autocommit_frame,
            text="Auto-commit file moves",
            state=tk.DISABLED,
        )
        _ac_chk.pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(
            self._fm_autocommit_frame,
            text="Coming Soon — The tool will file documents automatically without human oversight.",
            foreground="#999999",
            font=(APP_FONT, 8, "italic"),
        ).pack(side=tk.LEFT)

        register("processing.suggest_location_enabled", self.s_suggest_loc_var, "bool")
        register("processing.suggest_location_parent_folder", self.s_suggest_parent_var, "str")
        register("processing.file_mode_destination", self.fo_dest_var, "str")

        # Buttons + status — pinned to footer so they're always visible
        btn_row = ttk.Frame(_footer_frame)
        btn_row.pack(fill=tk.X)
        self._btn_test_conn = ttk.Button(btn_row, text="Test API Connection",
                   command=self._test_connection,
                   bootstyle="primary-outline")
        self._btn_test_conn.pack(side=tk.LEFT, padx=(0, 10))
        self._btn_save_settings = ttk.Button(btn_row, text="Save Settings",
                   command=self._save_settings,
                   bootstyle="primary")
        self._btn_save_settings.pack(side=tk.LEFT)
        self.conn_status_var = tk.StringVar(value="")
        self.conn_label = ttk.Label(btn_row, textvariable=self.conn_status_var)
        self.conn_label.pack(side=tk.LEFT, padx=14)

        # Version + update controls — right-aligned on the same footer row
        self._btn_check_updates = ttk.Button(
            btn_row, text="Check for Updates",
            command=lambda: self._check_for_updates_async(silent=False),
            bootstyle="secondary-outline",
        )
        self._btn_check_updates.pack(side=tk.RIGHT)
        ttk.Label(
            btn_row, text=f"v{APP_VERSION}",
            foreground="#888888",
            font=(APP_FONT, 9),
        ).pack(side=tk.RIGHT, padx=(0, 12))

        # Second footer row — settings hygiene: export/import/reset
        btn_row2 = ttk.Frame(_footer_frame)
        btn_row2.pack(fill=tk.X, pady=(6, 0))
        ttk.Button(btn_row2, text="Export Settings…", command=self._export_settings,
                   bootstyle="secondary-outline").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_row2, text="Import Settings…", command=self._import_settings,
                   bootstyle="secondary-outline").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_row2, text="Reset All Settings to Defaults…",
                   command=self._reset_all_settings,
                   bootstyle="danger-outline").pack(side=tk.LEFT)

    # ── Settings helpers ──────────────────────────────────────

    def _load_settings_to_ui(self):
        cfg = self.config_mgr.config
        # Every setting built via register() in _build_settings_tab round-trips
        # generically here — see the "kind" handling below. This is what
        # guarantees a new setting can't silently fail to load.
        for path, meta in self._settings_registry.items():
            section, key = path.split(".", 1)
            section_cfg = cfg.get(section, {}) or {}
            default_val = DEFAULT_CONFIG.get(section, {}).get(key)
            raw = section_cfg.get(key, default_val)
            kind = meta["kind"]
            var = meta["var"]
            if kind == "bool":
                var.set(bool(raw))
            elif kind == "str":
                var.set(raw if raw is not None else "")
            elif kind in ("int_str", "float_str"):
                var.set(str(raw if raw is not None else default_val))
            elif kind == "mode3":
                var.set(learning_mode_to_display(raw))
            elif kind == "mode2":
                var.set(retroactive_mode_to_display(raw))

        # Fields whose on-screen representation differs from their stored
        # form (not a plain scalar) are handled explicitly instead of via
        # the registry.
        _method = cfg["processing"].get("extraction_method", "ocr")
        self.s_extraction_method_var.set(
            "Use Vision Model" if _method == "vision" else "Use OCR"
        )
        if not self.s_report_folder_var.get().strip():
            self.s_report_folder_var.set(DEFAULT_REPORTS_FOLDER)

        self.after(0, self._apply_extraction_method_ui)
        self.after(0, self._apply_audit_mode)
        self.after(0, self._apply_file_mode)
        self.after(0, self._apply_manual_tab_visibility)
        self.after(0, self._refresh_settings_dependencies)
        self.after(0, self._apply_dry_run_banner)
        self.after(400, self._apply_round_styling)
        # Populate model list in background (won't block startup)
        self.after(300, self._refresh_models)

    def _save_settings(self):
        try:
            for path, meta in self._settings_registry.items():
                kind = meta["kind"]
                if kind not in ("int_str", "float_str"):
                    continue
                raw = meta["var"].get().strip()
                label = meta.get("label", path)
                try:
                    val = int(raw) if kind == "int_str" else float(raw)
                except ValueError:
                    raise ValueError(f"{label} must be a number.")
                lo, hi = meta.get("min"), meta.get("max")
                if lo is not None and val < lo:
                    raise ValueError(f"{label} must be at least {lo}.")
                if hi is not None and val > hi:
                    raise ValueError(f"{label} must be at most {hi}.")
        except ValueError as e:
            messagebox.showerror("Invalid Value", str(e))
            return

        cfg = self.config_mgr.config
        for path, meta in self._settings_registry.items():
            section, key = path.split(".", 1)
            cfg.setdefault(section, {})
            kind = meta["kind"]
            var = meta["var"]
            if kind == "bool":
                cfg[section][key] = bool(var.get())
            elif kind == "str":
                cfg[section][key] = var.get().strip()
            elif kind == "int_str":
                cfg[section][key] = int(var.get().strip())
            elif kind == "float_str":
                cfg[section][key] = float(var.get().strip())
            elif kind == "mode3":
                cfg[section][key] = learning_mode_to_config(var.get())
            elif kind == "mode2":
                cfg[section][key] = retroactive_mode_to_config(var.get())

        # Extraction method — force OCR if the selected model can't do vision
        _method_label = self.s_extraction_method_var.get()
        _method = "vision" if _method_label == "Use Vision Model" else "ocr"
        if _method == "vision" and not model_supports_vision(cfg["api"]["model"]):
            _method = "ocr"
        cfg["processing"]["extraction_method"] = _method

        if not cfg["reports"].get("report_folder", "").strip():
            cfg["reports"]["report_folder"] = DEFAULT_REPORTS_FOLDER

        self.config_mgr.save(cfg)
        self._apply_audit_mode()
        self._apply_file_mode()
        self._apply_manual_tab_visibility()
        self._refresh_settings_dependencies()
        self._apply_dry_run_banner()
        self._refresh_client_list_tab()
        self._refresh_unnamed_count()
        messagebox.showinfo("Saved", "Settings saved successfully.")

    # ── Dependency guarding (Pass 7 §2) ───────────────────────

    def _refresh_settings_dependencies(self):
        """Grey out (and explain) every Settings control whose prerequisite
        isn't met right now. Called on load, on save, and via trace_add
        whenever a prerequisite checkbox or the model changes."""
        if not hasattr(self, "_bool_setting_specs"):
            return
        cfg_snapshot = {
            "classification": {
                "structured_output": getattr(
                    self, "s_classification_structured_output_var", tk.BooleanVar(value=False)
                ).get(),
                "extract_recipient": getattr(
                    self, "s_classification_extract_recipient_var", tk.BooleanVar(value=False)
                ).get(),
            },
            "api": {"model": self.s_model_var.get() if hasattr(self, "s_model_var") else ""},
        }
        dep_state = compute_settings_dependency_state(cfg_snapshot)
        for spec in self._bool_setting_specs:
            dep_key = spec.get("dep_key")
            if not dep_key:
                continue
            disabled, reason = dep_state.get(dep_key, (False, ""))
            try:
                spec["checkbutton"].configure(state=(tk.DISABLED if disabled else tk.NORMAL))
                if spec.get("reason_label") is not None:
                    spec["reason_label"].configure(text=reason)
            except tk.TclError:
                pass  # widget destroyed (shouldn't happen, but never crash on a trace callback)

    # ── Collapsible Advanced section ──────────────────────────

    def _toggle_advanced_settings(self):
        self._advanced_expanded = not getattr(self, "_advanced_expanded", False)
        if self._advanced_expanded:
            self._advanced_frame.pack(fill=tk.X, pady=(4, 0))
            self._advanced_toggle_btn.configure(text="▼  Advanced Settings")
        else:
            self._advanced_frame.pack_forget()
            self._advanced_toggle_btn.configure(text="▶  Advanced Settings")

    # ── Preview-mode banner (Process tab) ─────────────────────

    def _apply_dry_run_banner(self):
        if not hasattr(self, "_preview_banner"):
            return
        dry_run = self.config_mgr.config.get("automation", {}).get("dry_run", False)
        if dry_run:
            self._preview_banner.pack(fill=tk.X, before=self._process_btn_row)
        else:
            self._preview_banner.pack_forget()

    # ── Reset to defaults (Pass 7 §5) ─────────────────────────

    def _reset_one_field(self, path):
        meta = self._settings_registry.get(path)
        if not meta:
            return
        section, key = path.split(".", 1)
        default_val = DEFAULT_CONFIG.get(section, {}).get(key)
        kind = meta["kind"]
        var = meta["var"]
        if kind == "bool":
            var.set(bool(default_val))
        elif kind == "str":
            var.set(default_val if default_val is not None else "")
        elif kind in ("int_str", "float_str"):
            var.set(str(default_val))
        elif kind == "mode3":
            var.set(learning_mode_to_display(default_val))
        elif kind == "mode2":
            var.set(retroactive_mode_to_display(default_val))

    def _reset_settings_section(self, title, paths):
        if not messagebox.askyesno(
            "Reset Section",
            f'Reset all settings in "{title}" to their defaults?\n\n'
            "This only changes what's on screen — click Save Settings to keep it.",
        ):
            return
        for path in paths:
            self._reset_one_field(path)
        self._refresh_settings_dependencies()

    def _reset_all_settings(self):
        if not messagebox.askyesno(
            "Reset All Settings",
            "Reset ALL settings on this tab to their defaults?\n\n"
            "This only changes what's on screen — click Save Settings to keep it.",
        ):
            return
        for path in list(self._settings_registry.keys()):
            self._reset_one_field(path)
        self._refresh_settings_dependencies()

    # ── Export / import settings (Pass 7 §5) ──────────────────

    def _export_settings(self):
        path = filedialog.asksaveasfilename(
            title="Export Settings",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            initialfile="scandocs_settings.json",
        )
        if not path:
            return
        try:
            data = ConfigManager._deep_copy(self.config_mgr.config)
            data.get("api", {})["api_key"] = ""  # never write the API key to a shared file
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            messagebox.showinfo(
                "Exported",
                f"Settings exported to:\n{path}\n\n"
                "The API key was left blank for safety — re-enter it after importing "
                "on another machine.",
            )
        except Exception as e:
            messagebox.showerror("Export Failed", str(e))

    def _import_settings(self):
        path = filedialog.askopenfilename(
            title="Import Settings",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                raise ValueError("File does not contain a settings object.")
        except Exception as e:
            messagebox.showerror("Import Failed", f"Could not read settings file:\n{e}")
            return
        if not messagebox.askyesno(
            "Import Settings",
            f"This will replace your current settings with the ones in\n{os.path.basename(path)}.\n\n"
            "Continue?",
        ):
            return
        merged = ConfigManager._deep_merge(ConfigManager._deep_copy(DEFAULT_CONFIG), data)
        self.config_mgr.save(merged)
        self._load_settings_to_ui()
        self._refresh_settings_dependencies()
        messagebox.showinfo("Imported", "Settings imported successfully.")

    # ── Undo last rename batch (Pass 7 §4) ────────────────────

    def _show_undo_dialog(self):
        log = RenameLog()
        batches = log.last_batches(15)
        dlg = tk.Toplevel(self)
        dlg.title("Undo Last Rename Batch")
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()
        self.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 540) // 2
        y = self.winfo_y() + (self.winfo_height() - 420) // 2
        dlg.geometry(f"540x420+{x}+{y}")

        ttk.Label(
            dlg,
            text="Choose a batch to undo. This reverses renames/moves, most-recent\n"
                 "first, back to their original names.",
            font=(APP_FONT, 10),
        ).pack(anchor="w", padx=16, pady=(16, 8))

        if not batches:
            ttk.Label(dlg, text="No rename batches recorded yet.",
                      foreground="gray").pack(padx=16, pady=20)
            ttk.Button(dlg, text="Close", command=dlg.destroy).pack(pady=(0, 16))
            return

        list_frame = ttk.Frame(dlg)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 8))
        lb = tk.Listbox(list_frame, font=(APP_FONT, 10), height=10)
        sb = ttk.Scrollbar(list_frame, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        lb.pack(fill=tk.BOTH, expand=True)
        for b in batches:
            ts = (b.get("last_ts") or "")[:19].replace("T", " ")
            n = b.get("count", 0)
            lb.insert(tk.END, f"{ts} — {n} file{'s' if n != 1 else ''} — batch {b.get('batch_id', '')}")
        lb.selection_set(0)
        self._bind_mousewheel(lb, lambda e: lb.yview_scroll(
            -1 * (e.delta // 120) if e.delta else (-1 if e.num == 4 else 1), "units"))

        def _do_undo():
            sel = lb.curselection()
            if not sel:
                return
            batch = batches[sel[0]]
            ts_disp = (batch.get("last_ts") or "")[:19].replace("T", " ")
            if not messagebox.askyesno(
                "Confirm Undo",
                f"Undo batch from {ts_disp}\n({batch.get('count', 0)} file(s))?\n\n"
                "A file already moved or renamed again since then will be skipped, "
                "not overwritten.",
            ):
                return
            undone, skipped, errors = log.undo_batch(batch.get("batch_id", ""))
            msg = f"Undone: {undone}\nSkipped: {skipped}"
            if errors:
                shown = errors[:10]
                msg += "\n\nErrors:\n" + "\n".join(shown)
                if len(errors) > 10:
                    msg += f"\n… and {len(errors) - 10} more"
            messagebox.showinfo("Undo Complete", msg)
            self._refresh_unnamed_count()
            dlg.destroy()

        btn_row = ttk.Frame(dlg)
        btn_row.pack(pady=(0, 16))
        ttk.Button(btn_row, text="Undo Selected Batch", bootstyle="danger",
                   command=_do_undo).pack(side=tk.LEFT, padx=6)
        ttk.Button(btn_row, text="Close", bootstyle="secondary-outline",
                   command=dlg.destroy).pack(side=tk.LEFT, padx=6)

    def _browse_dir(self, var: tk.StringVar):
        init = var.get() or SCRIPT_DIR
        d = filedialog.askdirectory(title="Select Folder", initialdir=init)
        if d:
            var.set(os.path.normpath(d))

    def _browse_file(self, var: tk.StringVar):
        init = os.path.dirname(var.get()) if var.get() else SCRIPT_DIR
        f = filedialog.askopenfilename(
            title="Select File",
            initialdir=init,
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
        )
        if f:
            var.set(os.path.normpath(f))

    def _refresh_models(self):
        """Fetch available models from Ollama and populate the Model combobox."""
        _FALLBACK_MODELS = [
            "llama3.2-vision", "llama3.2", "llama3.1", "llama3",
            "mistral", "mixtral", "phi3", "gemma2", "qwen2",
            "deepseek-r1", "llava", "bakllava",
        ]
        ollama_url = self.s_ollama_url_var.get().strip() or "http://localhost:11434"

        def _run():
            try:
                import requests as _req
                resp = _req.get(
                    f"{ollama_url.rstrip('/')}/api/tags",
                    timeout=5,
                )
                resp.raise_for_status()
                data = resp.json()
                models = sorted(m["name"] for m in data.get("models", []))
                if not models:
                    raise ValueError("Empty model list")
            except Exception:
                models = _FALLBACK_MODELS
            self.after(0, lambda: self._apply_model_list(models))

        threading.Thread(target=_run, daemon=True).start()

    def _apply_model_list(self, models: list):
        current = self.s_model_var.get().strip()
        self.s_model_combo["values"] = models
        # Keep the current selection if it's still valid, otherwise set first entry
        if current not in models and models:
            # Only overwrite if the field is blank or was a fallback default
            pass  # leave whatever the user typed / saved
        if not current and models:
            self.s_model_var.set(models[0])

    def _test_connection(self):
        self.conn_status_var.set("Testing…")
        self.conn_label.config(foreground="gray")
        self.update_idletasks()
        # Build a temporary config from the current (unsaved) field values
        cfg = ConfigManager._deep_copy(self.config_mgr.config)
        cfg["api"].update({
            "openwebui_url": self.s_owui_url_var.get().strip(),
            "ollama_url":    self.s_ollama_url_var.get().strip(),
            "model":         self.s_model_var.get().strip(),
            "api_key":       self.s_api_key_var.get().strip(),
        })

        def _run():
            ok, msg = APIClient.test_connection(cfg)
            self.after(0, lambda: self._show_conn_result(ok, msg))

        threading.Thread(target=_run, daemon=True).start()

    def _show_conn_result(self, ok: bool, msg: str):
        self.conn_status_var.set(msg)
        self.conn_label.config(foreground="green" if ok else "red")

    # ── Client list tab helpers ───────────────────────────────

    def _on_client_entry_key(self):
        """Show/hide the Add suggestion label as the user types."""
        typed = self.new_client_var.get().strip()
        if not typed:
            self._add_suggestion_lbl.config(text="")
            return
        existing = [c.lower() for c in self.client_listbox.get(0, tk.END)]
        if typed.lower() in existing:
            self._add_suggestion_lbl.config(text="")
        else:
            self._add_suggestion_lbl.config(text=f'  + Add "{typed}" to client list')

    def _refresh_client_list_tab(self):
        path = self.config_mgr.config["paths"]["client_list_file"]
        clients = ClientListManager.load(path)
        self.client_listbox.delete(0, tk.END)
        for c in clients:
            self.client_listbox.insert(tk.END, c)

    def _add_client(self):
        name = self.new_client_var.get().strip()
        if not name:
            return
        if not ClientListManager.is_valid_format(name):
            # Completely invalid (no comma, etc.) — hard warning, don't add
            messagebox.showwarning(
                "Invalid Format",
                f'"{name}" is not in LAST, First format.\n\nExample: GARCIA, Maria',
            )
            return
        # Check for the expected ALL-CAPS LAST, Title First convention
        _convention_ok = bool(re.match(
            r"^[A-Z][A-Z\-\'\.\s]+,\s+[A-Z][a-z]", name.strip()
        ))
        if not _convention_ok:
            if not messagebox.askyesno(
                "Format Warning",
                "The client name format doesn't match the expected format (LAST, First).\n\n"
                "Expected: ALL-CAPS last name, comma, Title-case first name\n"
                f'Example: GARCIA, Maria\n\nYou entered: "{name}"\n\n'
                "Would you like to proceed anyway?"
            ):
                return
        existing = list(self.client_listbox.get(0, tk.END))
        if name in existing:
            messagebox.showinfo("Duplicate", f'"{name}" is already in the list.')
            return
        existing.append(name)
        existing.sort()
        self.client_listbox.delete(0, tk.END)
        for item in existing:
            self.client_listbox.insert(tk.END, item)
        self.new_client_var.set("")
        self._add_suggestion_lbl.config(text="")
        self.client_status_var.set("Unsaved changes")

    def _remove_client(self):
        sel = self.client_listbox.curselection()
        if not sel:
            return
        name = self.client_listbox.get(sel[0])
        if messagebox.askyesno("Remove", f'Remove "{name}" from the list?'):
            self.client_listbox.delete(sel[0])
            self.client_status_var.set("Unsaved changes")

    def _save_client_list(self):
        path = self.config_mgr.config["paths"]["client_list_file"]
        if not path:
            messagebox.showerror(
                "Not Configured",
                "Client list file path is not set.\nConfigure it in Settings first.",
            )
            return
        clients = list(self.client_listbox.get(0, tk.END))
        try:
            ClientListManager.save(path, clients)
            self.client_status_var.set(f"Saved {len(clients)} client(s)")
            self.after(3000, lambda: self.client_status_var.set(""))
        except Exception as e:
            messagebox.showerror("Save Failed", str(e))

    # ── Manual Correction helpers ──────────────────────────────

    def _refresh_review_tab(self):
        """Reload the client-list cache used by the Manual Correction panel's
        client entry/listbox, and — if the legacy Manual Entry tab is enabled —
        rescan the scandocs folder for files still awaiting review. Called
        after processing finishes, after a correction is applied, and after
        an audit is submitted."""
        client_list_path = self.config_mgr.config["paths"]["client_list_file"]
        self._all_clients = ClientListManager.load(client_list_path)
        if hasattr(self, "corr_client_listbox"):
            self._filter_client_combo()

        if not hasattr(self, "review_listbox"):
            return
        scandocs = self.config_mgr.config["paths"]["scandocs_folder"]
        self._review_selected_file = ""
        self.review_listbox.delete(0, tk.END)
        if not os.path.isdir(scandocs):
            self.review_count_var.set("Scandocs folder not found")
        else:
            review_files = sorted(
                f for f in os.listdir(scandocs)
                if os.path.isfile(os.path.join(scandocs, f))
                and os.path.splitext(f)[1].lower() in SUPPORTED_EXTENSIONS
                and not FileProcessor._already_processed(f, self._all_clients, self.config_mgr.config.get("naming", {}))
            )
            for f in review_files:
                self.review_listbox.insert(tk.END, f)
            self.review_count_var.set(f"Files awaiting review: {len(review_files)}")
        self._filter_review_client_combo()

    def _apply_manual_tab_visibility(self):
        """Attach or detach the legacy Manual Entry tab based on the Settings toggle."""
        show = self.s_show_manual_tab_var.get()
        present = str(self._review_tab_frame) in self.notebook.tabs()
        if show and not present:
            self.notebook.insert(1, self._review_tab_frame, text="  Manual Entry  ")
            self._refresh_review_tab()
        elif not show and present:
            self.notebook.forget(self._review_tab_frame)

    def _on_review_select(self, _event):
        sel = self.review_listbox.curselection()
        if not sel:
            return
        filename = self.review_listbox.get(sel[0])
        self._review_selected_file = filename   # remember even if focus moves away
        # Pre-fill the subject field: extract whatever comes after " - " if present,
        # otherwise use the bare filename (without extension) as a starting point.
        m = re.match(r"^.+? - (.+)\.(pdf|jpg|jpeg)$", filename, re.IGNORECASE)
        if m:
            self.review_subject_var.set(m.group(1))
        else:
            self.review_subject_var.set(os.path.splitext(filename)[0])

        # If the preview popup is open, refresh it with the newly selected file
        if self._file_popup is not None and self._file_popup.winfo_exists():
            path = os.path.join(
                self.config_mgr.config["paths"]["scandocs_folder"], filename)
            if os.path.isfile(path):
                self._refresh_popup_content(path)

    def _open_review_file(self):
        sel = self.review_listbox.curselection()
        filename = (self.review_listbox.get(sel[0]) if sel
                    else self._review_selected_file)
        if not filename:
            messagebox.showinfo("No Selection", "Please select a file from the list first.")
            return
        path = os.path.join(
            self.config_mgr.config["paths"]["scandocs_folder"], filename
        )
        if os.path.isfile(path):
            self._open_file_popup(path)
        else:
            messagebox.showerror("File Not Found", f"{filename} no longer exists.")
            self._refresh_review_tab()

    def _assign_review_file(self):
        # Use the stored selection — it persists even when focus is on the Assign fields
        sel = self.review_listbox.curselection()
        filename = (self.review_listbox.get(sel[0]) if sel
                    else self._review_selected_file)
        if not filename:
            messagebox.showinfo("No Selection", "Please select a file from the list first.")
            return
        client  = self.review_client_var.get().strip()
        subject = self.review_subject_var.get().strip()

        if not client:
            messagebox.showwarning("Missing Client", "Please select a client from the dropdown.")
            return
        if not subject:
            messagebox.showwarning("Missing Subject", "Please enter a subject for the document.")
            return

        ext      = os.path.splitext(filename)[1].lower()
        safe_sub = FileProcessor._safe_subject(subject) or "Document"
        new_name = f"{client} - {safe_sub}{ext}"

        scandocs = self.config_mgr.config["paths"]["scandocs_folder"]
        new_name = FileProcessor._resolve_collision(scandocs, new_name, filename)
        src = os.path.join(scandocs, filename)
        dst = os.path.join(scandocs, new_name)
        try:
            os.rename(src, dst)
            # If file mode is on and a destination is set, move the renamed file there
            if self.s_file_mode_var.get():
                dest = self.fo_dest_var.get().strip()
                if dest and os.path.isdir(dest):
                    import shutil
                    final_name = new_name
                    final_dst = os.path.join(dest, final_name)
                    counter = 1
                    base2, ext2 = os.path.splitext(final_name)
                    while os.path.exists(final_dst):
                        final_name = f"{base2} ({counter}){ext2}"
                        final_dst = os.path.join(dest, final_name)
                        counter += 1
                    shutil.move(dst, final_dst)
                    messagebox.showinfo("Success",
                        f"Renamed and moved to:\n{os.path.basename(dest)}")
                else:
                    messagebox.showinfo("Success", f"Renamed to:\n{new_name}")
            else:
                messagebox.showinfo("Success", f"Renamed to:\n{new_name}")
            self._refresh_review_tab()
            self.review_client_var.set("")
            self.review_subject_var.set("")
        except Exception as e:
            messagebox.showerror("Rename Failed", str(e))

    def _review_prev(self, _event=None):
        """Move to the previous file in the legacy Manual Entry list."""
        n = self.review_listbox.size()
        if n == 0:
            return
        sel = self.review_listbox.curselection()
        idx = (sel[0] - 1) if sel else n - 1
        idx = max(idx, 0)
        self.review_listbox.selection_clear(0, tk.END)
        self.review_listbox.selection_set(idx)
        self.review_listbox.see(idx)
        self._on_review_select(None)

    def _review_next(self, _event=None):
        """Move to the next file in the legacy Manual Entry list."""
        n = self.review_listbox.size()
        if n == 0:
            return
        sel = self.review_listbox.curselection()
        idx = (sel[0] + 1) if sel else 0
        idx = min(idx, n - 1)
        self.review_listbox.selection_clear(0, tk.END)
        self.review_listbox.selection_set(idx)
        self.review_listbox.see(idx)
        self._on_review_select(None)

    def _on_review_return(self, _event=None):
        """Spacebar on the legacy Manual Entry list: toggle the file viewer."""
        sel = self.review_listbox.curselection()
        filename = (self.review_listbox.get(sel[0]) if sel
                    else self._review_selected_file)
        if not filename:
            return "break"
        path = os.path.join(
            self.config_mgr.config["paths"]["scandocs_folder"], filename)
        if os.path.isfile(path):
            self._toggle_file_popup(path)
        return "break"

    def _on_review_double_click(self, _event=None):
        """Double-click on the legacy Manual Entry list: always open the file viewer."""
        sel = self.review_listbox.curselection()
        filename = (self.review_listbox.get(sel[0]) if sel
                    else self._review_selected_file)
        if not filename:
            return "break"
        path = os.path.join(
            self.config_mgr.config["paths"]["scandocs_folder"], filename)
        if os.path.isfile(path):
            self._open_file_popup(path)
        return "break"

    def _filter_review_client_combo(self, _event=None):
        """Filter the legacy Manual Entry tab's client list as the user types."""
        typed = self.review_client_var.get().lower()
        self.review_client_listbox.delete(0, tk.END)
        for name in self._all_clients:
            if not typed or typed in name.lower():
                self.review_client_listbox.insert(tk.END, name)

    def _on_review_client_listbox_select(self, _event=None):
        """Clicking a name in the legacy client list copies it to the entry field."""
        sel = self.review_client_listbox.curselection()
        if sel:
            self.review_client_var.set(self.review_client_listbox.get(sel[0]))

    def _open_manual_correction(self, iid: str):
        """Show the Manual Correction panel, pre-filled with this row's
        auto-process result so the employee only has to fix what's wrong."""
        result = self._iid_to_result.get(iid)
        if not result:
            return

        scandocs = self.config_mgr.config["paths"]["scandocs_folder"]
        filename = result.final_name
        if not os.path.isfile(os.path.join(scandocs, filename)):
            filename = result.original_name

        self._correction_iid = iid
        self.results_tree.selection_set(iid)
        self.results_tree.see(iid)
        self.corr_file_var.set(f"Correcting:  {filename}")

        # Populate with the auto-process info first — if the client/subject
        # split is already right, the employee only needs to touch the field
        # that's wrong (often just the file name itself).
        m = re.match(r"^(.+?) - (.+)\.(pdf|jpg|jpeg)$", filename, re.IGNORECASE)
        if m:
            self.corr_client_var.set(m.group(1))
            self.corr_subject_var.set(m.group(2))
        else:
            self.corr_client_var.set(result.client or "")
            self.corr_subject_var.set(result.description or os.path.splitext(filename)[0])

        self._filter_client_combo()

        if not self.correction_panel.winfo_ismapped():
            self.correction_panel.pack(fill=tk.X, padx=10, pady=(6, 4),
                                        after=self._tree_frame)
            self._grow_window_for_correction_panel()

    def _hide_manual_correction(self):
        self.correction_panel.pack_forget()
        self._correction_iid = ""
        self._shrink_window_after_correction_panel()

    def _grow_window_for_correction_panel(self):
        """Grow the main window so the newly-shown Manual Correction panel
        doesn't just squeeze the results table — all its fields stay visible."""
        if self.state() == "zoomed":
            return   # already fullscreen; the tree just shrinks to make room
        self.update_idletasks()
        panel_h = self.correction_panel.winfo_reqheight()
        if self._pre_correction_height is None:
            self._pre_correction_height = self.winfo_height()
        screen_h = self.winfo_screenheight()
        new_h = min(self._pre_correction_height + panel_h, screen_h - 60)
        self.geometry(f"{self.winfo_width()}x{new_h}+{self.winfo_x()}+{self.winfo_y()}")

    def _shrink_window_after_correction_panel(self):
        """Restore the window to its pre-panel height, if we grew it."""
        if self._pre_correction_height is None or self.state() == "zoomed":
            self._pre_correction_height = None
            return
        self.geometry(
            f"{self.winfo_width()}x{self._pre_correction_height}"
            f"+{self.winfo_x()}+{self.winfo_y()}")
        self._pre_correction_height = None

    def _corr_open_preview(self):
        """The Manual Correction panel's big 'Open File Preview' button."""
        result = self._iid_to_result.get(self._correction_iid)
        if not result:
            return
        scandocs = self.config_mgr.config["paths"]["scandocs_folder"]
        path = os.path.join(scandocs, result.final_name)
        if not os.path.isfile(path):
            path = os.path.join(scandocs, result.original_name)
        if os.path.isfile(path):
            self._open_file_popup(path)
        else:
            messagebox.showwarning("File Not Found",
                f"Could not locate the file:\n{result.final_name}")

    # ── PASS 6 (learning): correction logging ───────────────────────────
    #
    # The office doesn't use the Audit checkboxes (too much hassle) but
    # DOES use the Manual Correction panel — so _corr_commit (below), not
    # _submit_audit, is where nearly all of the firm's real correction
    # signal comes from. Both "Close" and "Next →" route through
    # _corr_commit, so every correction the office makes passes through
    # this one chokepoint.

    def _log_correction_event(self, result: "ProcessResult",
                               predicted_client: str, predicted_desc: str,
                               corrected_client: str, corrected_desc: str,
                               source: str) -> None:
        """Log one correction/confirmation/audit event to the LearningStore,
        gated on learning.log_corrections. Never raises and never blocks a
        rename — a logging failure here must not affect the file operation
        that already happened."""
        if not self.config_mgr.config.get("learning", {}).get("log_corrections", True):
            return
        try:
            changed_client = (corrected_client or "").strip() != (predicted_client or "").strip()
            changed_desc = (corrected_desc or "").strip() != (predicted_desc or "").strip()
            store = _get_learning_store(self.config_mgr.config)
            entry_source = source
            if source == "correction" and not changed_client and not changed_desc:
                # A no-change commit is only an implicit confirmation when
                # there was a real prediction to confirm. On a row that is
                # still unresolved, "no change" means the employee looked
                # and moved on — clicking between rows with the Manual
                # Correction panel open does exactly this — so recording it
                # as a confirmation of "A-NEEDS REVIEW" would be inventing
                # agreement that never happened.
                if store.is_sentinel(corrected_client):
                    return
                # Free audit data from a workflow the office already does.
                entry_source = "confirmation"
            store.log_correction({
                "doc_hash": getattr(result, "doc_hash", ""),
                "original_name": getattr(result, "original_name", ""),
                "predicted_client": predicted_client or "",
                "raw_client": getattr(result, "raw_client", ""),
                "predicted_desc": predicted_desc or "",
                "predicted_doc_type": getattr(result, "doc_type", ""),
                "predicted_recipient": getattr(result, "recipient", ""),
                "corrected_client": corrected_client or "",
                "corrected_desc": corrected_desc or "",
                "claim_number": getattr(result, "claim_number", ""),
                "doc_date": getattr(result, "doc_date", ""),
                "text_excerpt": (getattr(result, "extracted_text", "") or "")[:600],
                "changed_client": changed_client,
                "changed_desc": changed_desc,
                "source": entry_source,
            })
        except Exception as e:
            logging.warning(f"Could not log correction: {e}")

    def _corr_commit(self) -> bool:
        """Apply the Manual Correction panel's fields to the file under
        correction: rename it (and move it, if Manual File Mode is on).

        Returns False (after showing a message) when the fields aren't
        valid yet. Callers that ADVANCE to another row must honor that and
        stay put — otherwise the warning flashes up and the half-finished
        edit is thrown away as the panel rebinds to a different file.
        Closing is the one exception: _corr_close deliberately ignores the
        result, because a user must never be trapped in the panel by a
        validation error they don't want to resolve."""
        iid = self._correction_iid
        if not iid or iid not in self._iid_to_result:
            return True
        result = self._iid_to_result[iid]

        client  = self.corr_client_var.get().strip()
        subject = self.corr_subject_var.get().strip()
        if not client:
            messagebox.showwarning("Missing Client", "Please select a client from the list.")
            return False
        if not subject:
            messagebox.showwarning("Missing Subject", "Please enter a subject for the document.")
            return False

        scandocs = self.config_mgr.config["paths"]["scandocs_folder"]
        current_name = result.final_name
        if not os.path.isfile(os.path.join(scandocs, current_name)):
            current_name = result.original_name
        if not os.path.isfile(os.path.join(scandocs, current_name)):
            # File's already gone (moved/renamed outside the app, or by a
            # prior correction) — nothing to rename. Don't block Close/Next
            # over it, just leave this row as-is.
            messagebox.showwarning(
                "File Not Found",
                f"Could not locate \"{current_name}\" on disk — "
                "skipping the rename for this file.")
            return True

        dry_run = self.config_mgr.config.get("automation", {}).get("dry_run", False)

        # Captured BEFORE the mutations below overwrite result.client /
        # result.description — these are the model's original prediction,
        # exactly what PASS 6 needs to backlog.
        predicted_client = result.client
        predicted_desc = result.description

        ext      = os.path.splitext(current_name)[1].lower()
        safe_sub = FileProcessor._safe_subject(subject) or "Document"
        new_name = f"{client} - {safe_sub}{ext}"

        if new_name != current_name:
            new_name = FileProcessor._resolve_collision(scandocs, new_name, current_name)
            if dry_run:
                self.status_var.set("Preview mode is on — no files were renamed.")
            else:
                old_path = os.path.join(scandocs, current_name)
                new_path = os.path.join(scandocs, new_name)
                try:
                    os.rename(old_path, new_path)
                    if self.config_mgr.config.get("safety", {}).get("undo_log", True):
                        log_rename(getattr(self.engine, "batch_id", "") or "manual",
                                   "rename", old_path, new_path, "correction")
                except Exception as e:
                    messagebox.showwarning(
                        "Rename Failed",
                        f"Could not rename the file:\n{e}\n\nMoving on without renaming.")
                    return True
        else:
            new_name = current_name

        # PASS 6: log against the ORIGINAL ProcessResult (still holding the
        # model's raw_client/doc_hash/etc.) before it's overwritten below.
        self._log_correction_event(
            result, predicted_client, predicted_desc, client, subject,
            source="correction",
        )

        result.final_name           = new_name
        result.client               = client
        result.description          = subject
        result.status               = "renamed"
        result.was_dry_run          = dry_run
        result.audit_corrected_name = new_name

        vals = list(self.results_tree.item(iid, "values"))
        vals[2] = new_name   # New Name
        vals[3] = "OK (preview)" if dry_run else "OK"   # Status
        vals[4] = client     # Client
        self.results_tree.item(iid, values=vals, tags=("renamed",))

        # Optional move, if Manual File Mode has a destination configured
        if self.s_file_mode_var.get() and self.s_file_mode_manual_var.get():
            dest = self.fo_dest_var.get().strip()
            if dest and os.path.isdir(dest):
                self._fo_do_move(result, dest, iid)

        return True

    def _corr_close(self):
        """Apply the correction (best-effort) and hide the panel — closing
        should never be blocked, even if the rename couldn't be applied."""
        self._corr_commit()
        self._hide_manual_correction()

    def _corr_next(self):
        """Apply the correction, then open the next row — but stay on this
        row if the fields didn't validate, so the edit isn't discarded."""
        if not self._corr_commit():
            return
        children = self.results_tree.get_children()
        if not children or self._correction_iid not in children:
            self._hide_manual_correction()
            return
        idx = children.index(self._correction_iid) + 1
        if idx >= len(children):
            self._hide_manual_correction()
            return
        self._open_manual_correction(children[idx])

    # ── Processing ────────────────────────────────────────────

    def _start_processing(self):
        errors = self.config_mgr.validate()
        if errors:
            messagebox.showerror(
                "Configuration Error",
                "\n".join(errors) + "\n\nPlease fix these in Settings.",
            )
            self.notebook.select(self._settings_tab_frame)
            return

        dry_run = self.config_mgr.config.get("automation", {}).get("dry_run", False)
        # Clearly-visible indicator when preview mode is on — window title,
        # the Process tab banner, and the status bar all reflect it,
        # re-evaluated at the start of every batch so it stays in sync if
        # the config changes between runs.
        self.title(APP_TITLE + (" — PREVIEW MODE (no files will be renamed)" if dry_run else ""))
        self._apply_dry_run_banner()

        # Drain anything left over from a previous run. _poll_queue stops
        # reading the moment it sees "error" or "stopped", so the "done"
        # that follows stays queued; the next run would then consume that
        # stale "done" immediately, report "Done. 0 file(s) processed",
        # re-enable the buttons and stop polling — while the batch it just
        # started carried on renaming files invisibly in the background.
        while True:
            try:
                self._queue.get_nowait()
            except queue.Empty:
                break

        # Clear old results
        self._hide_manual_correction()
        for row in self.results_tree.get_children():
            self.results_tree.delete(row)
        for btn in self._correction_buttons.values():
            btn.destroy()
        self._correction_buttons.clear()
        self._results.clear()
        self._iid_to_result.clear()
        self._total_files = 0
        self.progress_var.set(0)
        # Reset audit panel
        self.audit_file_label.config(
            text="Select a row above to audit it.", foreground="gray"
        )
        self._audit_updating = True
        self.audit_correct_var.set(False)
        self.audit_wrong_client_var.set(False)
        self.audit_bad_desc_var.set(False)
        self.audit_failed_client_var.set(False)
        self.audit_should_flag_var.set(False)
        self._audit_updating = False
        for w in (self.audit_open_btn, self.audit_prev_btn, self.audit_next_btn,
                  self.audit_correct_chk, self.audit_wrong_client_chk,
                  self.audit_bad_desc_chk, self.audit_failed_client_chk,
                  self.audit_should_flag_chk):
            w.config(state=tk.DISABLED)
        self.status_var.set(
            "Starting… — PREVIEW MODE: no files will be renamed" if dry_run else "Starting…"
        )
        self.btn_process.config(state=tk.DISABLED)
        self.btn_stop.config(state=tk.NORMAL)
        # Reset audit submit button for new run
        self.btn_submit_audit.config(state=tk.NORMAL, text="Submit Audit")
        self._processing_active = True

        t = threading.Thread(
            target=self.engine.run_batch,
            args=(self.config_mgr.config, self._queue),
            daemon=True,
        )
        t.start()
        self.after(100, self._poll_queue)

    def _stop_processing(self):
        self.engine.stop()
        self.status_var.set("Stopping after current file…")
        self.btn_stop.config(state=tk.DISABLED)

    def _poll_queue(self):
        try:
            while True:
                msg = self._queue.get_nowait()
                mtype = msg["type"]

                if mtype == "total":
                    self._total_files = msg["count"]
                    self.status_var.set(
                        f"Found {self._total_files} file(s). Processing…"
                    )

                elif mtype == "progress":
                    n = msg["current"]
                    pct = (n / self._total_files * 100) if self._total_files else 0
                    self.progress_var.set(pct)
                    self.status_var.set(
                        f"Processing {n} / {self._total_files}: {msg['filename']}"
                    )

                elif mtype == "result":
                    result = msg["result"]
                    # Suggest Location: pre-fill pending_dest if enabled
                    if (self.s_suggest_loc_var.get()
                            and self.s_file_mode_var.get()
                            and result.status in ("renamed", "needs_review")
                            and result.client
                            and not result.pending_dest):
                        _parent = self.s_suggest_parent_var.get().strip()
                        _desc = result.description
                        suggested = self._suggest_location(result.client, _desc, _parent)
                        if suggested:
                            result.pending_dest = suggested
                    self._add_result_row(result)

                elif mtype == "error":
                    messagebox.showerror("Error", msg["message"])
                    self._finish_processing()
                    return

                elif mtype == "stopped":
                    self.status_var.set("Stopped by user.")
                    self._finish_processing()
                    return

                elif mtype == "done":
                    self.progress_var.set(100)
                    self.status_var.set(
                        f"Done. {len(self._results)} file(s) processed."
                    )
                    self._finish_processing(auto_save=True)
                    self._refresh_review_tab()
                    self._refresh_unnamed_count()
                    return

        except queue.Empty:
            pass
        self.after(100, self._poll_queue)

    def _finish_processing(self, auto_save: bool = False):
        self._processing_active = False
        self.btn_process.config(state=tk.NORMAL)
        self.btn_stop.config(state=tk.DISABLED)
        # If the user finished auditing while processing was still running,
        # the audit-complete check was deferred — run it now.
        self._check_audit_complete()
        if auto_save and self._results and self.s_auto_save_var.get():
            folder = self.s_report_folder_var.get().strip() or DEFAULT_REPORTS_FOLDER
            try:
                path = self._write_report(folder)
                self.status_var.set(
                    self.status_var.get() + f"  |  Report saved: {os.path.basename(path)}"
                )
            except Exception as e:
                messagebox.showerror("Auto-Save Failed", f"Could not save report:\n{e}")

    # ── Report helpers ────────────────────────────────────────

    _REPORT_HEADERS = [
        "Original File", "New Name", "Status",
        "Client", "Description", "Confidence",
        "Extraction Method", "Renamed At", "Error",
        "Moved To",
        "Audit: Correct", "Audit: Wrong Client", "Audit: Bad Description",
        "Audit: Failed to Identify Client", "Audit: Should Have Flagged",
        "Audit: Corrected Name",
    ]

    def _results_as_rows(self) -> list:
        rows = []
        for r in self._results:
            rows.append([
                r.original_name, r.final_name, r.status,
                r.client, r.description, r.confidence,
                r.extraction_method, r.renamed_at or "", r.error_message or "",
                r.moved_to,
                "Yes" if r.audit_correct         else "",
                "Yes" if r.audit_wrong_client    else "",
                "Yes" if r.audit_bad_description else "",
                "Yes" if r.audit_failed_client   else "",
                "Yes" if r.audit_should_review   else "",
                r.audit_corrected_name,
            ])
        # Summary row
        total      = len(self._results)
        renamed    = sum(1 for r in self._results if r.status == "renamed")
        flagged    = sum(1 for r in self._results if r.status == "needs_review")
        moved      = sum(1 for r in self._results if r.moved_to)
        correct    = sum(1 for r in self._results if r.audit_correct)
        wrong_cl   = sum(1 for r in self._results if r.audit_wrong_client)
        bad_desc   = sum(1 for r in self._results if r.audit_bad_description)
        failed_cl  = sum(1 for r in self._results if r.audit_failed_client)
        sh_review  = sum(1 for r in self._results if r.audit_should_review)
        rows.append([])   # blank separator
        rows.append([
            "SUMMARY", "", "",
            f"Total: {total}", f"Renamed: {renamed}", f"Auto-flagged: {flagged}",
            "", "", "",
            f"Moved: {moved}",
            f"Confirmed correct: {correct}",
            f"Wrong client: {wrong_cl}",
            f"Bad description: {bad_desc}",
            f"Failed to identify client: {failed_cl}",
            f"Should have flagged: {sh_review}",
            "",
        ])
        return rows

    @staticmethod
    def _config_flags_snapshot(cfg: dict) -> list:
        """Flatten the feature-flag portions of `cfg` into (label, value)
        pairs, stamped into every report (Pass 7 §5) — 'what was turned on'
        is the first question when a batch looks wrong."""
        def yn(v):
            return "On" if v else "Off"
        naming = cfg.get("naming", {}) or {}
        reading = cfg.get("reading", {}) or {}
        classification = cfg.get("classification", {}) or {}
        learning = cfg.get("learning", {}) or {}
        processing = cfg.get("processing", {}) or {}
        automation = cfg.get("automation", {}) or {}
        api = cfg.get("api", {}) or {}
        return [
            ("Model", api.get("model", "")),
            ("Extraction Method",
             "Vision" if processing.get("extraction_method") == "vision" else "OCR"),
            ("Preview Mode (dry run)", yn(automation.get("dry_run", False))),
            ("Require High Confidence", yn(processing.get("require_high_confidence", True))),
            ("", ""),
            ("Naming: Preserve Acronyms", yn(naming.get("preserve_acronyms", False))),
            ("Naming: Use Templates", yn(naming.get("use_templates", False))),
            ("Naming: Include Recipient", yn(naming.get("include_recipient", False))),
            ("Naming: Include Doc Date", yn(naming.get("include_doc_date", False))),
            ("Naming: Date Disambiguation", yn(naming.get("date_disambiguation", False))),
            ("Naming: Split Unknown States", yn(naming.get("split_unknown_states", False))),
            ("", ""),
            ("Reading: Skip Fax Cover Pages", yn(reading.get("skip_fax_cover_pages", False))),
            ("Reading: Deskew Photos", yn(reading.get("deskew_photos", False))),
            ("Reading: Vision Escalation", yn(reading.get("vision_escalation", False))),
            ("Reading: Extract Claim Numbers", yn(reading.get("extract_claim_numbers", False))),
            ("", ""),
            ("Classification: Structured Output", yn(classification.get("structured_output", False))),
            ("Classification: Use Document Types", yn(classification.get("use_document_types", False))),
            ("Classification: Use Providers", yn(classification.get("use_providers", False))),
            ("Classification: Extract Recipient", yn(classification.get("extract_recipient", False))),
            ("Classification: Grounding Check", yn(classification.get("grounding_check", False))),
            ("Classification: Evidence Confidence", yn(classification.get("evidence_confidence", False))),
            ("Classification: Candidate Shortlist", yn(classification.get("use_candidate_shortlist", False))),
            ("", ""),
            ("Learning: Document Types", learning_mode_to_display(learning.get("document_types", "off"))),
            ("Learning: Client Relationships",
             learning_mode_to_display(learning.get("client_relationships", "off"))),
            ("Learning: Claim Linking", learning_mode_to_display(learning.get("claim_linking", "off"))),
            ("Learning: Log Corrections", yn(learning.get("log_corrections", True))),
            ("Learning: Retroactive Rename",
             retroactive_mode_to_display(learning.get("retroactive_rename", "off"))),
        ]

    @staticmethod
    def _save_xlsx(path: str, headers: list, rows: list, results: list = None,
                    config: dict = None):
        """Write an Excel workbook with auto-fitted columns and a styled header row.
        Rows where 'Audit: Wrong Client' is 'Yes' are highlighted in red — these
        are the most critical errors and must be easy to spot.
        If `results` is provided, a Summary sheet is added. If `config` is
        provided, a Configuration sheet stamps which feature flags were on
        for this run."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"

        # Locate the "Audit: Wrong Client" column index (0-based in the data row)
        try:
            wrong_client_idx = headers.index("Audit: Wrong Client")
        except ValueError:
            wrong_client_idx = -1

        # Header row
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="1F497D")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 18

        # Fill styles
        light     = PatternFill("solid", fgColor="EEF2F7")
        red_fill  = PatternFill("solid", fgColor="FFCCCC")  # wrong-client flag
        red_font  = Font(bold=True, color="990000")

        for i, row in enumerate(rows, start=2):
            ws.append(row)
            # Detect wrong-client flag
            is_wrong_client = (
                wrong_client_idx >= 0
                and len(row) > wrong_client_idx
                and row[wrong_client_idx] == "Yes"
            )
            if is_wrong_client:
                for cell in ws[i]:
                    cell.fill = red_fill
                    cell.font = red_font
            elif i % 2 == 0:
                for cell in ws[i]:
                    cell.fill = light

        # Auto-fit column widths (cap between 12 and 60 chars wide)
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            letter = col[0].column_letter
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)

        ws.freeze_panes = "A2"  # keep header visible while scrolling

        # ── Summary sheet ─────────────────────────────────────
        if results is not None:
            ss = wb.create_sheet("Summary", 0)  # insert as first tab

            total       = len(results)
            skipped     = sum(1 for r in results if r.status == "skipped")
            errors      = sum(1 for r in results if r.status == "error")
            identified  = sum(1 for r in results if r.status == "renamed")
            needs_rev   = sum(1 for r in results if r.status == "needs_review")
            processed   = total - skipped - errors  # docs that went through AI

            # Audit counts (only processed files can be audited)
            _was_audited = lambda r: (r.audit_correct or r.audit_wrong_client
                                      or r.audit_bad_description or r.audit_failed_client
                                      or r.audit_should_review)
            audited       = sum(1 for r in results if _was_audited(r))
            audit_correct = sum(1 for r in results if r.audit_correct)
            audit_wrong   = sum(1 for r in results if r.audit_wrong_client)
            audit_bad_desc = sum(1 for r in results if r.audit_bad_description)
            audit_failed  = sum(1 for r in results if r.audit_failed_client)
            audit_should  = sum(1 for r in results if r.audit_should_review)

            # Identified-client subset audit
            id_results    = [r for r in results if r.status == "renamed"]
            id_audited    = sum(1 for r in id_results if _was_audited(r))
            id_correct    = sum(1 for r in id_results if r.audit_correct)

            def _pct(num, denom):
                return f"{num / denom * 100:.1f}%" if denom else "N/A"

            summary_data = [
                ("Documents Total",                     total),
                ("Documents Processed",                 processed),
                ("Documents Skipped",                   skipped),
                ("Documents With Errors",               errors),
                ("Documents Identified a Client",       identified),
                ("Documents Marked Needs Review",       needs_rev),
                ("", ""),
                ("Audit Completion (All Processed)",    _pct(audited, processed)),
                ("Audit Completion (Identified Only)",  _pct(id_audited, identified)),
                ("", ""),
                ("Success Rate (All Audited)",          _pct(audit_correct, audited)),
                ("Success Rate (Identified Only)",      _pct(id_correct, id_audited)),
                ("", ""),
                ("Audit Flags",                         ""),
                ("  Marked Correct",                    audit_correct),
                ("  Wrong Client",                      audit_wrong),
                ("  Bad Description",                   audit_bad_desc),
                ("  Failed to Identify Client",         audit_failed),
                ("  Should Have Been Flagged",          audit_should),
            ]

            # Style the summary sheet
            title_font = Font(bold=True, size=14, color="1F497D")
            label_font = Font(bold=True, size=11)
            value_font = Font(size=11)
            ss.append(["Audit Summary"])
            ss["A1"].font = title_font
            ss.append([])  # blank row

            for label, value in summary_data:
                ss.append([label, value])
            for row_cells in ss.iter_rows(min_row=3, max_row=ss.max_row, max_col=2):
                row_cells[0].font = label_font
                row_cells[1].font = value_font
                row_cells[1].alignment = Alignment(horizontal="right")

            ss.column_dimensions["A"].width = 42
            ss.column_dimensions["B"].width = 18

        # ── Configuration sheet — what was turned on for this run ──
        if config is not None:
            cs = wb.create_sheet("Configuration")
            title_font = Font(bold=True, size=14, color="1F497D")
            label_font = Font(bold=True, size=11)
            value_font = Font(size=11)
            cs.append(["Configuration"])
            cs["A1"].font = title_font
            cs.append([])
            for label, value in ScandocsApp._config_flags_snapshot(config):
                cs.append([label, value])
            for row_cells in cs.iter_rows(min_row=3, max_row=cs.max_row, max_col=2):
                row_cells[0].font = label_font
                row_cells[1].font = value_font
            cs.column_dimensions["A"].width = 38
            cs.column_dimensions["B"].width = 22

        wb.save(path)

    @staticmethod
    def _save_csv(path: str, headers: list, rows: list, config: dict = None):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for row in rows:
                w.writerow(row)
            if config is not None:
                w.writerow([])
                w.writerow(["Configuration for this run"])
                for label, value in ScandocsApp._config_flags_snapshot(config):
                    w.writerow([label, value])

    def _write_report(self, folder: str) -> str:
        """Write results to a timestamped report in `folder`. Returns the saved path."""
        os.makedirs(folder, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
        headers = self._REPORT_HEADERS
        rows = self._results_as_rows()
        config = self.config_mgr.config
        if _XLSX_AVAILABLE:
            path = os.path.join(folder, f"scandocs_report_{ts}.xlsx")
            self._save_xlsx(path, headers, rows, results=self._results, config=config)
        else:
            path = os.path.join(folder, f"scandocs_report_{ts}.csv")
            self._save_csv(path, headers, rows, config=config)
        return path


    def _add_result_row(self, result: ProcessResult):
        tag = result.status
        label = {
            "renamed":      "OK",
            "needs_review": "REVIEW",
            "skipped":      "Skipped",
            "error":        "ERROR",
        }.get(result.status, result.status)
        if result.was_dry_run and result.status in ("renamed", "needs_review"):
            label += " (preview)"

        # skip_reason takes priority — it's the whole point of a skipped
        # row — then the resolved client, then the error message.
        client_cell = (
            result.skip_reason if result.skip_reason
            else result.client if result.client
            else (result.error_message or "")[:60]
        )
        iid = self.results_tree.insert(
            "", tk.END,
            values=(
                "",   # audited checkmark — filled in by _on_audit_check
                result.original_name,
                result.final_name,
                label,
                client_cell,
                "",   # new_location — filled in by _fo_do_move
                "",   # correction — covered by a real overlaid button, see below
            ),
            tags=(tag,),
        )
        self._iid_to_result[iid] = result
        self._results.append(result)
        self._make_correction_button(iid)
        self.results_tree.yview_moveto(1.0)
        self._reposition_all_correction_buttons()

    def _make_correction_button(self, iid: str):
        """Create the blue 'Manual Correction' button overlaid on this row's
        correction cell. ttk.Treeview can't style an individual cell, so a
        real button is placed on top of it instead — see
        _position_correction_button for how it tracks the row."""
        btn = tk.Button(
            self.results_tree,
            text="✎  Manual Correction",
            command=lambda i=iid: self._open_manual_correction(i),
            bg="#1565c0", fg="white",
            activebackground="#0d47a1", activeforeground="white",
            disabledforeground="#cfd8dc",
            relief="solid", bd=1,
            font=(APP_FONT, 8, "bold"),
            cursor="hand2",
        )
        self._correction_buttons[iid] = btn
        self._position_correction_button(iid)

    def _position_correction_button(self, iid: str):
        """Move/show/hide the given row's correction button to track the
        treeview's current scroll position and column layout."""
        btn = self._correction_buttons.get(iid)
        if btn is None or not btn.winfo_exists():
            return
        bbox = self.results_tree.bbox(iid, "correction")
        if not bbox:
            btn.place_forget()
            return
        x, y, w, h = bbox
        btn.place(x=x + 2, y=y + 2, width=max(w - 4, 10), height=max(h - 4, 10))

    def _reposition_all_correction_buttons(self):
        for iid in list(self._correction_buttons.keys()):
            self._position_correction_button(iid)

    # ── Audit helpers ─────────────────────────────────────────

    @staticmethod
    def _is_audited(result: ProcessResult) -> bool:
        return (result.audit_correct or result.audit_wrong_client or
                result.audit_bad_description or result.audit_failed_client or
                result.audit_should_review)

    def _audit_mode_on(self) -> bool:
        return self.s_audit_mode_var.get()

    def _apply_extraction_method_ui(self):
        """Enable/disable the extraction-method dropdown based on whether the
        currently selected model is vision-capable. Also greys out the
        Max Vision Pages entry when OCR mode is selected."""
        if not hasattr(self, "s_method_combo"):
            return
        model = self.s_model_var.get().strip()
        vision_ok = model_supports_vision(model)

        if vision_ok:
            self.s_method_combo.configure(
                values=["Use OCR", "Use Vision Model"], state="readonly"
            )
            self._method_hint_lbl.configure(
                text=f"  '{model}' supports vision — both modes available."
            )
        else:
            # Force OCR and lock the dropdown
            if self.s_extraction_method_var.get() == "Use Vision Model":
                self.s_extraction_method_var.set("Use OCR")
            self.s_method_combo.configure(
                values=["Use OCR"], state="disabled"
            )
            if model:
                self._method_hint_lbl.configure(
                    text=f"  '{model}' is not on the vision allowlist "
                         f"(llama3.2-vision, gemma4). Vision mode disabled."
                )
            else:
                self._method_hint_lbl.configure(
                    text="  Select a model first. Vision mode requires "
                         "llama3.2-vision or gemma4."
                )

        # Grey out Max Vision Pages when not in vision mode
        in_vision = self.s_extraction_method_var.get() == "Use Vision Model"
        if hasattr(self, "s_vision_pages_entry"):
            self.s_vision_pages_entry.configure(
                state=("normal" if in_vision else "disabled")
            )

    def _apply_file_mode(self):
        """Show or hide File Mode panels based on the master toggle and per-tab sub-settings."""
        master_on = self.s_file_mode_var.get()
        auto_on   = master_on and self.s_file_mode_auto_var.get()
        manual_on = master_on and self.s_file_mode_manual_var.get()

        # Show / hide the per-tab sub-checkboxes
        if hasattr(self, "_file_mode_sub_frame"):
            if master_on:
                self._file_mode_sub_frame.grid()
            else:
                self._file_mode_sub_frame.grid_remove()

        # Suggest Location and Auto-commit are always visible but permanently disabled (Coming Soon)

        # Process tab — Move Files panel
        if auto_on:
            self.file_ops_panel.pack(fill=tk.X, padx=10, pady=(6, 10))
        else:
            self.file_ops_panel.pack_forget()

        # Manual Correction panel — destination row
        if hasattr(self, "_corr_dest_widgets"):
            for w in self._corr_dest_widgets:
                if manual_on:
                    w.grid()
                else:
                    w.grid_remove()

        # Legacy Manual Entry tab — destination row
        if hasattr(self, "_review_dest_widgets"):
            for w in self._review_dest_widgets:
                if manual_on:
                    w.grid()
                else:
                    w.grid_remove()

    def _browse_suggest_parent(self):
        folder = filedialog.askdirectory(
            title="Select Client Folders Parent Directory",
            initialdir=self.s_suggest_parent_var.get() or SCRIPT_DIR,
        )
        if folder:
            self.s_suggest_parent_var.set(os.path.normpath(folder))

    @staticmethod
    def _suggest_location(client: str, description: str, parent_folder: str) -> Optional[str]:
        """Find the single client folder in parent_folder that matches client, then pick
        the best subfolder based on the document description.
        Returns a path string, or None if no unambiguous match is found.
        """
        if not client or not parent_folder or not os.path.isdir(parent_folder):
            return None

        try:
            all_dirs = [
                d for d in os.listdir(parent_folder)
                if os.path.isdir(os.path.join(parent_folder, d))
            ]
        except Exception as e:
            logging.warning(f"Suggest location: could not list {parent_folder}: {e}")
            return None

        # Match the client name against folder names — look for the last name
        # (before the comma) plus partial first name to avoid false positives
        client_lower = client.lower()
        last_name = client.split(",")[0].strip().lower()
        first_part = client.split(",")[1].strip().lower().split()[0] if "," in client else ""

        matching_dirs = []
        for d in all_dirs:
            d_lower = d.lower()
            # Must contain the last name
            if last_name not in d_lower:
                continue
            # If we have a first name part, folder should contain it too
            if first_part and first_part not in d_lower:
                continue
            # Fuzzy fallback: ratio against the full client name
            ratio = difflib.SequenceMatcher(None, client_lower, d_lower).ratio()
            if ratio >= 0.7 or (last_name in d_lower and (not first_part or first_part in d_lower)):
                matching_dirs.append(d)

        if len(matching_dirs) != 1:
            # 0 matches → no folder found; 2+ matches → ambiguous, don't guess
            logging.debug(
                f"Suggest location: {len(matching_dirs)} matches for '{client}' "
                f"in {parent_folder} — skipping"
            )
            return None

        client_folder = os.path.join(parent_folder, matching_dirs[0])

        # Look for best subfolder using description keywords
        try:
            subfolders = [
                d for d in os.listdir(client_folder)
                if os.path.isdir(os.path.join(client_folder, d))
            ]
        except Exception:
            return client_folder

        if not subfolders:
            return client_folder

        desc_words = [w.lower() for w in re.split(r"\W+", description) if len(w) >= 3]
        if not desc_words:
            return client_folder

        best_sub = None
        best_score = 0
        for sub in subfolders:
            sub_lower = sub.lower()
            score = sum(1 for w in desc_words if w in sub_lower)
            if score > best_score:
                best_score = score
                best_sub = sub

        if best_sub and best_score >= 1:
            return os.path.join(client_folder, best_sub)
        return client_folder

    def _fo_browse_dest(self):
        folder = filedialog.askdirectory(
            title="Select Destination Folder",
            initialdir=self.fo_dest_var.get() or SCRIPT_DIR,
        )
        if folder:
            self.fo_dest_var.set(os.path.normpath(folder))
            # Persist immediately
            cfg = self.config_mgr.config
            cfg["processing"]["file_mode_destination"] = self.fo_dest_var.get()
            self.config_mgr.save(cfg)

    def _fo_resolve_dest(self) -> str | None:
        """Return the destination path if valid, else show an error and return None."""
        dest = self.fo_dest_var.get().strip()
        if not dest:
            messagebox.showwarning(
                "No Destination",
                "Please select a destination folder before moving files."
            )
            return None
        if not os.path.isdir(dest):
            create = messagebox.askyesno(
                "Folder Not Found",
                f"The destination folder does not exist:\n{dest}\n\nCreate it?",
            )
            if not create:
                return None
            try:
                os.makedirs(dest, exist_ok=True)
            except Exception as e:
                messagebox.showerror("Error", f"Could not create folder:\n{e}")
                return None
        return dest

    def _fo_do_move(self, result: ProcessResult, dest: str, iid: str) -> bool:
        """Move a single file to dest. Returns True on success (including a
        successful *preview* when dry-run is on — nothing is touched)."""
        dry_run = self.config_mgr.config.get("automation", {}).get("dry_run", False)
        if dry_run:
            self.status_var.set("Preview mode is on — no files were moved.")
            vals = list(self.results_tree.item(iid, "values"))
            vals[5] = f"{os.path.basename(dest)} (preview)"
            self.results_tree.item(iid, values=vals)
            result.was_dry_run = True
            return True

        if not os.path.isdir(dest):
            try:
                os.makedirs(dest, exist_ok=True)
            except Exception as e:
                messagebox.showerror("Error", f"Could not create destination folder:\n{e}")
                return False
        scandocs = self.config_mgr.config["paths"]["scandocs_folder"]
        src = os.path.join(scandocs, result.final_name)
        if not os.path.isfile(src):
            messagebox.showwarning(
                "File Not Found",
                f"Could not locate:\n{result.final_name}\n\n"
                "It may have already been moved or renamed."
            )
            return False
        # Collision avoidance at destination
        dst_name = result.final_name
        base, ext = os.path.splitext(dst_name)
        counter = 1
        dst = os.path.join(dest, dst_name)
        while os.path.exists(dst):
            dst_name = f"{base} ({counter}){ext}"
            dst = os.path.join(dest, dst_name)
            counter += 1
        try:
            import shutil
            shutil.move(src, dst)
            result.moved_to = dst
            if self.config_mgr.config.get("safety", {}).get("undo_log", True):
                log_rename(getattr(self.engine, "batch_id", "") or "manual",
                           "move", src, dst, "move")
            # Update the New Location column (index 5) in the treeview row
            vals = list(self.results_tree.item(iid, "values"))
            vals[5] = os.path.basename(dest)
            self.results_tree.item(iid, values=vals, tags=("moved",))
            return True
        except Exception as e:
            messagebox.showerror("Move Failed", f"Could not move file:\n{e}")
            return False

    def _fo_apply_to_selected(self):
        """Stage the current destination for all selected rows without moving them yet.
        The staged location is shown in the New Location column as '(pending)'."""
        dest = self.fo_dest_var.get().strip()
        if not dest:
            self.fo_status_var.set("Set a destination folder first.")
            return
        sel = self.results_tree.selection()
        if not sel:
            self.fo_status_var.set("Select one or more files first.")
            return
        folder_label = os.path.basename(dest) or dest
        audit_on = self._audit_mode_on()
        count = 0
        skipped_audit = 0
        for iid in sel:
            result = self._iid_to_result.get(iid)
            if result and result.status == "renamed" and not result.moved_to:
                # In audit mode a file must be marked Correct before it can be moved
                if audit_on and not result.audit_correct:
                    skipped_audit += 1
                    continue
                result.pending_dest = dest
                vals = list(self.results_tree.item(iid, "values"))
                vals[5] = f"{folder_label} (pending)"
                self.results_tree.item(iid, values=vals)
                count += 1
        if count:
            msg = f"{count} file{'s' if count != 1 else ''} staged — click Move Files to commit."
            if skipped_audit:
                msg += f"  ({skipped_audit} skipped — not marked Correct in audit.)"
            self.fo_status_var.set(msg)
        else:
            if skipped_audit:
                self.fo_status_var.set(
                    f"No files staged — {skipped_audit} file(s) must be marked Correct in audit first.")
            else:
                self.fo_status_var.set(
                    "No eligible files selected (files must be renamed and not yet moved).")

    def _fo_move_all(self):
        """Move all files that have been staged via Apply to Selected.
        In audit mode, only files marked Correct are eligible."""
        audit_on = self._audit_mode_on()
        moveable = [
            (iid, r) for iid, r in self._iid_to_result.items()
            if r.status == "renamed" and not r.moved_to and r.pending_dest
            and (not audit_on or r.audit_correct)
        ]
        if not moveable:
            self.fo_status_var.set("No files staged — select files and click Apply to Selected first.")
            return
        moved, failed = 0, 0
        for iid, result in moveable:
            if self._fo_do_move(result, result.pending_dest, iid):
                moved += 1
            else:
                failed += 1
        msg = f"Moved {moved} file{'s' if moved != 1 else ''}."
        if failed:
            msg += f"  {failed} failed."
        self.fo_status_var.set(msg)

    # ── Palette & rounding ────────────────────────────────────────────────────

    def _apply_default_styling(self):
        """Apply fixed default blue styling across all UI elements."""
        s = self.style
        primary = _APP_PRIMARY
        light   = _APP_LIGHT
        mid     = _APP_MID

        # Primary buttons
        s.configure("primary.TButton",
            background=primary, bordercolor=primary,
            darkcolor=primary,  lightcolor=primary,
            foreground="#ffffff",
        )
        s.map("primary.TButton",
            background=[("active !disabled", mid),
                        ("pressed !disabled", primary),
                        ("disabled", "#cccccc")],
            bordercolor=[("active !disabled", mid),
                         ("focus !disabled", mid)],
            darkcolor=[("active !disabled", mid),
                       ("pressed !disabled", primary)],
            lightcolor=[("active !disabled", mid),
                        ("pressed !disabled", primary)],
        )

        # Outline buttons
        s.configure("primary-outline.TButton",
            foreground=primary, bordercolor=primary,
        )
        s.map("primary-outline.TButton",
            background=[("active !disabled", light)],
            foreground=[("active !disabled", primary)],
        )

        # Progressbar
        s.configure("primary.Horizontal.TProgressbar",
            troughcolor="#e0e0e0", background=primary)
        s.configure("Horizontal.TProgressbar",
            troughcolor="#e0e0e0", background=primary)

        # Entry & Combobox focus highlight
        s.map("TEntry",
            bordercolor=[("focus !disabled", primary), ("hover !disabled", mid)],
            lightcolor=[("focus !disabled", light)],
            darkcolor=[("focus !disabled", light)],
        )
        s.map("TCombobox",
            bordercolor=[("focus !disabled", primary), ("hover !disabled", mid)],
            lightcolor=[("focus !disabled", light)],
        )

        # Notebook active tab
        s.map("TNotebook.Tab",
            background=[("selected", light), ("active", "#f8f9fa")],
            foreground=[("selected", primary)],
        )

        # LabelFrame title
        s.configure("TLabelframe.Label", foreground=primary)

        # Checkbutton / Radiobutton indicators
        s.map("TCheckbutton",  indicatorcolor=[("selected", primary)])
        s.map("Checkbutton",   indicatorcolor=[("selected", primary)])
        s.map("TRadiobutton",  indicatorcolor=[("selected", primary)])

        # Scrollbar thumb
        s.configure("Vertical.TScrollbar",   troughcolor="#f0f0f0", background=mid)
        s.configure("Horizontal.TScrollbar", troughcolor="#f0f0f0", background=mid)

    def _apply_round_styling(self):
        """Apply soft visual tweaks: padded entries/tabs and DWM rounded window corners."""
        s = self.style

        s.configure("TEntry",        padding=[8, 6])
        s.configure("TCombobox",     padding=[8, 6])
        s.configure("TNotebook.Tab", padding=[14, 6])
        s.configure("TButton",       padding=[10, 6])

        self._apply_default_styling()

        # Rounded window corners (Windows 11 DWM — silent no-op elsewhere)
        self.after(200, self._try_round_window)

    # ── Scroll helper ─────────────────────────────────────────

    @staticmethod
    def _bind_mousewheel(widget, scroll_fn):
        """Recursively bind mousewheel/trackpad scroll to widget and all descendants."""
        widget.bind("<MouseWheel>",  lambda e: scroll_fn(e), add="+")
        widget.bind("<Button-4>",    lambda e: scroll_fn(e), add="+")  # Linux scroll-up
        widget.bind("<Button-5>",    lambda e: scroll_fn(e), add="+")  # Linux scroll-down
        for child in widget.winfo_children():
            ScandocsApp._bind_mousewheel(child, scroll_fn)

    def _try_round_window(self):
        """Ask Windows 11 DWM to use rounded window corners."""
        try:
            import ctypes
            DWMWA_WINDOW_CORNER_PREFERENCE = 33
            DWMWCP_ROUND = 2
            hwnd = self.winfo_id()
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd,
                DWMWA_WINDOW_CORNER_PREFERENCE,
                ctypes.byref(ctypes.c_int(DWMWCP_ROUND)),
                ctypes.sizeof(ctypes.c_int))
        except Exception:
            pass  # Not Windows 11 or DWM unavailable — ignore

    def _apply_audit_mode(self):
        """Show or hide all audit UI based on the current Audit Mode setting."""
        on = self._audit_mode_on()
        # ✓ column in treeview
        if on:
            self.results_tree["displaycolumns"] = (
                "audited", "original", "new_name", "status", "client",
                "new_location", "correction")
        else:
            self.results_tree["displaycolumns"] = (
                "original", "new_name", "status", "client",
                "new_location", "correction")
        if hasattr(self, "_correction_buttons"):
            self.results_tree.after_idle(self._reposition_all_correction_buttons)
        # Audit panel
        if on:
            self.audit_panel.pack(fill=tk.X, padx=10, pady=(6, 0))
        else:
            self.audit_panel.pack_forget()
        # Top-right button: Submit Audit in audit mode, Open Report otherwise
        if on:
            self.btn_open_report.pack_forget()
            self.btn_submit_audit.pack(anchor="e")
        else:
            self.btn_submit_audit.pack_forget()
            self.btn_open_report.pack(anchor="e")

    def _check_audit_complete(self):
        """If all auditable rows are reviewed AND processing is done, prompt to submit."""
        if not self._audit_mode_on() or not self._results:
            return
        if getattr(self, "_processing_active", False):
            return  # Still processing — re-checked automatically when processing finishes
        auditable = [r for r in self._results if r.status in ("renamed", "needs_review")]
        if not auditable:
            return
        if all(self._is_audited(r) for r in auditable):
            folder = self.s_report_folder_var.get().strip() or DEFAULT_REPORTS_FOLDER
            try:
                self._write_report(folder)
            except Exception as e:
                messagebox.showerror("Report Error", f"Could not save report:\n{e}")
                return
            # Disable the submit button so it can't be double-submitted via auto-path
            self.btn_submit_audit.config(state=tk.DISABLED, text="Audit Saved")
            # Custom prompt — two explicit choices instead of a plain showinfo
            dlg = tk.Toplevel(self)
            dlg.title("Audit Complete")
            dlg.resizable(False, False)
            dlg.transient(self)
            dlg.grab_set()
            # Centre over the main window
            self.update_idletasks()
            x = self.winfo_x() + (self.winfo_width() - 380) // 2
            y = self.winfo_y() + (self.winfo_height() - 150) // 2
            dlg.geometry(f"380x150+{x}+{y}")
            ttk.Label(
                dlg,
                text="You finished the audit.\nWould you like to submit the audit now?",
                font=(APP_FONT, 11),
                anchor="center",
                justify="center",
            ).pack(pady=(24, 16))
            btn_row = ttk.Frame(dlg)
            btn_row.pack()
            def _do_submit():
                dlg.destroy()
                self.btn_submit_audit.config(state=tk.NORMAL, text="Submit Audit")
                self._submit_audit()
            def _make_changes():
                dlg.destroy()
                self.btn_submit_audit.config(state=tk.NORMAL, text="Submit Audit")
            ttk.Button(btn_row, text="Submit Audit", bootstyle="primary",
                       command=_do_submit).pack(side=tk.LEFT, padx=(0, 12))
            ttk.Button(btn_row, text="Make Changes", bootstyle="secondary-outline",
                       command=_make_changes).pack(side=tk.LEFT)

    def _on_close(self):
        """Warn if audit mode is on and items are still unreviewed."""
        if self._audit_mode_on() and self._results:
            auditable = [r for r in self._results
                         if r.status in ("renamed", "needs_review")]
            unreviewed = [r for r in auditable if not self._is_audited(r)]
            if unreviewed:
                n = len(unreviewed)
                answer = messagebox.askyesno(
                    "Audit Incomplete",
                    f"{n} file{'s' if n != 1 else ''} still "
                    f"{'have' if n != 1 else 'has'} not been audited.\n\n"
                    "You should complete the audit before closing so the report\n"
                    "includes accurate quality data.\n\n"
                    "Close anyway?",
                    icon="warning",
                )
                if not answer:
                    return
        self.destroy()

    def _sort_treeview(self, col: str):
        """Sort the results treeview by *col*, toggling A→Z / Z→A on repeated clicks."""
        col_index = {
            "audited": 0, "original": 1, "new_name": 2,
            "status": 3, "client": 4, "new_location": 5,
        }
        col_labels = {
            "audited": "✓", "original": "Original File", "new_name": "New Name",
            "status": "Status", "client": "Client",
            "new_location": "New Location",
        }
        sortable = {"original", "new_name", "status", "client"}

        if self._sort_col == col:
            self._sort_reverse = not self._sort_reverse
        else:
            self._sort_col = col
            self._sort_reverse = False

        idx = col_index[col]
        items = [(self.results_tree.item(iid, "values"), iid)
                 for iid in self.results_tree.get_children()]
        items.sort(key=lambda x: (x[0][idx] or "").lower(),
                   reverse=self._sort_reverse)
        for i, (_, iid) in enumerate(items):
            self.results_tree.move(iid, "", i)
        self._reposition_all_correction_buttons()

        # Update heading arrows
        arrow = " ↓" if self._sort_reverse else " ↑"
        for c, lbl in col_labels.items():
            if c not in sortable:
                continue
            self.results_tree.heading(
                c, text=lbl + (arrow if c == col else ""),
                command=lambda _c=c: self._sort_treeview(_c))

    def _on_result_select(self, _event=None):
        """Update the audit panel when the user selects a row."""
        sel = self.results_tree.selection()
        if not sel:
            return
        iid = sel[0]
        result = self._iid_to_result.get(iid)
        if not result:
            return

        # If the Manual Correction panel is open on a DIFFERENT row and the
        # user clicks a row directly in the table (rather than using the
        # panel's own Next/Close buttons), the panel used to stay silently
        # bound to the old row while the table showed the new one selected
        # -- any edit made after that was then applied to the wrong file.
        # Commit the row being left, then switch the panel to the newly
        # selected one, exactly like the "Next" button does.
        #
        # If the commit doesn't validate, put the selection back on the row
        # under correction instead of switching. Clicking around a results
        # table is browsing, not saving, so an unfinished edit must not be
        # silently dropped just because the user looked at another row.
        if (self.correction_panel.winfo_ismapped()
                and self._correction_iid and self._correction_iid != iid):
            leaving_iid = self._correction_iid
            if not self._corr_commit():
                self.results_tree.selection_set(leaving_iid)
                return
            self._open_manual_correction(iid)

        # Update file label
        self.audit_file_label.config(
            text=result.final_name or result.original_name, foreground="#212529"
        )
        # Enable controls
        for w in (self.audit_open_btn, self.audit_prev_btn, self.audit_next_btn,
                  self.audit_correct_chk, self.audit_wrong_client_chk,
                  self.audit_bad_desc_chk, self.audit_failed_client_chk,
                  self.audit_should_flag_chk):
            w.config(state=tk.NORMAL)

        # Sync checkboxes without triggering callbacks
        self._audit_updating = True
        self.audit_correct_var.set(result.audit_correct)
        self.audit_wrong_client_var.set(result.audit_wrong_client)
        self.audit_bad_desc_var.set(result.audit_bad_description)
        self.audit_failed_client_var.set(result.audit_failed_client)
        self.audit_should_flag_var.set(result.audit_should_review)
        self._audit_updating = False

        # Sync the rename hint
        self._update_audit_rename_hint(result)

        # If the preview popup is open, refresh it with the newly selected file
        if self._file_popup is not None and self._file_popup.winfo_exists():
            scandocs = self.config_mgr.config["paths"]["scandocs_folder"]
            path = os.path.join(scandocs, result.final_name)
            if not os.path.isfile(path):
                path = os.path.join(scandocs, result.original_name)
            if os.path.isfile(path):
                self._refresh_popup_content(path)


    def _audit_open_file(self):
        """Open the currently selected file in a popup viewer."""
        sel = self.results_tree.selection()
        if not sel:
            return
        result = self._iid_to_result.get(sel[0])
        if not result:
            return
        scandocs = self.config_mgr.config["paths"]["scandocs_folder"]
        path = os.path.join(scandocs, result.final_name)
        if not os.path.isfile(path):
            path = os.path.join(scandocs, result.original_name)
        if os.path.isfile(path):
            self._open_file_popup(path)
        else:
            messagebox.showwarning("File Not Found",
                f"Could not locate the file:\n{result.final_name}")

    def _update_audit_rename_hint(self, result: ProcessResult):
        """Update the italic rename-hint label below the audit checkboxes.
        Wrong client name → client portion becomes A-NEEDS REVIEW.
        Bad description   → description portion becomes Scanned Document.
        Both flags can apply simultaneously."""
        any_bad = (result.audit_wrong_client or result.audit_bad_description
                   or result.audit_failed_client or result.audit_should_review)
        if not any_bad:
            self.audit_rename_hint_var.set("")
            return

        fname = result.final_name
        base, ext = os.path.splitext(fname)
        if " - " in base:
            orig_client, orig_desc = base.split(" - ", 1)
        else:
            orig_client, orig_desc = base, ""

        # Which part changes?
        client_bad = result.audit_wrong_client or result.audit_failed_client or result.audit_should_review
        no_client_label = self.config_mgr.config.get("naming", {}).get("no_client_label", "A-NEEDS REVIEW")
        new_client = no_client_label if client_bad else orig_client
        new_desc   = "Scanned Document" if result.audit_bad_description else orig_desc

        proposed = f"{new_client} - {new_desc}{ext}" if new_desc else f"{new_client}{ext}"
        self.audit_rename_hint_var.set(
            f"This document will be renamed \"{proposed}\" after the audit is complete.")

    def _on_audit_check(self, flag: str):
        """Called when an audit checkbox is toggled."""
        if self._audit_updating:
            return
        sel = self.results_tree.selection()
        if not sel:
            return
        iid = sel[0]
        result = self._iid_to_result.get(iid)
        if not result:
            return

        if flag == "correct":
            result.audit_correct = self.audit_correct_var.get()
            # Correct is mutually exclusive with all problem flags — clear them
            if result.audit_correct:
                result.audit_wrong_client    = False
                result.audit_bad_description = False
                result.audit_failed_client   = False
                result.audit_should_review   = False
                self._audit_updating = True
                self.audit_wrong_client_var.set(False)
                self.audit_bad_desc_var.set(False)
                self.audit_failed_client_var.set(False)
                self.audit_should_flag_var.set(False)
                self._audit_updating = False
        elif flag == "wrong_client":
            result.audit_wrong_client = self.audit_wrong_client_var.get()
        elif flag == "bad_description":
            result.audit_bad_description = self.audit_bad_desc_var.get()
        elif flag == "failed_client":
            result.audit_failed_client = self.audit_failed_client_var.get()
        elif flag == "should_review":
            result.audit_should_review = self.audit_should_flag_var.get()

        # Orange if any problem flag set; keep green if marked correct; else original
        any_flagged = (result.audit_wrong_client or result.audit_bad_description
                       or result.audit_failed_client or result.audit_should_review)
        if any_flagged:
            tags = ("audited",)
        elif result.audit_correct:
            tags = ("renamed",)
        else:
            tags = (result.status,)
        self.results_tree.item(iid, tags=tags)

        # Update the ✓ column
        is_audited = self._is_audited(result)
        vals = list(self.results_tree.item(iid, "values"))
        vals[0] = "✓" if is_audited else ""
        self.results_tree.item(iid, values=vals)

        # Check if all auditable rows are now done
        self._check_audit_complete()

        # Update the rename hint below the checkboxes
        self._update_audit_rename_hint(result)

    def _show_correction_dialog(self, iid: str, result: ProcessResult):
        """Modal dialog asking the employee what the file should be named."""
        dialog = tk.Toplevel(self)
        dialog.title("Correct File Name")
        dialog.resizable(False, False)
        dialog.grab_set()   # modal
        dialog.focus_set()

        # Centre over the main window
        self.update_idletasks()
        x = self.winfo_x() + self.winfo_width()  // 2 - 260
        y = self.winfo_y() + self.winfo_height() // 2 - 80
        dialog.geometry(f"520x160+{x}+{y}")

        ttk.Label(dialog, text="What should this file be named?",
                  font=(APP_FONT, 10, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Label(dialog, text="Include the file extension  (e.g. GARCIA, Maria - Invoice.pdf)",
                  foreground="gray", font=(APP_FONT, 8)).pack(anchor="w", padx=16)

        entry_var = tk.StringVar(value=result.audit_corrected_name or result.final_name)
        entry = ttk.Entry(dialog, textvariable=entry_var, width=62)
        entry.pack(padx=16, pady=(6, 0), fill=tk.X)
        entry.select_range(0, tk.END)
        entry.focus_set()

        def _submit():
            corrected = entry_var.get().strip()
            if corrected:
                result.audit_corrected_name = corrected
                # Update the New Name cell in the treeview to show the correction
                vals = list(self.results_tree.item(iid, "values"))
                vals[1] = f"{corrected}  ✎"
                self.results_tree.item(iid, values=vals)
            dialog.destroy()

        def _cancel():
            # Uncheck whichever box triggered this if no correction was previously saved
            if not result.audit_corrected_name:
                self._audit_updating = True
                if result.audit_wrong_client and not self.audit_bad_desc_var.get():
                    self.audit_wrong_client_var.set(False)
                    result.audit_wrong_client = False
                elif result.audit_bad_description:
                    self.audit_bad_desc_var.set(False)
                    result.audit_bad_description = False
                self._audit_updating = False
                # Re-evaluate row colour
                any_flagged = (result.audit_wrong_client or result.audit_bad_description
                               or result.audit_should_review)
                tags = ("audited",) if any_flagged else (result.status,)
                self.results_tree.item(iid, tags=tags)
            dialog.destroy()

        btn_row = ttk.Frame(dialog)
        btn_row.pack(pady=(10, 0))
        ttk.Button(btn_row, text="Submit", bootstyle="primary",
                   command=_submit).pack(side=tk.LEFT, padx=6)
        ttk.Button(btn_row, text="Cancel", bootstyle="dark-outline",
                   command=_cancel).pack(side=tk.LEFT, padx=6)

        dialog.bind("<Return>", lambda e: _submit())
        dialog.bind("<Escape>", lambda e: _cancel())
        dialog.wait_window()

    # ── Audit navigation ──────────────────────────────────────

    def _audit_prev(self, _event=None):
        """Move selection to the previous row in the results tree."""
        children = self.results_tree.get_children()
        sel = self.results_tree.selection()
        if not children:
            return
        if not sel:
            self.results_tree.selection_set(children[-1])
            self.results_tree.see(children[-1])
        else:
            idx = children.index(sel[0])
            if idx > 0:
                self.results_tree.selection_set(children[idx - 1])
                self.results_tree.see(children[idx - 1])

    def _audit_next(self, _event=None):
        """Move selection to the next row in the results tree."""
        children = self.results_tree.get_children()
        sel = self.results_tree.selection()
        if not children:
            return
        if not sel:
            self.results_tree.selection_set(children[0])
            self.results_tree.see(children[0])
        else:
            idx = children.index(sel[0])
            if idx < len(children) - 1:
                self.results_tree.selection_set(children[idx + 1])
                self.results_tree.see(children[idx + 1])

    # ── File popup viewer ─────────────────────────────────────

    # ── Document popup viewer ─────────────────────────────────

    def _render_doc_image(self, path: str):
        """Render up to the first three PDF pages (or a JPEG) into a single PIL image.
        Returns None on any error so callers can fall back gracefully."""
        ext = os.path.splitext(path)[1].lower()
        try:
            import io
            if ext == ".pdf":
                doc = fitz.open(path)
                mat = fitz.Matrix(1.5, 1.5)
                n = min(doc.page_count, 3)
                imgs = [
                    PILImage.open(io.BytesIO(doc[i].get_pixmap(matrix=mat).tobytes("png")))
                    for i in range(n)
                ]
                doc.close()
                if len(imgs) == 1:
                    return imgs[0]
                # Stack pages vertically with a light-grey gap
                gap = 8
                w = max(p.width  for p in imgs)
                h = sum(p.height for p in imgs) + gap
                combined = PILImage.new("RGB", (w, h), color=(200, 200, 200))
                y = 0
                for p in imgs:
                    combined.paste(p, (0, y))
                    y += p.height + gap
                return combined
            elif ext in (".jpg", ".jpeg"):
                return PILImage.open(path)
        except Exception as e:
            logging.warning(f"Could not render {path}: {e}")
        return None

    def _open_file_popup(self, path: str):
        """Open the preview popup for *path*.
        If the popup is already visible, refresh its content instead of opening a second window."""
        if fitz is None or PILImage is None or PILImageTk is None:
            _open_file(path)
            return

        img = self._render_doc_image(path)
        if img is None:
            _open_file(path)
            return

        # Reuse existing window if it is still open
        if self._file_popup is not None and self._file_popup.winfo_exists():
            self._refresh_popup_content(path, img)
            return

        # ── Create new popup ──────────────────────────────────
        popup = tk.Toplevel(self)
        popup.title(os.path.basename(path))
        popup.resizable(True, True)

        self.update_idletasks()
        screen_w = self.winfo_screenwidth()
        screen_h = self.winfo_screenheight()
        ui_cfg = self.config_mgr.config.get("ui", {})
        saved_w = ui_cfg.get("preview_popup_width")  or 0
        saved_h = ui_cfg.get("preview_popup_height") or 0
        if saved_w > 0 and saved_h > 0:
            width, height = saved_w, saved_h
        else:
            # Default: large, centered popup rather than a small side panel
            width  = int(screen_w * 0.8)
            height = int(screen_h * 0.85)
        x = max(0, (screen_w - width) // 2)
        y = max(0, (screen_h - height) // 2)
        popup.geometry(f"{width}x{height}+{x}+{y}")

        def _save_popup_size(_event=None):
            try:
                w = popup.winfo_width()
                h = popup.winfo_height()
            except tk.TclError:
                return
            if w <= 1 or h <= 1:
                return
            self.config_mgr.config["ui"]["preview_popup_width"]  = w
            self.config_mgr.config["ui"]["preview_popup_height"] = h
            self.config_mgr.save()

        self._popup_resize_after_id = None

        def _on_popup_configure(event):
            if event.widget is not popup:
                return
            if self._popup_resize_after_id is not None:
                popup.after_cancel(self._popup_resize_after_id)
            self._popup_resize_after_id = popup.after(500, _save_popup_size)

        popup.bind("<Configure>", _on_popup_configure)

        canvas = tk.Canvas(popup, bg="white")
        vsb = ttk.Scrollbar(popup, orient="vertical",   command=canvas.yview)
        hsb = ttk.Scrollbar(popup, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side=tk.RIGHT,  fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        canvas.pack(fill=tk.BOTH, expand=True)

        photo = PILImageTk.PhotoImage(img)
        canvas.create_image(0, 0, anchor="nw", image=photo)
        canvas.configure(scrollregion=(0, 0, img.width, img.height))
        canvas.image = photo

        canvas.bind("<MouseWheel>",
                    lambda e: canvas.yview_scroll(-1 * (e.delta // 120), "units"))

        self._file_popup        = popup
        self._file_popup_canvas = canvas

        def _on_popup_close():
            if self._popup_resize_after_id is not None:
                popup.after_cancel(self._popup_resize_after_id)
                self._popup_resize_after_id = None
            _save_popup_size()
            self._file_popup        = None
            self._file_popup_canvas = None
            popup.destroy()

        popup.protocol("WM_DELETE_WINDOW", _on_popup_close)

        # Force keyboard focus onto the popup itself — otherwise, if it was
        # opened while an Entry field had focus (e.g. from the Manual
        # Correction panel), spacebar would keep typing into that field
        # instead of reaching the popup's <space>-to-close binding below.
        popup.lift()
        popup.focus_force()

        # Spacebar while the popup is focused closes it, mirroring the
        # tree/listbox toggle behaviour (the popup steals keyboard focus,
        # so the tree's own <space> binding never fires while it's open).
        popup.bind("<space>", lambda e: _on_popup_close())

        # Arrow keys while the popup is focused still navigate the active list
        # (results tree on the Process tab, or the legacy Manual Entry list).
        popup.bind("<Left>",  lambda e: self._popup_prev())
        popup.bind("<Right>", lambda e: self._popup_next())
        popup.bind("<Up>",    lambda e: self._popup_prev())
        popup.bind("<Down>",  lambda e: self._popup_next())

    def _refresh_popup_content(self, path: str, img=None):
        """Swap the canvas image in the already-open popup for a new file."""
        if self._file_popup is None or not self._file_popup.winfo_exists():
            return
        if img is None:
            img = self._render_doc_image(path)
        if img is None:
            return
        canvas = self._file_popup_canvas
        photo  = PILImageTk.PhotoImage(img)
        canvas.delete("all")
        canvas.create_image(0, 0, anchor="nw", image=photo)
        canvas.configure(scrollregion=(0, 0, img.width, img.height))
        canvas.image = photo          # prevent GC
        canvas.yview_moveto(0)        # scroll back to top for the new document
        self._file_popup.title(os.path.basename(path))

    def _toggle_file_popup(self, path: str):
        """Open the popup if it is closed; close it if it is already open."""
        if self._file_popup is not None and self._file_popup.winfo_exists():
            self._file_popup.destroy()
            self._file_popup        = None
            self._file_popup_canvas = None
        else:
            self._open_file_popup(path)

    def _popup_prev(self):
        """Arrow-key navigation while the preview popup is focused — routes to
        whichever list is currently on screen (Process tab or the legacy
        Manual Entry tab)."""
        current = self.notebook.select()
        if current == str(self._process_tab_frame):
            self._audit_prev()
        elif current == str(self._review_tab_frame):
            self._review_prev()

    def _popup_next(self):
        current = self.notebook.select()
        if current == str(self._process_tab_frame):
            self._audit_next()
        elif current == str(self._review_tab_frame):
            self._review_next()

    def _on_tree_return(self, _event=None):
        """Spacebar on the Auto-Process results tree: toggle the file viewer."""
        sel = self.results_tree.selection()
        if not sel:
            return "break"
        result = self._iid_to_result.get(sel[0])
        if not result:
            return "break"
        scandocs = self.config_mgr.config["paths"]["scandocs_folder"]
        path = os.path.join(scandocs, result.final_name)
        if not os.path.isfile(path):
            path = os.path.join(scandocs, result.original_name)
        if os.path.isfile(path):
            self._toggle_file_popup(path)
        return "break"   # prevent treeview default Enter behaviour

    def _on_tree_double_click(self, _event=None):
        """Double-click on the Auto-Process results tree: always open the file viewer."""
        sel = self.results_tree.selection()
        if not sel:
            return "break"
        result = self._iid_to_result.get(sel[0])
        if not result:
            return "break"
        scandocs = self.config_mgr.config["paths"]["scandocs_folder"]
        path = os.path.join(scandocs, result.final_name)
        if not os.path.isfile(path):
            path = os.path.join(scandocs, result.original_name)
        if os.path.isfile(path):
            self._open_file_popup(path)
        return "break"

    # ── Submit Audit ──────────────────────────────────────────

    def _submit_audit(self):
        """Apply audit-flagged renames and save the report."""
        if not self._results:
            messagebox.showinfo("No Results", "No results to submit. Run processing first.")
            return
        scandocs = self.config_mgr.config["paths"]["scandocs_folder"]
        dry_run = self.config_mgr.config.get("automation", {}).get("dry_run", False)
        errors = []

        for result in self._results:
            if result.status not in ("renamed", "needs_review"):
                continue
            if not (result.audit_wrong_client or result.audit_bad_description):
                continue

            current_name = result.final_name
            base, ext = os.path.splitext(current_name)

            # Determine new client segment
            m = re.match(r"^(.+?) - (.+)$", base)
            old_client = m.group(1) if m else base
            old_desc   = m.group(2) if m else ""

            no_client_label = self.config_mgr.config.get("naming", {}).get("no_client_label", "A-NEEDS REVIEW")
            new_client = no_client_label if result.audit_wrong_client else old_client
            new_desc   = "Scanned Document" if result.audit_bad_description else old_desc

            new_name = f"{new_client} - {new_desc}{ext}" if old_desc else f"{new_client}{ext}"

            if new_name == current_name:
                continue

            if dry_run:
                # Preview mode — do not touch disk.
                continue

            src = os.path.join(scandocs, current_name)
            if os.path.isfile(src):
                resolved = FileProcessor._resolve_collision(scandocs, new_name, current_name)
                try:
                    dst = os.path.join(scandocs, resolved)
                    os.rename(src, dst)
                    if self.config_mgr.config.get("safety", {}).get("undo_log", True):
                        log_rename(getattr(self.engine, "batch_id", "") or "manual",
                                   "rename", src, dst, "audit")
                    # PASS 6: the office doesn't really use Audit, but it's
                    # cheap to log here too in case they ever do.
                    self._log_correction_event(
                        result, result.client, result.description,
                        new_client, new_desc, source="audit",
                    )
                    result.final_name = resolved
                    result.audit_corrected_name = resolved
                    # Refresh the treeview row
                    for iid, r in self._iid_to_result.items():
                        if r is result:
                            vals = list(self.results_tree.item(iid, "values"))
                            vals[2] = resolved
                            self.results_tree.item(iid, values=vals)
                            break
                except Exception as e:
                    errors.append(f"{current_name}: {e}")

        # Build audit folder: ScandocsAudit_YYYY-MM-DD_NN
        import shutil
        report_root = self.s_report_folder_var.get().strip() or DEFAULT_REPORTS_FOLDER
        os.makedirs(report_root, exist_ok=True)
        today_str = datetime.datetime.now().strftime("%Y-%m-%d")
        audit_num = 1
        while True:
            audit_folder_name = f"ScandocsAudit_{today_str}_{audit_num:02d}"
            audit_folder = os.path.join(report_root, audit_folder_name)
            if not os.path.exists(audit_folder):
                break
            audit_num += 1
        os.makedirs(audit_folder)

        # Save report into the audit folder
        try:
            path = self._write_report(audit_folder)
        except Exception as e:
            messagebox.showerror("Report Error", f"Could not save report:\n{e}")
            return

        # Copy files marked as Wrong Client into the audit folder
        wrong_client_copies = []
        for result in self._results:
            if not result.audit_wrong_client:
                continue
            src_path = os.path.join(scandocs, result.final_name)
            if not os.path.isfile(src_path):
                src_path = os.path.join(scandocs, result.original_name)
            if os.path.isfile(src_path):
                try:
                    dest = os.path.join(audit_folder, os.path.basename(src_path))
                    shutil.copy2(src_path, dest)
                    wrong_client_copies.append(os.path.basename(src_path))
                except Exception as e:
                    errors.append(f"Copy {os.path.basename(src_path)}: {e}")

        # Grey out the button so it can't be submitted twice
        self.btn_submit_audit.config(state=tk.DISABLED, text="Audit Submitted ✓")

        msg = f"Audit submitted.\nReport saved to:\n{audit_folder_name}/"
        if dry_run:
            msg += "\n\nPreview mode is on — flagged files were not renamed."
        if wrong_client_copies:
            msg += f"\n\n{len(wrong_client_copies)} Wrong Client file(s) copied to audit folder."
        if errors:
            msg += f"\n\nWarnings ({len(errors)}):\n" + "\n".join(errors[:5])
        messagebox.showinfo("Audit Submitted", msg)
        self._refresh_review_tab()

    # ── Client combo filter ───────────────────────────────────

    def _filter_client_combo(self, _event=None):
        """Filter the visible client list as the user types in the entry field."""
        typed = self.corr_client_var.get().lower()
        self.corr_client_listbox.delete(0, tk.END)
        for name in self._all_clients:
            if not typed or typed in name.lower():
                self.corr_client_listbox.insert(tk.END, name)

    def _on_client_listbox_select(self, _event=None):
        """Clicking a name in the client list copies it to the entry field."""
        sel = self.corr_client_listbox.curselection()
        if sel:
            self.corr_client_var.set(self.corr_client_listbox.get(sel[0]))

    def _open_report(self):
        folder = self.s_report_folder_var.get().strip() or DEFAULT_REPORTS_FOLDER
        if not os.path.isdir(folder):
            messagebox.showinfo("No Reports", f"Reports folder not found:\n{folder}")
            return
        reports = sorted(
            [f for f in os.listdir(folder)
             if f.lower().endswith(".xlsx") or f.lower().endswith(".csv")],
            reverse=True,
        )
        if not reports:
            messagebox.showinfo(
                "No Reports",
                f"No reports found in:\n{folder}\n\nProcess some documents first.",
            )
            return
        os.startfile(os.path.join(folder, reports[0]))

    def _export_csv(self):
        if not self._results:
            messagebox.showinfo("No Data", "No results to export. Run processing first.")
            return
        ts = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
        init_dir = self.s_report_folder_var.get().strip() or DEFAULT_REPORTS_FOLDER
        if _XLSX_AVAILABLE:
            default_ext = ".xlsx"
            filetypes = [("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
            default_name = f"scandocs_report_{ts}.xlsx"
        else:
            default_ext = ".csv"
            filetypes = [("CSV files", "*.csv")]
            default_name = f"scandocs_report_{ts}.csv"
        path = filedialog.asksaveasfilename(
            title="Save Report",
            initialdir=init_dir,
            initialfile=default_name,
            defaultextension=default_ext,
            filetypes=filetypes,
        )
        if not path:
            return
        try:
            headers = self._REPORT_HEADERS
            rows = self._results_as_rows()
            config = self.config_mgr.config
            if path.lower().endswith(".xlsx") and _XLSX_AVAILABLE:
                self._save_xlsx(path, headers, rows, results=self._results, config=config)
            else:
                self._save_csv(path, headers, rows, config=config)
            messagebox.showinfo("Exported", f"Report saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export Failed", str(e))

    # ── First-run wizard ──────────────────────────────────────

    def _check_first_run(self):
        if not self.config_mgr.config["paths"]["scandocs_folder"]:
            self._run_first_run_wizard()

    def _run_first_run_wizard(self):
        messagebox.showinfo(
            "Welcome to Speedy Scandocs",
            "Let's get you set up.\n\nFirst, select your Scandocs folder "
            "(where the scanner drops files).",
        )
        scandocs = filedialog.askdirectory(title="Select Scandocs Folder")
        if scandocs:
            p = os.path.normpath(scandocs)
            self.config_mgr.config["paths"]["scandocs_folder"] = p
            self.s_scandocs_var.set(p)

        messagebox.showinfo(
            "Client List File",
            "Now select your client_list.txt file, or choose a location to create one.\n\n"
            "This file stores client names in LAST, First format.",
        )
        client_file = filedialog.asksaveasfilename(
            title="Select or Create Client List",
            initialfile="client_list.txt",
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt")],
        )
        if client_file:
            p = os.path.normpath(client_file)
            self.config_mgr.config["paths"]["client_list_file"] = p
            self.s_client_list_var.set(p)
            if not os.path.exists(client_file):
                open(client_file, "w").close()

        self.config_mgr.save()
        messagebox.showinfo(
            "Setup Complete",
            "Setup complete!\n\n"
            "Next steps:\n"
            "1. Go to the Client List tab → add your clients → Save.\n"
            "2. Go to Settings → verify API URLs → Test Connection.\n"
            "3. Return to Process → click Auto-Process Documents.",
        )
        self.notebook.select(self._settings_tab_frame)

    # ── Auto-update ───────────────────────────────────────────

    def _maybe_check_for_updates_async(self):
        """Called on startup. Skips the check if disabled or if <24h since
        the last attempt. Runs the actual network call on a daemon thread
        so we never block the UI."""
        if IS_TEST_BUILD:
            logging.info("Auto-update check skipped: test build.")
            return
        cfg = self.config_mgr.config.get("updates", {})
        if not cfg.get("check_on_startup", True):
            return
        last = cfg.get("last_check_iso") or ""
        if last:
            try:
                last_dt = datetime.datetime.fromisoformat(last)
                elapsed = (datetime.datetime.now() - last_dt).total_seconds()
                if elapsed < UPDATE_CHECK_INTERVAL_SEC:
                    return
            except ValueError:
                pass
        self._check_for_updates_async(silent=True)

    def _check_for_updates_async(self, silent: bool = False):
        """Kick off the release lookup on a background thread.
        silent=True: no popup if we're up-to-date or offline.
        silent=False (manual "Check now"): always inform the user."""
        if IS_TEST_BUILD:
            logging.info("Auto-update check skipped: test build.")
            if not silent:
                messagebox.showinfo(
                    "Check for Updates",
                    "Auto-update is disabled in test builds.")
            return
        t = threading.Thread(
            target=self._do_update_check, args=(silent,), daemon=True,
        )
        t.start()

    def _do_update_check(self, silent: bool):
        release = fetch_latest_release()
        if not release:
            # Don't stamp last_check_iso on failure — otherwise a one-time
            # network blip (or a previously private repo returning 404) locks
            # auto-update out for 24h. The check only runs once per launch,
            # so retry-on-next-launch is fine.
            if not silent:
                self.after(0, lambda: messagebox.showwarning(
                    "Check for Updates",
                    "Couldn't reach the update server.\n"
                    "Check your internet connection and try again."))
            return

        self.config_mgr.config["updates"]["last_check_iso"] = \
            datetime.datetime.now().isoformat(timespec="seconds")
        try:
            self.config_mgr.save()
        except Exception:
            pass

        tag = release.get("tag_name") or ""
        latest = _parse_version(tag)
        current = _parse_version(APP_VERSION)
        if latest <= current:
            if not silent:
                self.after(0, lambda: messagebox.showinfo(
                    "Check for Updates",
                    f"You're up to date.\n\nInstalled version: {APP_VERSION}"))
            return

        asset = _pick_release_asset(release)
        if not asset:
            if not silent:
                self.after(0, lambda: messagebox.showinfo(
                    "Update Available",
                    f"Version {tag} is available, but no installer was found "
                    "for this platform. Visit the Releases page to download "
                    "it manually."))
            return

        # Respect a "skip this version" choice from a previous prompt.
        skip = self.config_mgr.config["updates"].get("skip_version", "")
        if silent and skip == tag:
            return

        notes = (release.get("body") or "").strip()
        self.after(0, lambda: self._prompt_update(tag, asset, notes))

    def _prompt_update(self, tag: str, asset: dict, notes: str):
        """Ask the user whether to install the newer version now."""
        short_notes = notes if len(notes) < 600 else notes[:600].rstrip() + "…"
        message = (
            f"A new version of Speedy Scandocs is available.\n\n"
            f"Installed: {APP_VERSION}\n"
            f"Available: {tag.lstrip('vV')}\n\n"
        )
        if short_notes:
            message += f"What's new:\n{short_notes}\n\n"
        message += "Install now?\n\n(The app will close and the installer will open.)"

        # yesnocancel: Yes = install, No = remind later, Cancel = skip this version
        resp = messagebox.askyesnocancel("Update Available", message)
        if resp is None:
            self.config_mgr.config["updates"]["skip_version"] = tag
            self.config_mgr.save()
            return
        if resp is False:
            return
        self._download_and_install(asset)

    def _download_and_install(self, asset: dict):
        """Show a progress dialog, stream the installer to a temp file,
        then launch it and exit the app."""
        import tempfile
        url = asset.get("browser_download_url")
        name = asset.get("name") or "SpeedyScandocsInstaller"
        if not url:
            return
        dest = os.path.join(tempfile.gettempdir(), name)

        # Progress dialog
        dlg = tk.Toplevel(self)
        dlg.title("Downloading Update")
        dlg.geometry("420x140")
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()
        ttk.Label(dlg, text=f"Downloading {name}…", padding=(15, 12, 15, 0)).pack(anchor="w")
        pbar = ttk.Progressbar(dlg, mode="determinate", maximum=100, length=390)
        pbar.pack(padx=15, pady=10)
        status_var = tk.StringVar(value="Starting…")
        ttk.Label(dlg, textvariable=status_var, padding=(15, 0)).pack(anchor="w")

        cancelled = {"v": False}
        def _cancel():
            cancelled["v"] = True
            dlg.destroy()
        cancel_btn = ttk.Button(dlg, text="Cancel", command=_cancel,
                                bootstyle="secondary-outline")
        cancel_btn.pack(pady=6)

        def _progress(done, total):
            pct = int(done * 100 / total) if total else 0
            mb_done = done / (1024 * 1024)
            mb_total = total / (1024 * 1024) if total else 0
            self.after(0, lambda: (
                pbar.configure(value=pct),
                status_var.set(
                    f"{mb_done:.1f} MB / {mb_total:.1f} MB ({pct}%)"
                    if total else f"{mb_done:.1f} MB"
                ),
            ))

        def _worker():
            ok = download_file(url, dest, progress_cb=_progress,
                               cancel_flag=lambda: cancelled["v"])
            self.after(0, lambda: self._on_download_done(dlg, ok, dest, cancelled["v"]))

        threading.Thread(target=_worker, daemon=True).start()

    def _on_download_done(self, dlg, ok: bool, path: str, was_cancelled: bool):
        try:
            dlg.destroy()
        except Exception:
            pass
        if was_cancelled:
            return
        if not ok:
            messagebox.showerror(
                "Update Failed",
                "The update could not be downloaded. Please try again later "
                "or download the installer manually from the Releases page.",
            )
            return
        try:
            if sys.platform == "win32":
                # Inno Setup installer handles UAC + overwrite-in-place itself.
                os.startfile(path)
            elif sys.platform == "darwin":
                # Opens the DMG in Finder; user drags the new .app to
                # Applications. Not seamless, but the standard Mac pattern.
                subprocess.Popen(["open", path])
            else:
                _open_file(path)
        except Exception as e:
            messagebox.showerror(
                "Update Failed",
                f"Could not launch the installer:\n{e}\n\n"
                f"Installer location:\n{path}",
            )
            return
        # Give the installer a moment to spawn before we exit.
        self.after(600, self._on_close)


# ─────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────

def _missing_dep_error(package: str):
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror(
        "Missing Dependency",
        f"{package} is not installed.\n\n"
        "Please run:\n    pip install -r requirements.txt\n\n"
        "Then restart Speedy Scandocs.",
    )
    sys.exit(1)


if __name__ == "__main__":
    logging.basicConfig(
        filename=LOG_PATH,
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    if fitz is None:
        _missing_dep_error("PyMuPDF")
    if requests is None:
        _missing_dep_error("requests")

    _load_bundled_fonts()

    app = ScandocsApp()
    app.mainloop()
